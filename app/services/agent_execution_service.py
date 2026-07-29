from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from fastapi import HTTPException

from app.erp.base import ERPProviderError
from app.erp.providers import get_active_provider
from app.erp.resources import provider_fields_for, provider_resource_for
from app.permissions import ensure_erp_resource_allowed, is_valid_position
from app.services.automation_flow_version_service import resolve_flow_execution_reference
from app.services.finance_salary_service import FinanceSalaryExportResult, recognize_salary_export_intent
from app.services.finance_salary_wechat_service import (
    FINANCE_WECHAT_STATUS_LABELS,
    build_wechat_prepare_confirmation_task,
    extract_wechat_recipient,
    recognize_salary_wechat_send_intent,
    run_record_status_for_salary_wechat,
)
from app.services.enterprise_wechat_service import search_enterprise_wechat_recipients
from app.services.generated_file_service import save_generated_file
from app.services.logging_service import write_audit_log
from app.services.run_record_service import elapsed_ms, finish_run, now_ms, record_artifact, record_step, start_run
from app.services.user_ai_app_permission_service import is_ai_app_allowed
from app.skills.executor import execute_skill


PLAN_EXECUTE_WORKFLOW_ID = "finance_monthly_package_wechat_send"
PLAN_EXECUTE_APP_ID = "agent-plan-execute-finance-package"
EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True)
class AgentTaskRoute:
    mode: str
    workflow_id: str | None
    intent: str
    reason: str
    confidence: float
    estimated_step_count: int
    requires_plan_execute: bool


@dataclass(frozen=True)
class AgentPlanStep:
    key: str
    label: str
    tool_type: str
    description: str
    sensitive: bool = False


@dataclass(frozen=True)
class FinanceReportWorkbookResult:
    filename: str
    content: bytes
    metadata: dict[str, Any]
    resource_summaries: list[dict[str, Any]]


ProgressCallback = Callable[[dict[str, Any]], None]


def classify_agent_task(message: str, current_user: dict) -> AgentTaskRoute:
    text = " ".join((message or "").strip().split())
    if _looks_like_finance_monthly_package_wechat(text, current_user):
        return AgentTaskRoute(
            mode="plan_execute",
            workflow_id=PLAN_EXECUTE_WORKFLOW_ID,
            intent="finance_monthly_package_wechat_send",
            reason="需要整理财务报表和工资表，并准备通过微信发送，属于超过 5 步的跨能力任务。",
            confidence=0.96,
            estimated_step_count=len(build_finance_monthly_package_plan(message)),
            requires_plan_execute=True,
        )

    return AgentTaskRoute(
        mode="react",
        workflow_id=None,
        intent="simple_react_task",
        reason="单个 ReAct/Skill 快路径可以处理，优先低延迟执行。",
        confidence=0.72,
        estimated_step_count=1,
        requires_plan_execute=False,
    )


def build_finance_monthly_package_plan(message: str) -> list[dict[str, Any]]:
    intent = recognize_salary_wechat_send_intent(message)
    recipient_name = intent.recipient_name or extract_wechat_recipient(message) or "待识别联系人"
    period_label = intent.salary_intent.period_label
    steps = [
        AgentPlanStep(
            key="intent",
            label="识别复杂任务",
            tool_type="rules",
            description=f"识别期间为 {period_label}，微信联系人为 {recipient_name}。",
        ),
        AgentPlanStep(
            key="permission",
            label="检查权限",
            tool_type="backend_policy",
            description="检查财务岗位、AI 应用启用状态和 ERP 资源权限。",
            sensitive=True,
        ),
        AgentPlanStep(
            key="salary_export",
            label="生成工资表",
            tool_type="skill.finance_salary_export",
            description="通过财务工资表 Skill 查询 ERPNext Salary Slip 并生成 Excel。",
            sensitive=True,
        ),
        AgentPlanStep(
            key="finance_report",
            label="整理财务报表",
            tool_type="python.erp_report",
            description="通过内部 Python 服务查询财务 ERP 摘要并生成报表 Excel。",
            sensitive=True,
        ),
        AgentPlanStep(
            key="merge_workbook",
            label="合并汇总表",
            tool_type="python.openpyxl",
            description="把工资表和财务报表整理为一个汇总 Excel 包。",
            sensitive=True,
        ),
        AgentPlanStep(
            key="save_files",
            label="保存文件",
            tool_type="file_center",
            description="保存工资表、财务报表和合并汇总表到文档下载中心。",
            sensitive=True,
        ),
        AgentPlanStep(
            key="wechat_prepare",
            label="企业微信发送确认",
            tool_type="enterprise_wechat_api",
            description="文件生成后等待确认；确认后由后端企业微信应用发送文件，不附带正文说明。",
            sensitive=True,
        ),
        AgentPlanStep(
            key="final_result",
            label="返回结果",
            tool_type="assistant",
            description="只向普通用户返回最终结果、文件和待人工发送状态。",
        ),
    ]
    return [
        {
            "key": step.key,
            "label": step.label,
            "tool_type": step.tool_type,
            "description": step.description,
            "sensitive": step.sensitive,
        }
        for step in steps
    ]


def execute_finance_monthly_package_wechat(
    *,
    message: str,
    current_user: dict,
    thread_id: str | None = None,
    parent_run_id: str | None = None,
    source: str = "chat_stream",
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    route = classify_agent_task(message, current_user)
    if route.workflow_id != PLAN_EXECUTE_WORKFLOW_ID:
        raise ValueError("当前消息不属于财务月度报表工资表微信发送复杂任务。")

    plan = build_finance_monthly_package_plan(message)
    salary_intent = recognize_salary_export_intent(message)
    wechat_intent = recognize_salary_wechat_send_intent(message)
    recipient_name = wechat_intent.recipient_name or extract_wechat_recipient(message)
    if not recipient_name:
        raise ValueError("请说明要发送给哪个微信联系人，例如：发给张三。")

    _emit(progress_callback, plan, "intent", "running")
    run_id = start_run(
        run_type="agent_plan_execute",
        app_id=PLAN_EXECUTE_APP_ID,
        app_name="财务月度资料微信发送",
        entrypoint="/chat/stream" if source == "chat_stream" else "/chat",
        current_user=current_user,
        thread_id=thread_id,
        resource_type="automation",
        resource_id=PLAN_EXECUTE_WORKFLOW_ID,
        flow_reference=_resolve_optional_flow_reference(
            flow_key=f"automation:finance:{PLAN_EXECUTE_WORKFLOW_ID}",
            current_user=current_user,
            execution_source=source,
        ),
        input_text=message,
        metadata={
            "mode": route.mode,
            "workflow_id": route.workflow_id,
            "parent_run_id": parent_run_id,
            "plan": plan,
            "recipient_name": recipient_name,
            "period_label": salary_intent.period_label,
        },
    )
    started_ms = now_ms()
    generated_artifacts: list[dict[str, Any]] = []

    try:
        _emit(progress_callback, plan, "permission", "running")
        permission_started_ms = now_ms()
        _ensure_finance_plan_execute_allowed(current_user)
        record_step(
            run_id=run_id,
            step_name="plan.permission_check",
            step_order=1,
            status_value="succeeded",
            provider="backend_policy",
            resource_type="automation",
            resource_id=PLAN_EXECUTE_WORKFLOW_ID,
            input_text={"position": current_user.get("position"), "role": current_user.get("role")},
            output_text="财务复杂任务权限检查通过",
            duration_ms=elapsed_ms(permission_started_ms),
            metadata={
                "required_apps": ["automation-salary_summary", "automation-salary_wechat_send", "automation-report_analysis"],
                "erp_resources": ["Salary Slip", "GL Entry", "Payment Entry", "Sales Invoice", "Purchase Invoice"],
            },
        )
        _emit(progress_callback, plan, "permission", "succeeded")

        _emit(progress_callback, plan, "salary_export", "running")
        salary_started_ms = now_ms()
        salary_skill_result = execute_skill(
            skill_id="finance_salary_export",
            payload={
                "message": message,
                "intent": salary_intent,
                "requested_erp_resources": ["Salary Slip"],
                "metadata": {
                    "run_id": run_id,
                    "entrypoint": "agent_plan_execute",
                    "parent_run_id": parent_run_id,
                },
            },
            current_user=current_user,
            source="agent_plan_execute",
        )
        salary_attachment = _skill_attachment(salary_skill_result)
        salary_result = _salary_result_from_skill(
            attachment=salary_attachment,
            metadata=salary_skill_result.metadata,
            intent=salary_intent,
        )
        record_step(
            run_id=run_id,
            step_name="plan.salary_export",
            step_order=2,
            status_value="succeeded",
            provider=str((salary_skill_result.metadata.get("provider") or "skill.finance_salary_export")),
            resource_type="erp",
            resource_id="Salary Slip",
            input_text=message,
            output_text=salary_attachment["filename"],
            duration_ms=elapsed_ms(salary_started_ms),
            metadata=salary_skill_result.metadata,
        )
        _emit(progress_callback, plan, "salary_export", "succeeded", {"filename": salary_attachment["filename"]})

        _emit(progress_callback, plan, "finance_report", "running")
        report_started_ms = now_ms()
        report_result = build_monthly_finance_report_workbook(
            message=message,
            current_user=current_user,
            period_label=salary_intent.period_label,
            start_date=salary_intent.start_date,
            end_date=salary_intent.end_date,
        )
        record_step(
            run_id=run_id,
            step_name="plan.finance_report",
            step_order=3,
            status_value="succeeded",
            provider="python.erp_report",
            resource_type="erp",
            resource_id="GL Entry,Payment Entry,Sales Invoice,Purchase Invoice",
            input_text=message,
            output_text=report_result.filename,
            duration_ms=elapsed_ms(report_started_ms),
            metadata=report_result.metadata,
        )
        _emit(progress_callback, plan, "finance_report", "succeeded", {"filename": report_result.filename})

        _emit(progress_callback, plan, "merge_workbook", "running")
        merge_started_ms = now_ms()
        package_filename, package_content, package_metadata = build_finance_package_workbook(
            message=message,
            salary_result=salary_result,
            report_result=report_result,
            recipient_name=recipient_name,
        )
        record_step(
            run_id=run_id,
            step_name="plan.merge_workbook",
            step_order=4,
            status_value="succeeded",
            provider="python.openpyxl",
            resource_type="generated_file",
            resource_id=package_filename,
            input_text={"salary": salary_result.filename, "report": report_result.filename},
            output_text=package_filename,
            duration_ms=elapsed_ms(merge_started_ms),
            metadata=package_metadata,
        )
        _emit(progress_callback, plan, "merge_workbook", "succeeded", {"filename": package_filename})

        _emit(progress_callback, plan, "save_files", "running")
        salary_artifact_id = save_generated_file(
            run_id=run_id,
            content=salary_attachment["content"],
            artifact_type="excel_file",
            mime_type=salary_attachment.get("mime_type") or EXCEL_MIME_TYPE,
            filename=salary_attachment["filename"],
            current_user=current_user,
            metadata=salary_skill_result.metadata,
        )
        report_artifact_id = save_generated_file(
            run_id=run_id,
            content=report_result.content,
            artifact_type="excel_file",
            mime_type=EXCEL_MIME_TYPE,
            filename=report_result.filename,
            current_user=current_user,
            metadata=report_result.metadata,
        )
        package_artifact_id = save_generated_file(
            run_id=run_id,
            content=package_content,
            artifact_type="excel_file",
            mime_type=EXCEL_MIME_TYPE,
            filename=package_filename,
            current_user=current_user,
            metadata=package_metadata,
        )
        generated_artifacts = [
            _artifact_item("工资表", salary_artifact_id, salary_attachment["filename"]),
            _artifact_item("财务报表", report_artifact_id, report_result.filename),
            _artifact_item("合并汇总表", package_artifact_id, package_filename),
        ]
        record_step(
            run_id=run_id,
            step_name="plan.save_files",
            step_order=5,
            status_value="succeeded",
            provider="generated_file_service",
            resource_type="generated_file",
            resource_id=package_artifact_id,
            input_text=[item["filename"] for item in generated_artifacts],
            output_text=f"已保存 {len(generated_artifacts)} 个文件",
            duration_ms=0,
            metadata={"artifacts": generated_artifacts},
        )
        _emit(progress_callback, plan, "save_files", "succeeded", {"artifacts": generated_artifacts})

        _emit(progress_callback, plan, "wechat_prepare", "running")
        wechat_started_ms = now_ms()
        dispatch = _build_package_dispatch(
            message=message,
            recipient_name=recipient_name,
            salary_result=salary_result,
            package_filename=package_filename,
            generated_artifacts=generated_artifacts,
            plan=plan,
            source=source,
            current_user=current_user,
        )
        wechat_execution = build_wechat_prepare_confirmation_task(
            dispatch=dispatch,
            artifact_id=package_artifact_id,
            artifact_filename=package_filename,
            current_user=current_user,
        )
        record_step(
            run_id=run_id,
            step_name="plan.wechat_prepare",
            step_order=6,
            status_value=run_record_status_for_salary_wechat(wechat_execution["status"]),
            provider=str(wechat_execution.get("executor_type") or "confirmation_required"),
            resource_type="external_automation",
            resource_id="enterprise_wechat",
            input_text={"recipient_name": recipient_name, "package_artifact_id": package_artifact_id},
            output_text=wechat_execution.get("message"),
            duration_ms=elapsed_ms(wechat_started_ms),
            metadata=wechat_execution,
        )
        _record_mcp_traces(run_id=run_id, traces=wechat_execution.get("mcp_tool_calls"), start_order=7)
        _emit(progress_callback, plan, "wechat_prepare", "blocked", {"status": wechat_execution["status"]})

        answer = _finance_package_answer(
            period_label=salary_intent.period_label,
            recipient_name=recipient_name,
            generated_artifacts=generated_artifacts,
            wechat_execution=wechat_execution,
            report_result=report_result,
            salary_result=salary_result,
        )
        metadata = {
            "mode": "plan_execute",
            "workflow_id": PLAN_EXECUTE_WORKFLOW_ID,
            "plan": plan,
            "generated_artifacts": generated_artifacts,
            "wechat_send": wechat_execution,
            "business_status": wechat_execution["status"],
            "business_status_label": wechat_execution.get("status_label"),
            "manual_final_send_required": False,
            "parent_run_id": parent_run_id,
        }
        finish_run(
            run_id,
            status_value=run_record_status_for_salary_wechat(wechat_execution["status"]),
            output_text=answer,
            duration_ms=elapsed_ms(started_ms),
            metadata=metadata,
        )
        write_audit_log(
            user_id=current_user.get("id"),
            action="agent.plan_execute.finance_monthly_package_wechat_send",
            resource_type="automation",
            resource_id=PLAN_EXECUTE_WORKFLOW_ID,
            metadata={
                "run_id": run_id,
                "thread_id": thread_id,
                "username": current_user.get("username"),
                "position": current_user.get("position"),
                "recipient_name": recipient_name,
                "generated_artifact_count": len(generated_artifacts),
                "business_status": wechat_execution["status"],
            },
        )
        _emit(progress_callback, plan, "final_result", "succeeded")
        return {
            "thread_id": thread_id,
            "answer": answer,
            "intent": "finance_monthly_package_wechat_send",
            "risk_level": "high",
            "attachments": [
                {
                    "type": "excel_file",
                    "filename": package_filename,
                    "mime_type": EXCEL_MIME_TYPE,
                    "size_bytes": len(package_content),
                    "content_base64": None,
                    "metadata": {
                        **package_metadata,
                        "artifact_id": package_artifact_id,
                        "download_path": f"/files/{package_artifact_id}/download" if package_artifact_id else None,
                    },
                }
            ],
            "approval_result": {
                "status": wechat_execution["status"],
                "status_label": wechat_execution.get("status_label") or FINANCE_WECHAT_STATUS_LABELS.get(wechat_execution["status"]),
                "manual_final_send_required": False,
            },
            "automation": {
                "type": "agent_plan_execute",
                "workflow_id": PLAN_EXECUTE_WORKFLOW_ID,
                "run_id": run_id,
                "status": wechat_execution["status"],
                "status_label": wechat_execution.get("status_label"),
                "plan": plan,
                "generated_artifacts": generated_artifacts,
                "wechat_send": wechat_execution,
                "recipient_name": recipient_name,
                "artifact_id": package_artifact_id,
                "filename": package_filename,
                "download_path": f"/files/{package_artifact_id}/download" if package_artifact_id else None,
                "source_message": message,
            },
        }
    except Exception as error:
        record_step(
            run_id=run_id,
            step_name="plan.failed",
            step_order=99,
            status_value="failed",
            provider="agent_plan_execute",
            resource_type="automation",
            resource_id=PLAN_EXECUTE_WORKFLOW_ID,
            input_text=message,
            error_message=error,
            duration_ms=elapsed_ms(started_ms),
            metadata={"generated_artifacts": generated_artifacts},
        )
        finish_run(
            run_id,
            status_value="failed",
            error_message=error,
            duration_ms=elapsed_ms(started_ms),
            metadata={
                "mode": "plan_execute",
                "workflow_id": PLAN_EXECUTE_WORKFLOW_ID,
                "parent_run_id": parent_run_id,
                "generated_artifacts": generated_artifacts,
            },
        )
        _emit(progress_callback, plan, "final_result", "failed", {"message": str(error)})
        raise


def build_monthly_finance_report_workbook(
    *,
    message: str,
    current_user: dict,
    period_label: str,
    start_date: date,
    end_date: date,
    resources: tuple[str, ...] | list[str] | None = None,
    source: str = "agent_plan_execute",
) -> FinanceReportWorkbookResult:
    provider = get_active_provider()
    resource_order = tuple(resources or ("GL Entry", "Payment Entry", "Sales Invoice", "Purchase Invoice"))
    max_workers = max(1, min(4, len(resource_order)))
    resource_summary_map: dict[str, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(
                _query_finance_resource_summary,
                resource=resource,
                provider_id=provider.provider_id,
                provider_label=provider.provider_label,
                start_date=start_date,
                end_date=end_date,
                current_user=current_user,
            ): resource
            for resource in resource_order
        }
        for future in as_completed(future_map):
            resource_summary_map[future_map[future]] = future.result()
    resource_summaries = [resource_summary_map[resource] for resource in resource_order]
    insights = _finance_report_insights(resource_summaries)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "财务报表摘要"
    summary.append(["期间", period_label])
    summary.append(["用户要求", message])
    summary.append(["ERP Provider", provider.provider_label])
    summary.append([])
    summary.append(["ERP 表", "状态", "记录数", "金额摘要", "说明"])
    for item in resource_summaries:
        summary.append([
            item["label"],
            item["status_label"],
            item["count"],
            item["amount_summary"],
            item["message"],
        ])
    _style_sheet(summary)

    insight_sheet = workbook.create_sheet("经营分析")
    insight_sheet.append(["指标", "数值", "说明"])
    for item in insights:
        insight_sheet.append([item["label"], item["value"], item["message"]])
    _style_sheet(insight_sheet)

    for item in resource_summaries:
        sheet = workbook.create_sheet(item["label"][:28])
        fields = item.get("fields") or []
        sheet.append(fields or ["结果"])
        for row in item.get("items") or []:
            sheet.append([row.get(field, "") for field in fields])
        _style_sheet(sheet)

    output = BytesIO()
    workbook.save(output)
    filename = f"finance_monthly_report_{start_date.strftime('%Y%m')}.xlsx"
    content = output.getvalue()
    return FinanceReportWorkbookResult(
        filename=filename,
        content=content,
        metadata={
            "output_filename": filename,
            "output_format": "xlsx",
            "period_label": period_label,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "source": source,
            "resource_summaries": [
                {key: value for key, value in item.items() if key != "items"}
                for item in resource_summaries
            ],
            "insights": insights,
            "output_bytes": len(content),
        },
        resource_summaries=resource_summaries,
    )


def build_finance_package_workbook(
    *,
    message: str,
    salary_result: FinanceSalaryExportResult,
    report_result: FinanceReportWorkbookResult,
    recipient_name: str,
) -> tuple[str, bytes, dict[str, Any]]:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "发送汇总"
    summary.append(["字段", "内容"])
    summary.append(["用户要求", message])
    summary.append(["发送对象", recipient_name])
    summary.append(["期间", salary_result.intent.period_label])
    summary.append(["工资表文件", salary_result.filename])
    summary.append(["财务报表文件", report_result.filename])
    summary.append(["员工数", salary_result.metadata.get("employee_count")])
    summary.append(["应发合计", salary_result.metadata.get("gross_pay_total")])
    summary.append(["实发合计", salary_result.metadata.get("net_pay_total")])
    summary.append(["安全要求", "微信最终发送必须由人工确认，系统不会自动点击发送。"])
    _style_sheet(summary)

    salary_sheet = workbook.create_sheet("工资摘要")
    salary_sheet.append(["工资单号", "员工编号", "员工姓名", "开始日期", "结束日期", "应发工资", "实发工资", "状态"])
    for item in salary_result.items:
        salary_sheet.append([
            item.get("name") or "",
            item.get("employee") or "",
            item.get("employee_name") or "",
            item.get("start_date") or "",
            item.get("end_date") or "",
            item.get("gross_pay") or 0,
            item.get("net_pay") or 0,
            item.get("status") or "",
        ])
    _style_sheet(salary_sheet)

    report_workbook = load_workbook(BytesIO(report_result.content), data_only=True)
    for source_sheet in report_workbook.worksheets[:5]:
        target = workbook.create_sheet(f"报表-{source_sheet.title}"[:31])
        for row in source_sheet.iter_rows(values_only=True):
            target.append(list(row))
        _style_sheet(target)

    output = BytesIO()
    workbook.save(output)
    filename = f"finance_monthly_package_{salary_result.intent.start_date.strftime('%Y%m')}.xlsx"
    content = output.getvalue()
    metadata = {
        "output_filename": filename,
        "output_format": "xlsx",
        "period_label": salary_result.intent.period_label,
        "salary_filename": salary_result.filename,
        "report_filename": report_result.filename,
        "recipient_name": recipient_name,
        "manual_final_send_required": True,
        "output_bytes": len(content),
    }
    return filename, content, metadata


def _looks_like_finance_monthly_package_wechat(text: str, current_user: dict) -> bool:
    lowered = text.lower()
    if current_user.get("role") != "admin" and current_user.get("position") != "finance":
        return False
    has_salary = any(keyword in lowered for keyword in ["工资", "薪资", "工资表", "salary", "payroll"])
    has_report = any(keyword in lowered for keyword in ["财务报表", "报表", "月报", "利润", "总账", "发票", "收付款"])
    has_send = any(keyword in lowered for keyword in ["发送", "发给", "传给", "转发", "send"])
    has_wechat = any(keyword in lowered for keyword in ["微信", "个人微信", "wechat", "weixin"])
    has_merge_or_multi = any(keyword in lowered for keyword in ["合并", "整理", "一起", "两个表", "多个", "汇总"])
    return has_salary and has_report and has_send and has_wechat and has_merge_or_multi


def _resolve_optional_flow_reference(
    *,
    flow_key: str,
    current_user: dict,
    execution_source: str,
) -> dict[str, Any] | None:
    try:
        return resolve_flow_execution_reference(
            flow_key=flow_key,
            current_user=current_user,
            execution_source=execution_source,
        )
    except HTTPException as error:
        if error.status_code == 404:
            return None
        raise


def _ensure_finance_plan_execute_allowed(current_user: dict) -> None:
    if current_user.get("role") != "admin":
        if current_user.get("position") != "finance" or not is_valid_position(current_user.get("position")):
            raise ValueError("只有财务岗位或管理员可以执行财务复杂自动化。")
        for app_id, label in [
            ("automation-salary_summary", "统计工资"),
            ("automation-salary_wechat_send", "工资表微信发送"),
            ("automation-report_analysis", "分析财务报表"),
        ]:
            if not _is_app_allowed(current_user, app_id):
                raise ValueError(f"{label}应用已被管理员禁用。")

    for resource in ("Salary Slip", "GL Entry", "Payment Entry", "Sales Invoice", "Purchase Invoice"):
        ensure_erp_resource_allowed(current_user, resource)


def _query_finance_resource_summary(
    *,
    resource: str,
    provider_id: str,
    provider_label: str,
    start_date: date,
    end_date: date,
    current_user: dict,
) -> dict[str, Any]:
    ensure_erp_resource_allowed(current_user, resource)
    provider = get_active_provider()
    provider_resource = provider_resource_for(resource, provider_id)
    fields = provider_fields_for(resource, provider_id)
    label = _resource_label(resource)
    if provider_resource is None:
        return {
            "resource": resource,
            "label": label,
            "status": "not_supported",
            "status_label": "未映射",
            "count": 0,
            "amount_summary": "-",
            "message": f"{provider_label} 暂未映射 {resource}。",
            "fields": fields,
            "items": [],
        }

    try:
        result = provider.query_resource(
            resource=resource,
            provider_resource=provider_resource,
            query=None,
            filters=[["posting_date", ">=", start_date.isoformat()], ["posting_date", "<=", end_date.isoformat()]],
            fields=fields,
            limit=80,
        )
    except ERPProviderError as error:
        return {
            "resource": resource,
            "label": label,
            "status": "failed",
            "status_label": "查询失败",
            "count": 0,
            "amount_summary": "-",
            "message": error.message,
            "fields": fields,
            "items": [],
        }

    items = result.get("items") if isinstance(result.get("items"), list) else []
    return {
        "resource": resource,
        "label": label,
        "status": "succeeded" if result.get("ok", True) else "failed",
        "status_label": "已读取" if result.get("ok", True) else "查询失败",
        "count": len(items),
        "amount_summary": _amount_summary(resource, items),
        "message": str(result.get("message") or "已读取 ERP 摘要。"),
        "fields": fields,
        "items": items,
    }


def _amount_summary(resource: str, items: list[dict[str, Any]]) -> str:
    if resource == "GL Entry":
        debit = _sum_number(items, "debit")
        credit = _sum_number(items, "credit")
        return f"借方 {debit:.2f} / 贷方 {credit:.2f}"
    if resource == "Payment Entry":
        return f"收付款合计 {_sum_number(items, 'paid_amount'):.2f}"
    if resource in {"Sales Invoice", "Purchase Invoice"}:
        total = _sum_number(items, "grand_total")
        outstanding = _sum_number(items, "outstanding_amount")
        return f"总额 {total:.2f} / 未结 {outstanding:.2f}"
    return "-"


def _finance_report_insights(resource_summaries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary_by_resource = {str(item.get("resource")): item for item in resource_summaries}
    gl_items = summary_by_resource.get("GL Entry", {}).get("items") or []
    payment_items = summary_by_resource.get("Payment Entry", {}).get("items") or []
    sales_items = summary_by_resource.get("Sales Invoice", {}).get("items") or []
    purchase_items = summary_by_resource.get("Purchase Invoice", {}).get("items") or []

    sales_total = _sum_number(sales_items, "grand_total")
    purchase_total = _sum_number(purchase_items, "grand_total")
    payment_total = _sum_number(payment_items, "paid_amount")
    sales_outstanding = _sum_number(sales_items, "outstanding_amount")
    purchase_outstanding = _sum_number(purchase_items, "outstanding_amount")
    gl_debit = _sum_number(gl_items, "debit")
    gl_credit = _sum_number(gl_items, "credit")
    operating_balance = sales_total - purchase_total
    cash_gap = payment_total - purchase_total

    return [
        {
            "label": "销售发票总额",
            "value": round(sales_total, 2),
            "message": f"本期读取 {len(sales_items)} 条销售发票，用于观察收入规模。",
        },
        {
            "label": "采购发票总额",
            "value": round(purchase_total, 2),
            "message": f"本期读取 {len(purchase_items)} 条采购发票，用于观察采购和成本压力。",
        },
        {
            "label": "经营差额",
            "value": round(operating_balance, 2),
            "message": "按销售发票总额减采购发票总额估算，不等同于正式利润表。",
        },
        {
            "label": "收付款合计",
            "value": round(payment_total, 2),
            "message": f"本期读取 {len(payment_items)} 条收付款单，用于观察现金流入流出。",
        },
        {
            "label": "现金覆盖差额",
            "value": round(cash_gap, 2),
            "message": "按收付款合计减采购发票总额估算，用于发现现金覆盖压力。",
        },
        {
            "label": "销售未结金额",
            "value": round(sales_outstanding, 2),
            "message": "销售发票 outstanding_amount 汇总，数值越高代表回款跟进压力越大。",
        },
        {
            "label": "采购未结金额",
            "value": round(purchase_outstanding, 2),
            "message": "采购发票 outstanding_amount 汇总，用于观察待付款压力。",
        },
        {
            "label": "总账借贷",
            "value": f"借方 {gl_debit:.2f} / 贷方 {gl_credit:.2f}",
            "message": f"本期读取 {len(gl_items)} 条总账分录，用于核对账务活动规模。",
        },
    ]


def _is_app_allowed(current_user: dict, app_id: str) -> bool:
    allowed_ids = current_user.get("allowed_ai_app_ids")
    if isinstance(allowed_ids, list):
        return app_id in {str(item) for item in allowed_ids}
    return is_ai_app_allowed(current_user, app_id)


def _sum_number(items: list[dict[str, Any]], key: str) -> float:
    total = 0.0
    for item in items:
        try:
            total += float(item.get(key) or 0)
        except (TypeError, ValueError):
            continue
    return total


def _skill_attachment(result) -> dict[str, Any]:
    if not result.attachments:
        raise ValueError("Skill 未返回文件产物。")
    attachment = result.attachments[0]
    if not isinstance(attachment.get("content"), bytes):
        raise ValueError("Skill 文件产物格式错误。")
    return attachment


def _salary_result_from_skill(
    *,
    attachment: dict[str, Any],
    metadata: dict[str, Any],
    intent,
) -> FinanceSalaryExportResult:
    salary_metadata = metadata if metadata.get("resource") == "Salary Slip" else metadata.get("salary") or metadata
    items = _salary_items_from_workbook(attachment["content"])
    return FinanceSalaryExportResult(
        filename=str(attachment["filename"]),
        content=attachment["content"],
        metadata=salary_metadata,
        intent=intent,
        items=items,
        provider=str(salary_metadata.get("provider") or "erp"),
        provider_label=str(salary_metadata.get("provider_label") or "ERP"),
        provider_resource=str(salary_metadata.get("provider_resource") or "Salary Slip"),
    )


def _salary_items_from_workbook(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        return []
    if "工资明细" not in workbook.sheetnames:
        return []
    sheet = workbook["工资明细"]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    key_map = {
        "工资单号": "name",
        "员工编号": "employee",
        "员工姓名": "employee_name",
        "开始日期": "start_date",
        "结束日期": "end_date",
        "应发工资": "gross_pay",
        "实发工资": "net_pay",
        "状态": "status",
    }
    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        item: dict[str, Any] = {}
        for index, value in enumerate(row):
            header = headers[index] if index < len(headers) else ""
            key = key_map.get(header)
            if key:
                item[key] = value
        if any(value not in (None, "") for value in item.values()):
            items.append(item)
    return items


def _build_package_dispatch(
    *,
    message: str,
    recipient_name: str,
    salary_result: FinanceSalaryExportResult,
    package_filename: str,
    generated_artifacts: list[dict[str, Any]],
    plan: list[dict[str, Any]],
    source: str,
    current_user: dict,
) -> dict[str, Any]:
    recipient_search = search_enterprise_wechat_recipients(
        recipient_name,
        current_user=current_user,
    )
    selected_recipient = recipient_search.get("selected_item") if isinstance(recipient_search.get("selected_item"), dict) else None
    requires_recipient_selection = bool(recipient_search.get("needs_selection"))
    return {
        "status": "waiting_recipient_selection" if requires_recipient_selection else "generated",
        "status_label": FINANCE_WECHAT_STATUS_LABELS["waiting_recipient_selection" if requires_recipient_selection else "generated"],
        "message": "财务报表和工资表已整理完成，正在准备企业微信发送确认。",
        "executor_type": "enterprise_wechat_api",
        "executor_mode": "confirm_before_enterprise_wechat_send",
        "configured": True,
        "platform": "enterprise_wechat",
        "target_app": "enterprise_wechat",
        "channel": "enterprise_wechat",
        "recipient_name": recipient_name,
        "recipient": selected_recipient,
        "recipient_search": recipient_search,
        "requires_recipient_selection": requires_recipient_selection,
        "manual_final_send_required": False,
        "requires_recipient_confirmation": True,
        "requires_sensitive_confirmation": True,
        "message_body": "",
        "payload": {
            "action": "send_enterprise_wechat_file_after_confirmation",
            "platform": "enterprise_wechat",
            "target_app": "enterprise_wechat",
            "channel": "enterprise_wechat",
            "recipient_name": recipient_name,
            "recipient": selected_recipient,
            "recipient_search": recipient_search,
            "filename": package_filename,
            "period_label": salary_result.intent.period_label,
            "manual_final_send_required": False,
            "message_body": "",
            "source": source,
            "requested_by": current_user.get("id"),
            "source_message": message,
            "attachment_filenames": [item["filename"] for item in generated_artifacts],
        },
        "plan": {
            "title": "财务月度资料微信发送",
            "summary": f"整理 {salary_result.intent.period_label} 财务报表和工资表，并准备通过企业微信发送给 {recipient_name}。",
            "period_label": salary_result.intent.period_label,
            "recipient_name": recipient_name,
            "requires_recipient_confirmation": True,
            "requires_sensitive_confirmation": True,
            "manual_final_send_required": False,
            "steps": plan,
            "warnings": [
                "工资和财务报表属于敏感数据。",
                "确认后由后端发送文件，不附带正文说明。",
            ],
        },
        "logs": [
            {"level": "info", "message": "Plan-and-Execute 已完成内部 Python 服务和 Skill 执行。"},
            {"level": "warning", "message": "企业微信发送前必须确认接收对象和敏感数据。"},
        ],
        "screenshots": [],
    }


def _finance_package_answer(
    *,
    period_label: str,
    recipient_name: str,
    generated_artifacts: list[dict[str, Any]],
    wechat_execution: dict[str, Any],
    report_result: FinanceReportWorkbookResult,
    salary_result: FinanceSalaryExportResult,
) -> str:
    names = "\n".join(f"- {item['label']}：{item['filename']}" for item in generated_artifacts)
    report_counts = "；".join(
        f"{item['label']} {item['count']} 条"
        for item in report_result.resource_summaries
    )
    return (
        f"已完成 {period_label} 财务资料整理，并准备通过企业微信发送给“{recipient_name}”。\n\n"
        f"生成文件：\n{names}\n\n"
        f"工资表：{salary_result.metadata.get('employee_count') or len(salary_result.items)} 名员工，"
        f"实发合计 {salary_result.metadata.get('net_pay_total', 0)}。\n"
        f"财务报表读取：{report_counts or '暂无可统计数据'}。\n"
        f"当前状态：{wechat_execution.get('status_label') or wechat_execution.get('status')}。\n"
        "工资和财务数据属于敏感内容，发送前需要你在聊天窗口确认接收对象和敏感数据。"
    )


def _artifact_item(label: str, artifact_id: str | None, filename: str) -> dict[str, Any]:
    return {
        "label": label,
        "artifact_id": artifact_id,
        "filename": filename,
        "download_path": f"/files/{artifact_id}/download" if artifact_id else None,
    }


def _record_mcp_traces(*, run_id: str, traces: Any, start_order: int) -> None:
    if not isinstance(traces, list):
        return
    for index, trace in enumerate(traces, start=start_order):
        if not isinstance(trace, dict):
            continue
        status = str(trace.get("status") or "").lower()
        record_step(
            run_id=run_id,
            step_name=f"mcp_tool.{trace.get('tool_id') or 'unknown'}",
            step_order=index,
            status_value="failed" if status in {"failed", "error"} else "blocked" if status.startswith("waiting") else "succeeded",
            provider="mcp",
            resource_type="mcp_tool",
            resource_id=str(trace.get("tool_id") or "mcp_tool"),
            input_text={"argument_keys": trace.get("argument_keys") or [], "source": trace.get("source")},
            output_text=trace.get("message") or trace.get("status"),
            duration_ms=trace.get("duration_ms") if isinstance(trace.get("duration_ms"), int) else None,
            metadata=trace,
        )


def _emit(
    callback: ProgressCallback | None,
    plan: list[dict[str, Any]],
    step_key: str,
    status: str,
    data: dict[str, Any] | None = None,
) -> None:
    if callback is None:
        return
    step = next((item for item in plan if item["key"] == step_key), None)
    callback({
        "workflow_id": PLAN_EXECUTE_WORKFLOW_ID,
        "step_key": step_key,
        "label": step["label"] if step else step_key,
        "status": status,
        "detail": step["description"] if step else None,
        "data": data or {},
    })


def _resource_label(resource: str) -> str:
    return {
        "GL Entry": "总账分录",
        "Payment Entry": "收付款单",
        "Sales Invoice": "销售发票",
        "Purchase Invoice": "采购发票",
    }.get(resource, resource)


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_font = Font(bold=True, color="17324D")
    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)
    sheet.freeze_panes = "A2"
