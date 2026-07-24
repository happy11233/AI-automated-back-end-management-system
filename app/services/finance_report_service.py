from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.llm import chat
from app.rag.loaders import EmptyDocumentError, UnsupportedDocumentTypeError, load_documents_from_bytes


MAX_REPORT_FILE_BYTES = 10 * 1024 * 1024
MAX_REPORT_FILES = 6
MAX_REPORT_TEXT_CHARS = 18000
SUPPORTED_REPORT_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx", ".xls", ".csv"}
WORD_MIME_TYPE = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass
class FinanceReportInputFile:
    filename: str
    content: bytes


@dataclass
class FinanceReportAnalysisResult:
    filename: str
    content: bytes
    mime_type: str
    output_format: str
    answer: str
    metadata: dict[str, Any]


def analyze_finance_report_files(
    *,
    files: list[FinanceReportInputFile],
    instruction: str,
    output_format: str,
) -> FinanceReportAnalysisResult:
    normalized_output = _normalize_output_format(output_format)
    if not files:
        raise ValueError("请至少上传 1 个财务报表文件。")
    if len(files) > MAX_REPORT_FILES:
        raise ValueError(f"一次最多上传 {MAX_REPORT_FILES} 个财务报表文件。")

    parsed_files = [_parse_report_file(item) for item in files]
    merged_context = _build_report_context(parsed_files)
    prompt = _build_analysis_prompt(
        instruction=instruction,
        merged_context=merged_context,
        output_format=normalized_output,
    )
    answer = chat(prompt)
    safe_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if normalized_output == "word":
        filename = f"finance_report_analysis_{safe_timestamp}.docx"
        content = _build_docx_report(answer=answer, files=parsed_files, instruction=instruction)
        mime_type = WORD_MIME_TYPE
    else:
        filename = f"finance_report_analysis_{safe_timestamp}.xlsx"
        content = _build_xlsx_report(answer=answer, files=parsed_files, instruction=instruction)
        mime_type = EXCEL_MIME_TYPE

    return FinanceReportAnalysisResult(
        filename=filename,
        content=content,
        mime_type=mime_type,
        output_format=normalized_output,
        answer=answer,
        metadata={
            "output_filename": filename,
            "output_format": normalized_output,
            "source_file_count": len(files),
            "source_filenames": [item.filename for item in files],
            "source_bytes": sum(len(item.content) for item in files),
            "parsed_document_count": sum(item["document_count"] for item in parsed_files),
            "parsed_text_chars": sum(len(item["text"]) for item in parsed_files),
            "analysis_chars": len(answer),
        },
    )


def _parse_report_file(item: FinanceReportInputFile) -> dict[str, Any]:
    filename = item.filename or "finance_report"
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_REPORT_EXTENSIONS:
        supported = "、".join(sorted(SUPPORTED_REPORT_EXTENSIONS))
        raise ValueError(f"{filename} 不支持，当前支持 {supported}。")
    if not item.content:
        raise ValueError(f"{filename} 是空文件。")
    if len(item.content) > MAX_REPORT_FILE_BYTES:
        raise ValueError(f"{filename} 超过 10MB。")

    try:
        documents = load_documents_from_bytes(filename, item.content)
    except (UnsupportedDocumentTypeError, EmptyDocumentError) as error:
        raise ValueError(str(error)) from error

    text_parts = []
    for document in documents:
        content = " ".join(document.page_content.split())
        if content:
            text_parts.append(content)
    text = "\n".join(text_parts).strip()
    if not text:
        raise ValueError(f"{filename} 解析后没有可分析内容。")

    return {
        "filename": filename,
        "suffix": suffix,
        "document_count": len(documents),
        "text": text[:MAX_REPORT_TEXT_CHARS],
    }


def _build_report_context(parsed_files: list[dict[str, Any]]) -> str:
    blocks = []
    remaining = MAX_REPORT_TEXT_CHARS
    for index, item in enumerate(parsed_files, start=1):
        head = f"文件 {index}：{item['filename']}，解析片段："
        available = max(500, remaining - len(head))
        snippet = item["text"][:available]
        blocks.append(f"{head}\n{snippet}")
        remaining -= len(snippet)
        if remaining <= 1000:
            break
    return "\n\n".join(blocks)


def _build_analysis_prompt(*, instruction: str, merged_context: str, output_format: str) -> str:
    user_instruction = instruction.strip() or "请分析上传的财务报表，输出摘要、关键指标、异常项、风险和建议。"
    return f"""你是跨境电商企业内部的财务报表分析 AI。
你只能基于上传文件中能解析到的内容和财务常识做分析，不要编造不存在的数据。

财务要求：
{user_instruction}

上传文件解析内容：
{merged_context}

输出文件格式：{output_format}

请用中文输出：
1. 报表摘要
2. 关键指标
3. 异常项和可能原因
4. 财务风险
5. 下一步复核建议
"""


def _build_xlsx_report(*, answer: str, files: list[dict[str, Any]], instruction: str) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "AI分析报告"
    summary.append(["模块", "内容"])
    for title, content in _split_report_sections(answer):
        summary.append([title, content])
    _style_two_column_sheet(summary)

    source_sheet = workbook.create_sheet("源文件摘要")
    source_sheet.append(["文件名", "类型", "解析文档数", "解析字符数"])
    for item in files:
        source_sheet.append([item["filename"], item["suffix"].lstrip("."), item["document_count"], len(item["text"])])
    _style_table(source_sheet)

    request_sheet = workbook.create_sheet("分析要求")
    request_sheet.append(["字段", "内容"])
    request_sheet.append(["用户要求", instruction.strip() or "默认财务报表分析"])
    request_sheet.append(["生成时间", datetime.now().isoformat(timespec="seconds")])
    _style_two_column_sheet(request_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_docx_report(*, answer: str, files: list[dict[str, Any]], instruction: str) -> bytes:
    paragraphs = [
        ("title", "财务报表 AI 分析报告"),
        ("normal", f"生成时间：{datetime.now().isoformat(timespec='seconds')}"),
        ("normal", f"分析要求：{instruction.strip() or '默认财务报表分析'}"),
        ("heading", "AI 分析结果"),
    ]
    for title, content in _split_report_sections(answer):
        paragraphs.append(("heading", title))
        paragraphs.append(("normal", content))
    paragraphs.append(("heading", "源文件"))
    for item in files:
        paragraphs.append(("normal", f"{item['filename']} / {item['suffix'].lstrip('.')} / 解析字符 {len(item['text'])}"))

    document_xml = _docx_document_xml(paragraphs)
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as docx:
        docx.writestr("[Content_Types].xml", _content_types_xml())
        docx.writestr("_rels/.rels", _root_rels_xml())
        docx.writestr("word/_rels/document.xml.rels", _document_rels_xml())
        docx.writestr("word/document.xml", document_xml)
        docx.writestr("word/styles.xml", _styles_xml())
    return output.getvalue()


def _split_report_sections(answer: str) -> list[tuple[str, str]]:
    text = answer.strip() or "AI 未返回分析内容。"
    pattern = re.compile(r"(?m)^\s*(?:\d+[\.、]|#+)\s*([^：:\n]+)[：:]?\s*$")
    matches = list(pattern.finditer(text))
    if not matches:
        return [("AI 分析", text)]

    sections: list[tuple[str, str]] = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        title = match.group(1).strip()
        content = text[start:end].strip()
        if title or content:
            sections.append((title or f"分析项 {index + 1}", content or "-"))
    return sections or [("AI 分析", text)]


def _style_two_column_sheet(sheet) -> None:
    _style_table(sheet)
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 90
    for cell in sheet["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _style_table(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_font = Font(bold=True, color="17324D")
    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)
    sheet.freeze_panes = "A2"


def _normalize_output_format(value: str) -> str:
    normalized = (value or "word").strip().lower()
    if normalized in {"docx", "word"}:
        return "word"
    if normalized in {"xlsx", "excel"}:
        return "excel"
    raise ValueError("输出格式只能选择 Word 或 Excel。")


def _xml_escape(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _docx_document_xml(paragraphs: list[tuple[str, str]]) -> str:
    body = []
    for style, text in paragraphs:
        lines = str(text).splitlines() or [""]
        for line in lines:
            p_style = ""
            if style == "title":
                p_style = '<w:pPr><w:pStyle w:val="Title"/></w:pPr>'
            elif style == "heading":
                p_style = '<w:pPr><w:pStyle w:val="Heading1"/></w:pPr>'
            body.append(
                "<w:p>"
                f"{p_style}"
                "<w:r><w:t xml:space=\"preserve\">"
                f"{_xml_escape(line)}"
                "</w:t></w:r>"
                "</w:p>"
            )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        + "".join(body)
        + '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/><w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440"/></w:sectPr>'
        + "</w:body></w:document>"
    )


def _content_types_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
  <Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
</Types>"""


def _root_rels_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""


def _document_rels_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>"""


def _styles_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:style w:type="paragraph" w:default="1" w:styleId="Normal">
    <w:name w:val="Normal"/>
    <w:rPr><w:sz w:val="22"/><w:szCs w:val="22"/></w:rPr>
  </w:style>
  <w:style w:type="paragraph" w:styleId="Title">
    <w:name w:val="Title"/>
    <w:rPr><w:b/><w:sz w:val="36"/><w:szCs w:val="36"/></w:rPr>
  </w:style>
  <w:style w:type="paragraph" w:styleId="Heading1">
    <w:name w:val="heading 1"/>
    <w:rPr><w:b/><w:sz w:val="28"/><w:szCs w:val="28"/></w:rPr>
  </w:style>
</w:styles>"""
