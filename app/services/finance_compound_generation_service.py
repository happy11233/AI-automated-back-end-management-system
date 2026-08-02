from __future__ import annotations

import base64
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.permissions import POSITION_LABELS, erp_scopes_for_position, is_valid_position
from app.services.agent_execution_service import (
    EXCEL_MIME_TYPE,
    FinanceReportWorkbookResult,
    build_monthly_finance_report_workbook,
)
from app.services.email_service import (
    EmailAttachment,
    email_result_metadata,
    resolve_email_recipient,
    send_email_with_attachments,
)
from app.services.finance_compound_intent_service import (
    FINANCE_COMPOUND_INTENT,
    FINANCE_REPORT_OUTPUT,
    FINANCE_REPORT_RESOURCES,
    FINANCE_SALARY_OUTPUT,
    OUTPUT_LABELS,
    FinanceCompoundIntent,
    recognize_finance_compound_intent,
)
from app.services.finance_salary_service import (
    FinanceSalaryExportResult,
    SalaryExportIntent,
    export_salary_workbook_from_erp,
)
from app.services.generated_file_service import save_generated_file
from app.services.run_record_service import elapsed_ms, now_ms, record_step
from app.services.user_ai_app_permission_service import is_ai_app_allowed


@dataclass
class FinanceCompoundAttachment:
    label: str
    filename: str
    content: bytes
    mime_type: str
    metadata: dict[str, Any]

    def to_chat_attachment(self) -> dict[str, Any]:
        return {
            "type": "excel_file",
            "filename": self.filename,
            "mime_type": self.mime_type,
            "size_bytes": len(self.content),
            "content_base64": base64.b64encode(self.content).decode("ascii"),
            "metadata": self.metadata,
        }

    def to_email_attachment(self) -> EmailAttachment:
        return EmailAttachment(filename=self.filename, content=self.content, mime_type=self.mime_type)


@dataclass
class FinanceCompoundGenerationResult:
    answer: str
    intent: FinanceCompoundIntent
    attachments: list[FinanceCompoundAttachment]
    metadata: dict[str, Any]
    salary_result: FinanceSalaryExportResult | None = None
    report_result: FinanceReportWorkbookResult | None = None


class FinanceCompoundClarification(ValueError):
    """Raised when the requested finance output lacks required data."""


def execute_finance_compound_generation(
    *,
    message: str,
    current_user: dict,
    intent: FinanceCompoundIntent | None = None,
    run_id: str | None = None,
    source: str = "chat",
) -> FinanceCompoundGenerationResult:
    intent = intent or recognize_finance_compound_intent(message)
    if intent.intent != FINANCE_COMPOUND_INTENT:
        raise ValueError("没有识别到财务报表生成请求，请说明要生成财务报表、工资表或两者都要。")

    started_ms = now_ms()
    _record_step(
        run_id,
        step_name="finance_compound_intent_recognition",
        step_order=1,
        status_value="succeeded",
        provider="rules",
        resource_type="automation",
        resource_id=FINANCE_COMPOUND_INTENT,
        input_text=message,
        output_text=", ".join(intent.output_labels),
        duration_ms=elapsed_ms(started_ms),
        metadata={
            "intent": intent.intent,
            "outputs": list(intent.outputs),
            "period_label": intent.period_label,
            "start_date": intent.start_date.isoformat(),
            "end_date": intent.end_date.isoformat(),
            "merge_requested": intent.merge_requested,
            "email_requested": intent.email_requested,
            "wechat_requested": intent.wechat_requested,
            "matched_keywords": list(intent.matched_keywords),
            "source": source,
        },
    )

    permission_started_ms = now_ms()
    _ensure_finance_compound_allowed(current_user=current_user, intent=intent)
    _record_step(
        run_id,
        step_name="finance_compound_permission_check",
        step_order=2,
        status_value="succeeded",
        provider="skill_registry",
        resource_type="automation",
        resource_id=FINANCE_COMPOUND_INTENT,
        input_text={
            "position": current_user.get("position"),
            "role": current_user.get("role"),
            "requested_outputs": list(intent.outputs),
        },
        output_text="财务复合生成权限检查通过",
        duration_ms=elapsed_ms(permission_started_ms),
        metadata={
            "required_apps": _required_apps(intent),
            "erp_resources": list(intent.requested_erp_resources),
        },
    )

    report_result: FinanceReportWorkbookResult | None = None
    salary_result: FinanceSalaryExportResult | None = None
    generated: list[FinanceCompoundAttachment] = []
    step_order = 3

    if FINANCE_REPORT_OUTPUT in intent.outputs:
        report_started_ms = now_ms()
        try:
            report_result = build_monthly_finance_report_workbook(
                message=message,
                current_user=current_user,
                period_label=intent.period_label,
                start_date=intent.start_date,
                end_date=intent.end_date,
                resources=FINANCE_REPORT_RESOURCES,
                source=source,
            )
        except HTTPException:
            raise
        except Exception as error:
            _record_step(
                run_id,
                step_name="finance_monthly_report_generation",
                step_order=step_order,
                status_value="failed",
                provider="python.erp_report",
                resource_type="erp",
                resource_id=",".join(FINANCE_REPORT_RESOURCES),
                input_text=message,
                error_message=error,
                duration_ms=elapsed_ms(report_started_ms),
            )
            raise

        if _report_has_no_data(report_result):
            raise FinanceCompoundClarification("缺少【财务报表所需 ERP 数据】数据，是否继续生成？")

        generated.append(
            FinanceCompoundAttachment(
                label=OUTPUT_LABELS[FINANCE_REPORT_OUTPUT],
                filename=report_result.filename,
                content=report_result.content,
                mime_type=EXCEL_MIME_TYPE,
                metadata=report_result.metadata,
            )
        )
        _record_step(
            run_id,
            step_name="finance_monthly_report_generation",
            step_order=step_order,
            status_value="succeeded",
            provider="python.erp_report",
            resource_type="erp",
            resource_id=",".join(FINANCE_REPORT_RESOURCES),
            input_text=message,
            output_text=report_result.filename,
            duration_ms=elapsed_ms(report_started_ms),
            metadata=report_result.metadata,
        )
        step_order += 1

    if FINANCE_SALARY_OUTPUT in intent.outputs:
        salary_started_ms = now_ms()
        try:
            salary_result = export_salary_workbook_from_erp(
                message=message,
                current_user=current_user,
                intent=_salary_intent_from_compound(intent),
            )
        except ValueError as error:
            if "没有查到" in str(error) or "工资单" in str(error):
                raise FinanceCompoundClarification("缺少【员工工资表】数据，是否继续生成？") from error
            _record_step(
                run_id,
                step_name="finance_salary_export",
                step_order=step_order,
                status_value="failed",
                provider="erp_provider",
                resource_type="erp",
                resource_id="Salary Slip",
                input_text=message,
                error_message=error,
                duration_ms=elapsed_ms(salary_started_ms),
            )
            raise

        generated.append(
            FinanceCompoundAttachment(
                label=OUTPUT_LABELS[FINANCE_SALARY_OUTPUT],
                filename=salary_result.filename,
                content=salary_result.content,
                mime_type=EXCEL_MIME_TYPE,
                metadata=salary_result.metadata,
            )
        )
        _record_step(
            run_id,
            step_name="finance_salary_export",
            step_order=step_order,
            status_value="succeeded",
            provider=salary_result.provider,
            resource_type="erp",
            resource_id="Salary Slip",
            input_text=message,
            output_text=salary_result.filename,
            duration_ms=elapsed_ms(salary_started_ms),
            metadata=salary_result.metadata,
        )
        step_order += 1

    attachments = generated
    if (intent.merge_requested or intent.wechat_requested) and len(generated) > 1:
        merge_started_ms = now_ms()
        merged = _build_merged_workbook(
            message=message,
            intent=intent,
            generated=generated,
        )
        attachments = [merged]
        _record_step(
            run_id,
            step_name="finance_compound_merge_workbook",
            step_order=step_order,
            status_value="succeeded",
            provider="python.openpyxl",
            resource_type="generated_file",
            resource_id=merged.filename,
            input_text=[item.filename for item in generated],
            output_text=merged.filename,
            duration_ms=elapsed_ms(merge_started_ms),
            metadata=merged.metadata,
        )
        step_order += 1

    save_started_ms = now_ms()
    _save_attachments(
        run_id=run_id,
        attachments=attachments,
        current_user=current_user,
    )
    _record_step(
        run_id,
        step_name="finance_compound_save_files",
        step_order=step_order,
        status_value="succeeded",
        provider="generated_file_service",
        resource_type="generated_file",
        resource_id=",".join(item.filename for item in attachments),
        input_text=[item.filename for item in attachments],
        output_text=f"已保存 {len(attachments)} 个文件",
        duration_ms=elapsed_ms(save_started_ms),
        metadata={"attachments": [_attachment_metadata(item) for item in attachments]},
    )

    email_metadata = {"email_requested": intent.email_requested, "email_sent": False}
    if intent.email_requested:
        email_recipient, email_recipient_source = resolve_email_recipient(
            message,
            current_user.get("email"),
        )
        email_result = send_email_with_attachments(
            to_email=email_recipient,
            subject=f"{intent.period_label}财务资料",
            body=_email_body(current_user=current_user, intent=intent, attachments=attachments),
            attachments=[item.to_email_attachment() for item in attachments],
        )
        email_metadata = {
            **email_result_metadata(email_result),
            "email_recipient_source": email_recipient_source,
        }
        _record_step(
            run_id,
            step_name="finance_compound_email_delivery",
            step_order=step_order + 1,
            status_value="succeeded" if email_result.sent else "failed",
            provider=email_result.provider,
            resource_type="email",
            resource_id=email_result.recipient,
            input_text=message,
            output_text="已发送财务资料邮件" if email_result.sent else email_result.error,
            error_message=email_result.error if not email_result.sent else None,
            duration_ms=0,
            metadata=email_metadata,
        )

    metadata = {
        "intent": intent.intent,
        "outputs": list(intent.outputs),
        "output_labels": list(intent.output_labels),
        "period_label": intent.period_label,
        "start_date": intent.start_date.isoformat(),
        "end_date": intent.end_date.isoformat(),
        "merge_requested": intent.merge_requested,
        "generated_count": len(attachments),
        "generated_files": [_attachment_metadata(item) for item in attachments],
        "source": source,
        **email_metadata,
    }

    return FinanceCompoundGenerationResult(
        answer=_build_answer(
            intent=intent,
            attachments=attachments,
            salary_result=salary_result,
            report_result=report_result,
            email_metadata=email_metadata,
        ),
        intent=intent,
        attachments=attachments,
        metadata=metadata,
        salary_result=salary_result,
        report_result=report_result,
    )


def _ensure_finance_compound_allowed(*, current_user: dict, intent: FinanceCompoundIntent) -> None:
    if current_user.get("role") != "admin":
        position = current_user.get("position")
        if position != "finance" or not is_valid_position(position):
            current_label = POSITION_LABELS.get(str(position), "当前")
            raise ValueError(f"{current_label}岗位没有权限生成财务资料。")

        for app_id, label in _required_app_pairs(intent):
            if not _is_app_allowed(current_user, app_id):
                raise ValueError(f"{label}应用已被管理员禁用。")

    allowed_resources = set(erp_scopes_for_position(current_user.get("position")))
    for resource in intent.requested_erp_resources:
        if current_user.get("role") != "admin" and resource not in allowed_resources:
            raise ValueError(f"你没有调用【{resource}】的权限，无法生成对应表。")


def _required_apps(intent: FinanceCompoundIntent) -> list[str]:
    return [app_id for app_id, _label in _required_app_pairs(intent)]


def _required_app_pairs(intent: FinanceCompoundIntent) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    if FINANCE_REPORT_OUTPUT in intent.outputs:
        pairs.append(("automation-report_analysis", "分析财务报表"))
    if FINANCE_SALARY_OUTPUT in intent.outputs:
        pairs.append(("automation-salary_summary", "统计工资"))
    return pairs


def _is_app_allowed(current_user: dict, app_id: str) -> bool:
    allowed_ids = current_user.get("allowed_ai_app_ids")
    if isinstance(allowed_ids, list):
        return app_id in {str(item) for item in allowed_ids}
    return is_ai_app_allowed(current_user, app_id)


def _salary_intent_from_compound(intent: FinanceCompoundIntent) -> SalaryExportIntent:
    return SalaryExportIntent(
        intent=FINANCE_SALARY_OUTPUT,
        period_label=intent.period_label,
        start_date=intent.start_date,
        end_date=intent.end_date,
        output_format="xlsx",
        confidence=intent.confidence,
        matched_keywords=list(intent.matched_keywords),
    )


def _report_has_no_data(result: FinanceReportWorkbookResult) -> bool:
    summaries = result.resource_summaries or []
    if not summaries:
        return True
    return all(int(item.get("count") or 0) <= 0 for item in summaries)


def _save_attachments(
    *,
    run_id: str | None,
    attachments: list[FinanceCompoundAttachment],
    current_user: dict,
) -> None:
    for attachment in attachments:
        artifact_id = save_generated_file(
            run_id=run_id,
            content=attachment.content,
            artifact_type="excel_file",
            mime_type=attachment.mime_type,
            filename=attachment.filename,
            current_user=current_user,
            metadata=attachment.metadata,
        )
        if artifact_id:
            attachment.metadata = {
                **attachment.metadata,
                "artifact_id": artifact_id,
                "download_path": f"/files/{artifact_id}/download",
            }


def _build_answer(
    *,
    intent: FinanceCompoundIntent,
    attachments: list[FinanceCompoundAttachment],
    salary_result: FinanceSalaryExportResult | None,
    report_result: FinanceReportWorkbookResult | None,
    email_metadata: dict[str, Any],
) -> str:
    output_names = "、".join(intent.output_labels)
    if len(intent.outputs) > 1:
        lines = [f"已识别为月度财务资料生成请求，本次需要生成：{output_names}。"]
    else:
        lines = [f"已识别为{output_names}生成请求，期间为 {intent.period_label}。"]

    lines.append("")
    lines.append("已完成：")
    for index, attachment in enumerate(attachments, start=1):
        lines.append(f"{index}. {attachment.label}：{attachment.filename}")

    if salary_result is not None:
        lines.append(
            f"工资表共 {len(salary_result.items)} 名员工，应发合计 "
            f"{salary_result.metadata['gross_pay_total']:.2f}，实发合计 "
            f"{salary_result.metadata['net_pay_total']:.2f}。"
        )

    if report_result is not None:
        lines.append(_report_summary_sentence(report_result))

    if (intent.merge_requested or intent.wechat_requested) and len(attachments) == 1:
        lines.append("你要求合并输出，本次已把多个结果整理到一个 Excel 文件。")
    elif intent.wechat_requested:
        lines.append("你提到了微信发送，本次已整理成可确认发送的文件。")

    if intent.email_requested:
        if email_metadata.get("email_sent"):
            lines.append(f"已按你的要求发送到邮箱：{email_metadata.get('recipient')}")
        else:
            lines.append(f"你要求发送到邮箱，但邮件未发送成功：{email_metadata.get('error') or '未配置邮箱地址'}")
    elif intent.wechat_requested:
        lines.append("你提到了微信发送，本次会在聊天窗口展示企业微信发送确认卡，确认后由后端发送文件。")
    else:
        lines.append("你没有要求发送邮箱或微信，本次只在对话中输出并生成可下载附件。")

    return "\n".join(lines)


def _report_summary_sentence(report_result: FinanceReportWorkbookResult) -> str:
    summaries = report_result.resource_summaries or []
    read_labels = [str(item.get("label") or item.get("resource")) for item in summaries if int(item.get("count") or 0) > 0]
    missing_labels = [str(item.get("label") or item.get("resource")) for item in summaries if int(item.get("count") or 0) <= 0]
    if read_labels and missing_labels:
        return f"财务报表已汇总 {', '.join(read_labels)} 数据；其中 {', '.join(missing_labels)} 本期暂无记录。"
    if read_labels:
        return f"财务报表已汇总 {', '.join(read_labels)} 数据。"
    return "财务报表已生成，但本期 ERP 财务表暂无可汇总记录。"


def _email_body(
    *,
    current_user: dict,
    intent: FinanceCompoundIntent,
    attachments: list[FinanceCompoundAttachment],
) -> str:
    display_name = current_user.get("display_name") or current_user.get("username") or "同事"
    files = "\n".join(f"- {item.filename}" for item in attachments)
    return (
        f"你好，{display_name}：\n\n"
        f"系统已根据你的 AI 对话请求自动生成 {intent.period_label} 财务资料。\n\n"
        f"附件文件：\n{files}\n\n"
        "请在公司权限范围内下载、查看和流转。"
    )


def _build_merged_workbook(
    *,
    message: str,
    intent: FinanceCompoundIntent,
    generated: list[FinanceCompoundAttachment],
) -> FinanceCompoundAttachment:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "合并汇总"
    summary.append(["字段", "内容"])
    summary.append(["用户要求", message])
    summary.append(["期间", intent.period_label])
    summary.append(["生成结果", "、".join(item.label for item in generated)])
    summary.append(["源文件", "；".join(item.filename for item in generated)])

    for item in generated:
        source_workbook = load_workbook(BytesIO(item.content), data_only=True)
        for source_sheet in source_workbook.worksheets[:6]:
            target_name = f"{item.label}-{source_sheet.title}"[:31]
            target = workbook.create_sheet(target_name)
            for row in source_sheet.iter_rows(values_only=True):
                target.append(list(row))

    output = BytesIO()
    workbook.save(output)
    content = output.getvalue()
    safe_period = intent.start_date.strftime("%Y%m")
    filename = f"finance_compound_{safe_period}.xlsx"
    metadata = {
        "output_filename": filename,
        "output_format": "xlsx",
        "period_label": intent.period_label,
        "start_date": intent.start_date.isoformat(),
        "end_date": intent.end_date.isoformat(),
        "source_filenames": [item.filename for item in generated],
        "output_bytes": len(content),
    }
    return FinanceCompoundAttachment(
        label="合并财务资料",
        filename=filename,
        content=content,
        mime_type=EXCEL_MIME_TYPE,
        metadata=metadata,
    )


def _attachment_metadata(attachment: FinanceCompoundAttachment) -> dict[str, Any]:
    return {
        "label": attachment.label,
        "filename": attachment.filename,
        "mime_type": attachment.mime_type,
        "size_bytes": len(attachment.content),
        **attachment.metadata,
    }


def _record_step(run_id: str | None, **kwargs: Any) -> None:
    if not run_id:
        return
    record_step(run_id=run_id, **kwargs)
