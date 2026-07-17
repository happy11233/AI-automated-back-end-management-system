from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ALLOWED_RECONCILIATION_EXTENSIONS = {".xlsx", ".xls"}
MAX_RECONCILIATION_FILE_BYTES = 8 * 1024 * 1024
MAX_RECONCILIATION_TOTAL_BYTES = 32 * 1024 * 1024
MAX_RECONCILIATION_FILES = 8
MAX_OUTPUT_ROWS = 10000


@dataclass
class FinanceReconciliationInputFile:
    filename: str
    content: bytes


@dataclass
class FinanceReconciliationResult:
    filename: str
    content: bytes
    metadata: dict[str, Any]


def reconcile_finance_workbooks(
    *,
    files: list[FinanceReconciliationInputFile],
    instruction: str = "",
    base_currency: str = "CNY",
) -> FinanceReconciliationResult:
    if not files:
        raise ValueError("请至少上传 1 个对账 Excel 文件。")

    if len(files) > MAX_RECONCILIATION_FILES:
        raise ValueError(f"一次最多上传 {MAX_RECONCILIATION_FILES} 个 Excel 文件。")

    total_bytes = sum(len(item.content) for item in files)
    if total_bytes > MAX_RECONCILIATION_TOTAL_BYTES:
        raise ValueError("对账文件总大小不能超过 32MB。")

    normalized_currency = _clean_currency(base_currency) or "CNY"
    normalized_instruction = instruction.strip() or "按订单号和 SKU 自动匹配结算、物流、采购、广告和汇率，生成订单利润表和异常账单。"
    sheets = _load_source_sheets(files)
    classified = [_classify_sheet(item, normalized_currency) for item in sheets]
    exchange_rates = _collect_exchange_rates(classified, normalized_currency)
    source_overview = [_source_overview(item) for item in classified]
    field_mappings = [_field_mapping(item) for item in classified]

    settlement_rows = _normalize_settlement(classified, exchange_rates, normalized_currency)
    if settlement_rows.empty:
        raise ValueError("未识别到可用于对账的 Amazon 结算/订单明细。请确认文件包含订单号、SKU 和销售金额字段。")

    profit_rows, anomalies = _build_profit_rows(
        settlement_rows=settlement_rows,
        logistics_rows=_normalize_cost_rows(classified, "logistics", exchange_rates, normalized_currency),
        purchase_rows=_normalize_cost_rows(classified, "purchase", exchange_rates, normalized_currency),
        ads_rows=_normalize_cost_rows(classified, "ads", exchange_rates, normalized_currency),
    )
    summary = _build_summary(
        profit_rows=profit_rows,
        anomalies=anomalies,
        files=files,
        classified=classified,
        base_currency=normalized_currency,
        instruction=normalized_instruction,
    )
    workbook = _build_reconciliation_workbook(
        summary=summary,
        profit_rows=profit_rows,
        anomalies=anomalies,
        field_mappings=field_mappings,
        source_overview=source_overview,
    )

    output = BytesIO()
    workbook.save(output)
    content = output.getvalue()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"finance_reconciliation_{timestamp}.xlsx"

    return FinanceReconciliationResult(
        filename=output_filename,
        content=content,
        metadata={
            "output_filename": output_filename,
            "source_file_count": len(files),
            "source_bytes": total_bytes,
            "base_currency": normalized_currency,
            "order_line_count": int(len(profit_rows)),
            "anomaly_count": int(len(anomalies)),
            "total_sales": _round_money(profit_rows["sales_amount_base"].sum()),
            "total_profit": _round_money(profit_rows["profit_base"].sum()),
            "negative_profit_count": int((profit_rows["profit_base"] < 0).sum()),
            "sheet_type_counts": _sheet_type_counts(classified),
            "output_bytes": len(content),
            "instruction_preview": normalized_instruction[:500],
        },
    )


def _load_source_sheets(files: list[FinanceReconciliationInputFile]) -> list[dict[str, Any]]:
    loaded: list[dict[str, Any]] = []

    for item in files:
        suffix = Path(item.filename or "").suffix.lower()
        if suffix not in ALLOWED_RECONCILIATION_EXTENSIONS:
            raise ValueError(f"{item.filename} 只支持 .xlsx 或 .xls 文件。")

        if not item.content:
            raise ValueError(f"{item.filename} 是空文件。")

        if len(item.content) > MAX_RECONCILIATION_FILE_BYTES:
            raise ValueError(f"{item.filename} 超过 8MB。")

        try:
            engine = "xlrd" if suffix == ".xls" else "openpyxl"
            raw_sheets = pd.read_excel(
                BytesIO(item.content),
                sheet_name=None,
                dtype=object,
                engine=engine,
            )
        except Exception as error:
            raise ValueError(f"{item.filename} 读取失败：{error}") from error

        for sheet_name, frame in raw_sheets.items():
            cleaned = frame.dropna(how="all").dropna(axis=1, how="all")
            cleaned = cleaned.where(pd.notna(cleaned), "")
            if cleaned.empty:
                continue
            loaded.append(
                {
                    "filename": item.filename,
                    "sheet_name": str(sheet_name),
                    "frame": cleaned,
                }
            )

    if not loaded:
        raise ValueError("上传文件没有可读取的数据 Sheet。")

    return loaded


def _classify_sheet(item: dict[str, Any], base_currency: str) -> dict[str, Any]:
    frame: pd.DataFrame = item["frame"]
    context = f"{item['filename']} {item['sheet_name']}"
    fields = _detect_fields(frame)
    source_type = _detect_source_type(context, fields)

    return {
        **item,
        "source_type": source_type,
        "fields": fields,
        "base_currency": base_currency,
    }


def _detect_fields(frame: pd.DataFrame) -> dict[str, Any]:
    fields: dict[str, Any] = {}
    aliases = {
        "order_no": [
            "amazon order id", "order id", "order no", "order number", "platform order id",
            "平台订单号", "订单号", "订单编号", "亚马逊订单号",
        ],
        "sku": [
            "sku", "seller sku", "msku", "merchant sku", "商品sku", "产品sku", "商家sku", "sku编码",
        ],
        "quantity": [
            "quantity", "qty", "shipped quantity", "数量", "件数", "订单数量", "发货数量",
        ],
        "currency": [
            "currency", "currency code", "币种", "货币",
        ],
        "sales_amount": [
            "item price", "principal", "sales amount", "sales", "gross sales", "order amount",
            "销售额", "销售金额", "订单金额", "商品金额", "收入", "销售收入", "平台订单金额",
        ],
        "refund_amount": [
            "refund", "refund amount", "refunds", "退款", "退款金额", "退货退款",
        ],
        "platform_fee": [
            "commission", "amazon fee", "amazon fees", "selling fee", "selling fees", "fba fee",
            "fulfillment fee", "transaction fee", "other fee", "平台手续费", "手续费", "佣金",
            "平台费", "fba费用", "交易费",
        ],
        "logistics_fee": [
            "logistics fee", "shipping cost", "freight", "delivery cost", "carrier fee",
            "物流费", "物流费用", "运费", "配送费", "头程", "尾程", "仓储配送费",
        ],
        "purchase_cost": [
            "purchase cost", "unit cost", "product cost", "cost", "采购成本", "产品成本",
            "单位成本", "单件成本", "成本单价", "采购单价",
        ],
        "ad_fee": [
            "ad spend", "ads spend", "advertising cost", "advertising fee", "ppc spend",
            "广告费", "广告花费", "广告费用", "推广费", "站内广告",
        ],
        "exchange_rate": [
            "exchange rate", "rate", "rate to base", "rate to cny", "汇率", "兑人民币",
            "折算汇率", "换算汇率",
        ],
    }

    normalized_columns = {column: _normalize_label(str(column)) for column in frame.columns}
    for field, values in aliases.items():
        matches = [
            column for column, normalized in normalized_columns.items()
            if _matches_alias(normalized, values)
        ]
        if field in {"platform_fee"}:
            fields[field] = matches
        elif matches:
            fields[field] = matches[0]

    if "currency" not in fields:
        fields["currency"] = None

    return fields


def _matches_alias(normalized_column: str, aliases: list[str]) -> bool:
    for alias in aliases:
        normalized_alias = _normalize_label(alias)
        if not normalized_alias:
            continue
        if normalized_column == normalized_alias:
            return True
        if len(normalized_alias) >= 4 and normalized_alias in normalized_column:
            return True
    return False


def _detect_source_type(context: str, fields: dict[str, Any]) -> str:
    normalized_context = _normalize_label(context)

    if fields.get("exchange_rate") and fields.get("currency"):
        return "exchange_rate"

    context_type_hints = [
        ("logistics", ["logistics", "freight", "shipping", "delivery", "物流", "运费", "配送"]),
        ("ads", ["advertising", "adspend", "ads", "ppc", "广告", "推广"]),
        ("purchase", ["purchase", "supplier", "cost", "采购", "供应商", "成本"]),
        ("settlement", ["settlement", "transaction", "order", "amazon", "结算", "订单", "销售"]),
    ]
    for source_type, hints in context_type_hints:
        if any(_normalize_label(hint) in normalized_context for hint in hints):
            return source_type

    if fields.get("logistics_fee"):
        return "logistics"
    if fields.get("ad_fee"):
        return "ads"
    if fields.get("purchase_cost") and fields.get("sku") and not fields.get("sales_amount"):
        return "purchase"
    if fields.get("order_no") and fields.get("sku") and (
        fields.get("sales_amount") or fields.get("refund_amount") or fields.get("platform_fee")
    ):
        return "settlement"

    return "unknown"


def _collect_exchange_rates(classified: list[dict[str, Any]], base_currency: str) -> dict[str, float]:
    rates = {base_currency: 1.0}
    for item in classified:
        if item["source_type"] != "exchange_rate":
            continue

        frame: pd.DataFrame = item["frame"]
        fields = item["fields"]
        currency_column = fields.get("currency")
        rate_column = fields.get("exchange_rate")
        if not currency_column or not rate_column:
            continue

        for _, row in frame.iterrows():
            currency = _clean_currency(row.get(currency_column))
            rate = _to_number(row.get(rate_column))
            if currency and rate > 0:
                rates[currency] = rate

    return rates


def _normalize_settlement(
    classified: list[dict[str, Any]],
    exchange_rates: dict[str, float],
    base_currency: str,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    for item in classified:
        if item["source_type"] != "settlement":
            continue

        frame: pd.DataFrame = item["frame"]
        fields = item["fields"]
        order_column = fields.get("order_no")
        sku_column = fields.get("sku")
        sales_column = fields.get("sales_amount")
        if not order_column or not sku_column or not sales_column:
            continue

        for _, row in frame.iterrows():
            order_no = _clean_text(row.get(order_column))
            sku = _clean_text(row.get(sku_column))
            if not order_no and not sku:
                continue

            currency = _row_currency(row, fields.get("currency"), base_currency)
            rate = exchange_rates.get(currency)
            quantity = _to_number(row.get(fields.get("quantity"))) if fields.get("quantity") else 1.0
            quantity = quantity if quantity > 0 else 1.0
            sales = _money_to_base(row.get(sales_column), currency, exchange_rates)
            refund = abs(_money_to_base(row.get(fields.get("refund_amount")), currency, exchange_rates))
            platform_fee = _sum_columns_to_base(row, fields.get("platform_fee") or [], currency, exchange_rates, absolute=True)

            rows.append(
                {
                    "order_no": order_no,
                    "sku": sku,
                    "quantity": quantity,
                    "currency": currency,
                    "rate_missing": rate is None,
                    "sales_amount_base": sales,
                    "refund_amount_base": refund,
                    "platform_fee_base": platform_fee,
                    "source": f"{item['filename']}::{item['sheet_name']}",
                }
            )

    if not rows:
        return pd.DataFrame()

    frame = pd.DataFrame(rows)
    grouped = frame.groupby(["order_no", "sku"], dropna=False, as_index=False).agg(
        {
            "quantity": "sum",
            "currency": lambda values: ",".join(sorted({str(value) for value in values if value})),
            "rate_missing": "max",
            "sales_amount_base": "sum",
            "refund_amount_base": "sum",
            "platform_fee_base": "sum",
            "source": lambda values: "; ".join(sorted({str(value) for value in values})),
        }
    )
    grouped["quantity"] = grouped["quantity"].where(grouped["quantity"] > 0, 1.0)
    return grouped


def _normalize_cost_rows(
    classified: list[dict[str, Any]],
    source_type: str,
    exchange_rates: dict[str, float],
    base_currency: str,
) -> pd.DataFrame:
    amount_field = {
        "logistics": "logistics_fee",
        "purchase": "purchase_cost",
        "ads": "ad_fee",
    }[source_type]
    amount_name = {
        "logistics": "logistics_fee_base",
        "purchase": "purchase_cost_base",
        "ads": "ad_fee_base",
    }[source_type]
    rows: list[dict[str, Any]] = []

    for item in classified:
        if item["source_type"] != source_type:
            continue

        frame: pd.DataFrame = item["frame"]
        fields = item["fields"]
        amount_column = fields.get(amount_field)
        if not amount_column:
            continue

        order_column = fields.get("order_no")
        sku_column = fields.get("sku")
        quantity_column = fields.get("quantity")

        for _, row in frame.iterrows():
            order_no = _clean_text(row.get(order_column)) if order_column else ""
            sku = _clean_text(row.get(sku_column)) if sku_column else ""
            if not order_no and not sku:
                continue

            currency = _row_currency(row, fields.get("currency"), base_currency)
            amount = abs(_money_to_base(row.get(amount_column), currency, exchange_rates))
            quantity = _to_number(row.get(quantity_column)) if quantity_column else 0.0
            rows.append(
                {
                    "order_no": order_no,
                    "sku": sku,
                    "quantity": quantity,
                    amount_name: amount,
                    "source": f"{item['filename']}::{item['sheet_name']}",
                }
            )

    return pd.DataFrame(rows)


def _build_profit_rows(
    *,
    settlement_rows: pd.DataFrame,
    logistics_rows: pd.DataFrame,
    purchase_rows: pd.DataFrame,
    ads_rows: pd.DataFrame,
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    settlement = settlement_rows.copy()
    settlement["line_key"] = settlement["order_no"] + "||" + settlement["sku"]
    settlement["purchase_cost_base"] = _purchase_costs(settlement, purchase_rows)
    settlement["logistics_fee_base"] = _allocated_costs(settlement, logistics_rows, "logistics_fee_base")
    settlement["ad_fee_base"] = _allocated_costs(settlement, ads_rows, "ad_fee_base")
    settlement["net_revenue_base"] = (
        settlement["sales_amount_base"]
        - settlement["refund_amount_base"]
        - settlement["platform_fee_base"]
    )
    settlement["total_cost_base"] = (
        settlement["purchase_cost_base"]
        + settlement["logistics_fee_base"]
        + settlement["ad_fee_base"]
    )
    settlement["profit_base"] = settlement["net_revenue_base"] - settlement["total_cost_base"]
    settlement["profit_margin"] = settlement.apply(
        lambda row: row["profit_base"] / row["sales_amount_base"] if row["sales_amount_base"] else 0,
        axis=1,
    )
    settlement["status"] = settlement.apply(_profit_status, axis=1)

    anomalies = _build_anomalies(settlement, purchase_rows, logistics_rows, ads_rows)
    output_columns = [
        "order_no",
        "sku",
        "quantity",
        "currency",
        "sales_amount_base",
        "refund_amount_base",
        "platform_fee_base",
        "logistics_fee_base",
        "ad_fee_base",
        "purchase_cost_base",
        "net_revenue_base",
        "total_cost_base",
        "profit_base",
        "profit_margin",
        "status",
        "source",
    ]
    return settlement[output_columns], anomalies


def _purchase_costs(settlement: pd.DataFrame, purchase_rows: pd.DataFrame) -> pd.Series:
    if purchase_rows.empty:
        return pd.Series([0.0] * len(settlement), index=settlement.index)

    purchase = purchase_rows.copy()
    purchase["line_key"] = purchase["order_no"] + "||" + purchase["sku"]
    exact = purchase[(purchase["order_no"] != "") & (purchase["sku"] != "")]
    exact_map = exact.groupby("line_key")["purchase_cost_base"].sum().to_dict() if not exact.empty else {}

    sku_only = purchase[(purchase["sku"] != "") & (purchase["order_no"] == "")]
    sku_cost_map: dict[str, float] = {}
    if not sku_only.empty:
        sku_cost_map = sku_only.groupby("sku")["purchase_cost_base"].mean().to_dict()

    values = []
    for _, row in settlement.iterrows():
        exact_value = exact_map.get(row["line_key"])
        if exact_value is not None:
            values.append(float(exact_value))
            continue
        unit_cost = sku_cost_map.get(row["sku"])
        values.append(float(unit_cost or 0) * float(row["quantity"] or 1))

    return pd.Series(values, index=settlement.index)


def _allocated_costs(settlement: pd.DataFrame, cost_rows: pd.DataFrame, amount_column: str) -> pd.Series:
    if cost_rows.empty:
        return pd.Series([0.0] * len(settlement), index=settlement.index)

    allocated = pd.Series([0.0] * len(settlement), index=settlement.index)
    cost = cost_rows.copy()
    cost["line_key"] = cost["order_no"] + "||" + cost["sku"]

    exact = cost[(cost["order_no"] != "") & (cost["sku"] != "")]
    if not exact.empty:
        exact_map = exact.groupby("line_key")[amount_column].sum().to_dict()
        allocated += settlement["line_key"].map(exact_map).fillna(0.0)

    order_only = cost[(cost["order_no"] != "") & (cost["sku"] == "")]
    for order_no, amount in order_only.groupby("order_no")[amount_column].sum().items():
        indexes = settlement.index[settlement["order_no"] == order_no].tolist()
        allocated = _allocate_to_indexes(allocated, settlement, indexes, float(amount))

    sku_only = cost[(cost["sku"] != "") & (cost["order_no"] == "")]
    for sku, amount in sku_only.groupby("sku")[amount_column].sum().items():
        indexes = settlement.index[settlement["sku"] == sku].tolist()
        allocated = _allocate_to_indexes(allocated, settlement, indexes, float(amount))

    return allocated


def _allocate_to_indexes(
    allocated: pd.Series,
    settlement: pd.DataFrame,
    indexes: list[int],
    amount: float,
) -> pd.Series:
    if not indexes:
        return allocated

    basis = settlement.loc[indexes, "sales_amount_base"].abs()
    if float(basis.sum()) <= 0:
        basis = settlement.loc[indexes, "quantity"].abs()
    if float(basis.sum()) <= 0:
        share = 1 / len(indexes)
        for index in indexes:
            allocated.loc[index] += amount * share
        return allocated

    total = float(basis.sum())
    for index in indexes:
        allocated.loc[index] += amount * float(basis.loc[index]) / total
    return allocated


def _profit_status(row: pd.Series) -> str:
    if row["profit_base"] < 0:
        return "亏损"
    if row["profit_margin"] < 0.08:
        return "低毛利"
    return "正常"


def _build_anomalies(
    profit_rows: pd.DataFrame,
    purchase_rows: pd.DataFrame,
    logistics_rows: pd.DataFrame,
    ads_rows: pd.DataFrame,
) -> list[dict[str, Any]]:
    anomalies: list[dict[str, Any]] = []

    for _, row in profit_rows.iterrows():
        if bool(row.get("rate_missing")):
            anomalies.append(_anomaly("missing_exchange_rate", "high", row, "币种缺少汇率，已按 1 处理。", row["sales_amount_base"]))
        if row["purchase_cost_base"] <= 0:
            anomalies.append(_anomaly("missing_purchase_cost", "high", row, "未匹配到采购成本。", row["sales_amount_base"]))
        if row["logistics_fee_base"] <= 0:
            anomalies.append(_anomaly("missing_logistics_fee", "medium", row, "未匹配到物流费用。", row["sales_amount_base"]))
        if row["profit_base"] < 0:
            anomalies.append(_anomaly("negative_profit", "high", row, "订单利润为负。", row["profit_base"]))
        elif row["profit_margin"] < 0.08:
            anomalies.append(_anomaly("low_margin", "medium", row, "订单利润率低于 8%。", row["profit_base"]))
        if row["refund_amount_base"] > row["sales_amount_base"] and row["refund_amount_base"] > 0:
            anomalies.append(_anomaly("refund_exceeds_sales", "high", row, "退款金额超过销售额。", row["refund_amount_base"]))

    line_keys = set((profit_rows["order_no"] + "||" + profit_rows["sku"]).tolist())
    order_keys = set(profit_rows["order_no"].tolist())
    sku_keys = set(profit_rows["sku"].tolist())
    for source_name, rows, amount_column in [
        ("logistics", logistics_rows, "logistics_fee_base"),
        ("ads", ads_rows, "ad_fee_base"),
        ("purchase", purchase_rows, "purchase_cost_base"),
    ]:
        if rows.empty:
            continue
        temp = rows.copy()
        temp["line_key"] = temp["order_no"] + "||" + temp["sku"]
        for _, row in temp.iterrows():
            if row["line_key"] in line_keys:
                continue
            if row["order_no"] and row["order_no"] in order_keys:
                continue
            if row["sku"] and row["sku"] in sku_keys:
                continue
            anomalies.append(
                {
                    "issue_type": f"unmatched_{source_name}",
                    "severity": "medium",
                    "order_no": row["order_no"],
                    "sku": row["sku"],
                    "message": "费用表中存在未匹配到结算订单的记录。",
                    "amount": _round_money(row.get(amount_column, 0)),
                }
            )

    return anomalies


def _anomaly(issue_type: str, severity: str, row: pd.Series, message: str, amount: float) -> dict[str, Any]:
    return {
        "issue_type": issue_type,
        "severity": severity,
        "order_no": row["order_no"],
        "sku": row["sku"],
        "message": message,
        "amount": _round_money(amount),
    }


def _build_summary(
    *,
    profit_rows: pd.DataFrame,
    anomalies: list[dict[str, Any]],
    files: list[FinanceReconciliationInputFile],
    classified: list[dict[str, Any]],
    base_currency: str,
    instruction: str,
) -> list[tuple[str, Any]]:
    return [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("基础币种", base_currency),
        ("处理要求", instruction),
        ("上传文件数", len(files)),
        ("识别 Sheet 数", len(classified)),
        ("订单利润行数", len(profit_rows)),
        ("异常数量", len(anomalies)),
        ("销售额合计", _round_money(profit_rows["sales_amount_base"].sum())),
        ("退款合计", _round_money(profit_rows["refund_amount_base"].sum())),
        ("平台费合计", _round_money(profit_rows["platform_fee_base"].sum())),
        ("物流费合计", _round_money(profit_rows["logistics_fee_base"].sum())),
        ("广告费合计", _round_money(profit_rows["ad_fee_base"].sum())),
        ("采购成本合计", _round_money(profit_rows["purchase_cost_base"].sum())),
        ("利润合计", _round_money(profit_rows["profit_base"].sum())),
        ("亏损订单数", int((profit_rows["profit_base"] < 0).sum())),
        ("低毛利订单数", int((profit_rows["status"] == "低毛利").sum())),
    ]


def _build_reconciliation_workbook(
    *,
    summary: list[tuple[str, Any]],
    profit_rows: pd.DataFrame,
    anomalies: list[dict[str, Any]],
    field_mappings: list[dict[str, Any]],
    source_overview: list[dict[str, Any]],
) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "对账摘要"
    _write_key_value_sheet(summary_sheet, ["项目", "内容"], summary)
    _write_records_sheet(workbook, "订单利润表", _profit_records(profit_rows))
    _write_records_sheet(workbook, "异常账单", anomalies or [{"issue_type": "无", "severity": "-", "order_no": "-", "sku": "-", "message": "未发现异常", "amount": 0}])
    _write_records_sheet(workbook, "字段识别", field_mappings)
    _write_records_sheet(workbook, "源文件概览", source_overview)
    return workbook


def _profit_records(profit_rows: pd.DataFrame) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    labels = {
        "order_no": "订单号",
        "sku": "SKU",
        "quantity": "数量",
        "currency": "源币种",
        "sales_amount_base": "销售额",
        "refund_amount_base": "退款",
        "platform_fee_base": "平台费",
        "logistics_fee_base": "物流费",
        "ad_fee_base": "广告费",
        "purchase_cost_base": "采购成本",
        "net_revenue_base": "净收入",
        "total_cost_base": "总成本",
        "profit_base": "利润",
        "profit_margin": "利润率",
        "status": "状态",
        "source": "来源",
    }
    for _, row in profit_rows.head(MAX_OUTPUT_ROWS).iterrows():
        item = {}
        for key, label in labels.items():
            value = row[key]
            if key.endswith("_base") or key in {"profit_margin", "quantity"}:
                value = _round_money(value)
            item[label] = value
        records.append(item)
    return records


def _write_key_value_sheet(sheet, headers: list[str], rows: list[tuple[str, Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(list(row))
    _style_sheet(sheet)
    sheet.freeze_panes = "A2"
    _auto_width(sheet, max_width=70)


def _write_records_sheet(workbook: Workbook, title: str, records: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    if not records:
        sheet.append(["无数据"])
        _style_sheet(sheet)
        return

    headers = list(records[0].keys())
    sheet.append(headers)
    for record in records:
        sheet.append([_clean_excel_value(record.get(header)) for header in headers])

    _style_sheet(sheet)
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1 and sheet.max_column > 1:
        sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        if cell.value in {"销售额", "退款", "平台费", "物流费", "广告费", "采购成本", "净收入", "总成本", "利润", "amount"}:
            continue

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if sheet.cell(row=1, column=cell.column).value == "利润率":
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "#,##0.00"

    if title == "订单利润表":
        status_column = _find_header_column(sheet, "状态")
        if status_column:
            for cell in sheet.iter_rows(min_row=2, min_col=status_column, max_col=status_column):
                target = cell[0]
                if target.value == "亏损":
                    target.fill = PatternFill("solid", fgColor="FEE2E2")
                    target.font = Font(color="991B1B", bold=True)
                elif target.value == "低毛利":
                    target.fill = PatternFill("solid", fgColor="FEF3C7")
                    target.font = Font(color="92400E", bold=True)

    _auto_width(sheet)


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    thin_side = Side(style="thin", color="D6DEE8")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            if cell.row != 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def _auto_width(sheet, *, max_width: int = 42) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 8
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), max_width))
        sheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, max_width))


def _find_header_column(sheet, header: str) -> int | None:
    for cell in sheet[1]:
        if cell.value == header:
            return cell.column
    return None


def _source_overview(item: dict[str, Any]) -> dict[str, Any]:
    frame: pd.DataFrame = item["frame"]
    return {
        "文件": item["filename"],
        "Sheet": item["sheet_name"],
        "识别类型": _type_label(item["source_type"]),
        "行数": int(len(frame)),
        "列数": int(len(frame.columns)),
        "字段": ", ".join(str(column) for column in frame.columns),
    }


def _field_mapping(item: dict[str, Any]) -> dict[str, Any]:
    fields = item["fields"]
    return {
        "文件": item["filename"],
        "Sheet": item["sheet_name"],
        "识别类型": _type_label(item["source_type"]),
        "订单字段": fields.get("order_no") or "",
        "SKU字段": fields.get("sku") or "",
        "数量字段": fields.get("quantity") or "",
        "币种字段": fields.get("currency") or "",
        "销售字段": fields.get("sales_amount") or "",
        "退款字段": fields.get("refund_amount") or "",
        "平台费字段": ", ".join(fields.get("platform_fee") or []),
        "物流费字段": fields.get("logistics_fee") or "",
        "采购成本字段": fields.get("purchase_cost") or "",
        "广告费字段": fields.get("ad_fee") or "",
        "汇率字段": fields.get("exchange_rate") or "",
    }


def _type_label(source_type: str) -> str:
    return {
        "settlement": "平台结算/订单",
        "logistics": "物流账单",
        "purchase": "采购成本",
        "ads": "广告费用",
        "exchange_rate": "汇率",
        "unknown": "未识别",
    }.get(source_type, source_type)


def _sheet_type_counts(classified: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in classified:
        counts[item["source_type"]] = counts.get(item["source_type"], 0) + 1
    return counts


def _money_to_base(value: Any, currency: str, exchange_rates: dict[str, float]) -> float:
    amount = _to_number(value)
    rate = exchange_rates.get(currency, 1.0)
    return amount * rate


def _sum_columns_to_base(
    row: pd.Series,
    columns: list[str],
    currency: str,
    exchange_rates: dict[str, float],
    *,
    absolute: bool = False,
) -> float:
    total = 0.0
    for column in columns:
        value = _money_to_base(row.get(column), currency, exchange_rates)
        total += abs(value) if absolute else value
    return total


def _row_currency(row: pd.Series, currency_column: str | None, base_currency: str) -> str:
    if not currency_column:
        return base_currency
    return _clean_currency(row.get(currency_column)) or base_currency


def _to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return 0.0
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -abs(number) if negative else number


def _clean_currency(value: Any) -> str:
    text = _clean_text(value).upper()
    if not text:
        return ""
    mapping = {
        "人民币": "CNY",
        "美元": "USD",
        "欧元": "EUR",
        "日元": "JPY",
        "英镑": "GBP",
    }
    return mapping.get(text, text[:3])


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def _normalize_label(value: str) -> str:
    return re.sub(r"[\s_\-:/\\().,，。]+", "", str(value).strip().lower())


def _round_money(value: Any) -> float:
    return round(float(value or 0), 2)


def _clean_excel_value(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    if isinstance(value, (int, float, str, bool, datetime)):
        return value
    return str(value)
