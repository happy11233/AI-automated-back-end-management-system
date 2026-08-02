from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

from app.llm import chat


ALLOWED_EXCEL_EXTENSIONS = {".xlsx", ".xls"}
MAX_EXCEL_BYTES = 8 * 1024 * 1024
MAX_OUTPUT_ROWS_PER_SHEET = 2000
MAX_PROMPT_ROWS_PER_SHEET = 12

READABLE_HEADER_KEYWORDS = (
    "美化",
    "美观",
    "好看",
    "排版",
    "格式",
    "可读",
    "易读",
    "中文",
    "翻译",
    "表头",
    "字段名",
    "字段名称",
    "英文单词",
    "不好读",
)

FINANCE_FIELD_LABELS = {
    "name": "名称",
    "id": "编号",
    "title": "标题",
    "status": "状态",
    "company": "公司",
    "customer": "客户",
    "customer_name": "客户名称",
    "supplier": "供应商",
    "supplier_name": "供应商名称",
    "employee": "员工",
    "employee_name": "员工姓名",
    "employee_id": "员工编号",
    "department": "部门",
    "designation": "职位",
    "account": "会计科目",
    "account_name": "会计科目名称",
    "posting_date": "记账日期",
    "transaction_date": "交易日期",
    "creation": "创建时间",
    "modified": "更新时间",
    "start_date": "开始日期",
    "end_date": "结束日期",
    "due_date": "到期日期",
    "invoice_date": "发票日期",
    "invoice_number": "发票号码",
    "voucher_type": "凭证类型",
    "voucher_no": "凭证编号",
    "party_type": "往来对象类型",
    "party": "往来对象",
    "remarks": "备注",
    "description": "描述",
    "currency": "币种",
    "mode_of_payment": "付款方式",
    "cost_center": "成本中心",
    "project": "项目",
    "debit": "借方金额",
    "credit": "贷方金额",
    "amount": "金额",
    "total": "合计",
    "net_total": "未税合计",
    "grand_total": "含税合计",
    "base_total": "本位币合计",
    "base_net_total": "本位币未税合计",
    "base_grand_total": "本位币含税合计",
    "paid_amount": "已付金额",
    "received_amount": "已收金额",
    "outstanding_amount": "未结金额",
    "gross_pay": "应发工资",
    "net_pay": "实发工资",
    "salary": "工资",
    "salary_structure": "薪资结构",
    "total_sales": "销售总额",
    "total_expenses": "费用总额",
    "total_qty": "总数量",
    "quantity": "数量",
    "qty": "数量",
    "rate": "单价",
    "price": "价格",
    "tax": "税额",
    "taxes": "税费",
    "taxes_and_charges": "税费规则",
    "base_total_taxes_and_charges": "本位币税费",
    "exchange_rate": "汇率",
}

FINANCE_FIELD_PART_LABELS = {
    "base": "本位币",
    "total": "合计",
    "net": "未税",
    "grand": "含税",
    "gross": "应发",
    "net": "实发",
    "pay": "工资",
    "paid": "已付",
    "received": "已收",
    "outstanding": "未结",
    "amount": "金额",
    "date": "日期",
    "time": "时间",
    "number": "编号",
    "no": "编号",
    "type": "类型",
    "code": "编码",
    "id": "编号",
    "name": "名称",
    "status": "状态",
    "account": "科目",
    "party": "往来对象",
    "employee": "员工",
    "customer": "客户",
    "supplier": "供应商",
    "department": "部门",
    "currency": "币种",
    "description": "描述",
    "remarks": "备注",
    "quantity": "数量",
    "qty": "数量",
    "rate": "单价",
    "price": "价格",
    "tax": "税额",
    "exchange": "汇率",
    "creation": "创建",
    "modified": "更新",
    "start": "开始",
    "end": "结束",
    "due": "到期",
    "invoice": "发票",
    "voucher": "凭证",
    "posting": "记账",
    "transaction": "交易",
}


@dataclass
class FinanceExcelTransformResult:
    filename: str
    content: bytes
    metadata: dict[str, Any]


def transform_finance_excel(
    *,
    source_filename: str,
    content: bytes | None,
    instruction: str,
    erp_context: list[dict[str, Any]] | None = None,
) -> FinanceExcelTransformResult:
    normalized_erp_context = erp_context or []
    has_uploaded_content = content is not None and len(content) > 0
    normalized_source_filename = source_filename or (
        "finance_uploaded.xlsx" if has_uploaded_content else "finance_erp_generated.xlsx"
    )
    source_mode = "uploaded_excel" if has_uploaded_content else "erp_context"

    if has_uploaded_content:
        suffix = Path(normalized_source_filename or "").suffix.lower()
        if suffix not in ALLOWED_EXCEL_EXTENSIONS:
            raise ValueError("只支持上传 .xlsx 或 .xls 文件。")

        if len(content) > MAX_EXCEL_BYTES:
            raise ValueError("Excel 文件不能超过 8MB。")

        sheets = _read_excel_sheets(content, suffix)
        if not sheets:
            raise ValueError("Excel 文件里没有可读取的 sheet。")
        sheet_summaries = _build_sheet_summaries(sheets)
    else:
        sheets = {}
        sheet_summaries = _build_erp_context_sheet_summaries(normalized_erp_context)

    if not sheets and not normalized_erp_context:
        raise ValueError("请上传 Excel 文件，或选择/说明要使用的财务 ERP 表。")

    normalized_instruction = (
        instruction.strip()
        or "请按财务复核要求整理表格，生成数值汇总，并指出需要人工复核的异常。"
    )
    output_sheets, field_mappings = _prepare_output_sheets(
        sheets,
        instruction=normalized_instruction,
    )
    output_sheet_summaries = _build_sheet_summaries(output_sheets)
    ai_suggestion = _build_ai_suggestion(
        instruction=normalized_instruction,
        source_filename=normalized_source_filename,
        sheets=output_sheets,
        sheet_summaries=output_sheet_summaries,
        erp_context=normalized_erp_context,
    )
    workbook = _build_output_workbook(
        source_filename=normalized_source_filename,
        instruction=normalized_instruction,
        sheets=output_sheets,
        sheet_summaries=output_sheet_summaries,
        ai_suggestion=ai_suggestion,
        erp_context=normalized_erp_context,
        field_mappings=field_mappings,
    )

    output = BytesIO()
    workbook.save(output)
    output_content = output.getvalue()

    safe_stem = _safe_filename(Path(normalized_source_filename).stem or "finance_excel")
    output_filename = f"{safe_stem}_ai_finance_result.xlsx"

    return FinanceExcelTransformResult(
        filename=output_filename,
        content=output_content,
        metadata={
            "source_filename": normalized_source_filename,
            "source_mode": source_mode,
            "output_filename": output_filename,
            "sheet_count": len(sheet_summaries),
            "uploaded_sheet_count": len(sheets),
            "total_rows": sum(summary["row_count"] for summary in sheet_summaries),
            "total_columns": sum(summary["column_count"] for summary in sheet_summaries),
            "erp_resource_count": len(normalized_erp_context),
            "erp_resources": [
                {
                    "resource": str(item.get("resource") or ""),
                    "label": str(item.get("label") or item.get("resource") or ""),
                    "ok": bool(item.get("ok")),
                    "status": str(item.get("status") or ""),
                    "result_count": len(item.get("items") or []),
                }
                for item in normalized_erp_context
            ],
            "readable_headers_applied": bool(field_mappings),
            "field_mapping_count": len(field_mappings),
            "field_mappings": field_mappings[:200],
            "instruction_preview": normalized_instruction[:500],
            "output_bytes": len(output_content),
        },
    )


def _read_excel_sheets(content: bytes, suffix: str) -> dict[str, pd.DataFrame]:
    try:
        engine = "xlrd" if suffix == ".xls" else "openpyxl"
        raw_sheets = pd.read_excel(
            BytesIO(content),
            sheet_name=None,
            dtype=object,
            engine=engine,
        )
    except Exception as error:
        raise ValueError(f"Excel 文件读取失败：{error}") from error

    sheets: dict[str, pd.DataFrame] = {}
    for name, frame in raw_sheets.items():
        cleaned = frame.dropna(how="all")
        cleaned = cleaned.dropna(axis=1, how="all")
        cleaned = cleaned.fillna("")
        sheets[str(name)[:31] or "Sheet"] = cleaned

    return sheets


def _build_sheet_summaries(sheets: dict[str, pd.DataFrame]) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []

    for sheet_name, frame in sheets.items():
        numeric_columns = []
        for column in frame.columns:
            numeric_values = pd.to_numeric(frame[column], errors="coerce")
            numeric_count = int(numeric_values.notna().sum())
            if numeric_count == 0:
                continue

            numeric_columns.append(
                {
                    "column": str(column),
                    "count": numeric_count,
                    "sum": float(numeric_values.sum()),
                    "mean": float(numeric_values.mean()),
                    "min": float(numeric_values.min()),
                    "max": float(numeric_values.max()),
                }
            )

        summaries.append(
            {
                "sheet_name": sheet_name,
                "row_count": int(len(frame)),
                "column_count": int(len(frame.columns)),
                "columns": [str(column) for column in frame.columns],
                "numeric_columns": numeric_columns,
            }
        )

    return summaries


def _build_erp_context_sheet_summaries(erp_context: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sheets: dict[str, pd.DataFrame] = {}

    for item in erp_context:
        items = item.get("items")
        if not isinstance(items, list) or not items:
            continue

        rows = [_flatten_record(row) for row in items if isinstance(row, dict)]
        if not rows:
            continue

        sheet_name = f"ERP_{item.get('label') or item.get('resource') or 'resource'}"
        sheets[sheet_name[:31] or "ERP"] = pd.DataFrame(rows).fillna("")

    return _build_sheet_summaries(sheets)


def _build_ai_suggestion(
    *,
    instruction: str,
    source_filename: str,
    sheets: dict[str, pd.DataFrame],
    sheet_summaries: list[dict[str, Any]],
    erp_context: list[dict[str, Any]],
) -> str:
    prompt = _build_ai_prompt(
        instruction=instruction,
        source_filename=source_filename,
        sheets=sheets,
        sheet_summaries=sheet_summaries,
        erp_context=erp_context,
    )

    try:
        return str(chat(prompt)).strip()
    except Exception as error:
        return (
            "AI 建议生成失败，已先输出程序整理结果。\n"
            f"失败原因：{error}\n"
            "人工复核建议：检查数值列总额、空值行、重复单号、异常负数、币种和期间是否符合本次处理要求。"
        )


def _build_ai_prompt(
    *,
    instruction: str,
    source_filename: str,
    sheets: dict[str, pd.DataFrame],
    sheet_summaries: list[dict[str, Any]],
    erp_context: list[dict[str, Any]],
) -> str:
    samples: list[str] = []
    for sheet_name, frame in sheets.items():
        sample_rows = frame.head(MAX_PROMPT_ROWS_PER_SHEET).astype(str).to_dict("records")
        samples.append(
            f"Sheet: {sheet_name}\n"
            f"Columns: {', '.join(str(column) for column in frame.columns)}\n"
            f"Sample rows: {sample_rows}"
        )

    erp_summaries = [
        {
            "resource": str(item.get("resource") or ""),
            "label": str(item.get("label") or item.get("resource") or ""),
            "provider": str(item.get("provider_label") or item.get("provider") or ""),
            "status": str(item.get("status") or ""),
            "ok": bool(item.get("ok")),
            "message": str(item.get("message") or "")[:300],
            "result_count": len(item.get("items") or []),
            "sample_rows": (item.get("items") or [])[:5],
        }
        for item in erp_context
    ]

    return f"""你是跨境电商企业内部财务 AI 助手。
请只围绕财务岗位可见数据做表格整理建议，不要输出越权内容。

源文件：{source_filename}
财务处理要求：{instruction}
数据模式：{"上传 Excel + ERP 辅助" if sheets else "未上传 Excel，基于财务权限内 ERP 表生成新工作簿"}

表格概览：
{sheet_summaries}

样例数据：
{chr(10).join(samples) if samples else "本次未上传 Excel，样例数据请参考下方 ERP 表样例。"}

财务权限内 ERP 表：
{erp_summaries or "本次未选择 ERP 表"}

请输出：
1. 本次新 Excel 应如何使用
2. 上传 Excel 与已选 ERP 表，或多张 ERP 表之间，可以如何合并或核对
3. 关键汇总指标应该重点看哪些列
4. 可能的异常和复核点
5. 如果要继续生成工资/利润/费用分析表，下一步需要补充哪些字段
"""


def _build_output_workbook(
    *,
    source_filename: str,
    instruction: str,
    sheets: dict[str, pd.DataFrame],
    sheet_summaries: list[dict[str, Any]],
    ai_suggestion: str,
    erp_context: list[dict[str, Any]],
    field_mappings: list[dict[str, str]] | None = None,
) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "处理摘要"

    _write_summary_sheet(
        summary_sheet,
        source_filename=source_filename,
        instruction=instruction,
        sheet_summaries=sheet_summaries,
    )
    if field_mappings:
        _write_field_mapping_sheet(workbook, field_mappings)
    _write_numeric_summary_sheet(workbook, sheet_summaries)
    _write_ai_suggestion_sheet(workbook, ai_suggestion)

    used_titles = set(workbook.sheetnames)
    if erp_context:
        _write_erp_summary_sheet(workbook, erp_context)
        _write_erp_combined_sheet(workbook, erp_context)
        used_titles = set(workbook.sheetnames)
        for item in erp_context:
            output_sheet = workbook.create_sheet(
                title=_unique_sheet_title(f"ERP_{item.get('resource') or 'resource'}", used_titles)
            )
            _write_erp_resource_sheet(output_sheet, item)

    for sheet_name, frame in sheets.items():
        output_sheet = workbook.create_sheet(
            title=_unique_sheet_title(f"整理_{sheet_name}", used_titles)
        )
        _write_dataframe_sheet(output_sheet, frame)

    return workbook


def _write_summary_sheet(
    sheet,
    *,
    source_filename: str,
    instruction: str,
    sheet_summaries: list[dict[str, Any]],
) -> None:
    total_rows = sum(summary["row_count"] for summary in sheet_summaries)
    total_columns = sum(summary["column_count"] for summary in sheet_summaries)
    rows = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("源文件", source_filename),
        ("处理要求", instruction),
        ("源 Sheet 数", len(sheet_summaries)),
        ("总数据行数", total_rows),
        ("总字段数", total_columns),
        ("输出规则", f"每个源 Sheet 最多输出 {MAX_OUTPUT_ROWS_PER_SHEET} 行，超出部分请回到源文件复核。"),
    ]

    sheet.append(["项目", "内容"])
    for row in rows:
        sheet.append(list(row))

    start_row = len(rows) + 4
    sheet.cell(row=start_row, column=1, value="Sheet")
    sheet.cell(row=start_row, column=2, value="行数")
    sheet.cell(row=start_row, column=3, value="列数")
    sheet.cell(row=start_row, column=4, value="字段")

    for index, summary in enumerate(sheet_summaries, start=start_row + 1):
        sheet.cell(row=index, column=1, value=summary["sheet_name"])
        sheet.cell(row=index, column=2, value=summary["row_count"])
        sheet.cell(row=index, column=3, value=summary["column_count"])
        sheet.cell(row=index, column=4, value=", ".join(summary["columns"]))

    _style_sheet(sheet)
    sheet.freeze_panes = "A2"
    _auto_width(sheet, max_width=70)


def _write_erp_summary_sheet(
    workbook: Workbook,
    erp_context: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("ERP数据摘要")
    sheet.append(["ERP资源", "资源名称", "提供方", "状态", "结果数", "说明"])

    for item in erp_context:
        sheet.append(
            [
                str(item.get("resource") or ""),
                str(item.get("label") or item.get("resource") or ""),
                str(item.get("provider_label") or item.get("provider") or ""),
                "成功" if item.get("ok") else str(item.get("status") or "失败"),
                len(item.get("items") or []),
                str(item.get("message") or ""),
            ]
        )

    _style_sheet(sheet)
    sheet.freeze_panes = "A2"
    _auto_width(sheet, max_width=60)


def _write_erp_combined_sheet(
    workbook: Workbook,
    erp_context: list[dict[str, Any]],
) -> None:
    rows: list[dict[str, Any]] = []
    field_names: list[str] = []

    for item in erp_context:
        items = item.get("items")
        if not isinstance(items, list):
            continue

        for raw_row in items[:MAX_OUTPUT_ROWS_PER_SHEET]:
            if not isinstance(raw_row, dict):
                continue

            flattened = _flatten_record(raw_row)
            for key in flattened:
                if key not in field_names:
                    field_names.append(key)

            rows.append(
                {
                    "ERP资源": str(item.get("resource") or ""),
                    "资源名称": str(item.get("label") or item.get("resource") or ""),
                    "查询状态": "成功" if item.get("ok") else str(item.get("status") or "失败"),
                    **flattened,
                }
            )

    sheet = workbook.create_sheet("ERP组合明细")
    headers = ["ERP资源", "资源名称", "查询状态", *field_names]
    sheet.append(headers)

    if not rows:
        sheet.append(["-", "-", "未返回可合并的 ERP 明细数据"])
    else:
        for row in rows[:MAX_OUTPUT_ROWS_PER_SHEET]:
            sheet.append([_clean_excel_value(row.get(header, "")) for header in headers])

    _style_sheet(sheet)
    sheet.freeze_panes = "A2"
    _auto_width(sheet, max_width=48)


def _write_erp_resource_sheet(sheet, item: dict[str, Any]) -> None:
    items = item.get("items")
    if not isinstance(items, list) or not items:
        sheet.append(["ERP资源", "状态", "说明"])
        sheet.append(
            [
                str(item.get("resource") or ""),
                str(item.get("status") or ""),
                str(item.get("message") or "未返回可写入的数据"),
            ]
        )
        _style_sheet(sheet)
        sheet.freeze_panes = "A2"
        _auto_width(sheet, max_width=60)
        return

    frame = pd.DataFrame([_flatten_record(row) for row in items if isinstance(row, dict)])
    if frame.empty:
        sheet.append(["ERP资源", "状态", "说明"])
        sheet.append([str(item.get("resource") or ""), str(item.get("status") or ""), "ERP 返回的数据格式无法写入表格"])
        _style_sheet(sheet)
        sheet.freeze_panes = "A2"
        _auto_width(sheet, max_width=60)
        return

    _write_dataframe_sheet(sheet, frame)


def _write_numeric_summary_sheet(
    workbook: Workbook,
    sheet_summaries: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("数值汇总")
    sheet.append(["Sheet", "字段", "有效数值数", "合计", "平均", "最小值", "最大值"])

    wrote_row = False
    for summary in sheet_summaries:
        for item in summary["numeric_columns"]:
            wrote_row = True
            sheet.append(
                [
                    summary["sheet_name"],
                    item["column"],
                    item["count"],
                    item["sum"],
                    item["mean"],
                    item["min"],
                    item["max"],
                ]
            )

    if not wrote_row:
        sheet.append(["-", "未识别到数值列", 0, 0, 0, 0, 0])

    _style_sheet(sheet)
    sheet.freeze_panes = "A2"
    _auto_width(sheet)


def _write_ai_suggestion_sheet(workbook: Workbook, ai_suggestion: str) -> None:
    sheet = workbook.create_sheet("AI建议")
    sheet.append(["AI 财务处理建议"])

    for line in ai_suggestion.splitlines() or ["暂无建议"]:
        sheet.append([line])

    _style_sheet(sheet)
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 120
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_field_mapping_sheet(workbook: Workbook, field_mappings: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet("字段说明")
    sheet.append(["来源 Sheet", "原字段名", "业务显示名", "说明"])
    for item in field_mappings:
        original_name = item["original_name"]
        display_name = item["display_name"]
        note = "已转换为中文业务字段名" if display_name != original_name else "原字段已具备可读名称"
        sheet.append([item["sheet_name"], original_name, display_name, note])

    _style_sheet(sheet)
    sheet.freeze_panes = "A2"
    _auto_width(sheet, max_width=60)


def _write_dataframe_sheet(sheet, frame: pd.DataFrame) -> None:
    headers = [str(column) if str(column) else f"Column {index}" for index, column in enumerate(frame.columns, start=1)]
    sheet.append(headers)

    for _, row in frame.head(MAX_OUTPUT_ROWS_PER_SHEET).iterrows():
        sheet.append([_clean_excel_value(value) for value in row.tolist()])

    if len(frame) > MAX_OUTPUT_ROWS_PER_SHEET:
        sheet.append([f"已截断：源 Sheet 共 {len(frame)} 行，本结果只输出前 {MAX_OUTPUT_ROWS_PER_SHEET} 行。"])

    _style_sheet(sheet)
    sheet.freeze_panes = "A2"
    _auto_width(sheet)


def _prepare_output_sheets(
    sheets: dict[str, pd.DataFrame],
    *,
    instruction: str,
) -> tuple[dict[str, pd.DataFrame], list[dict[str, str]]]:
    if not _requires_readable_headers(instruction):
        return sheets, []

    output_sheets: dict[str, pd.DataFrame] = {}
    mappings: list[dict[str, str]] = []
    for sheet_name, frame in sheets.items():
        used_names: set[str] = set()
        renamed_columns: list[str] = []
        for column in frame.columns:
            original_name = str(column)
            display_name = _readable_field_name(original_name)
            display_name = _unique_display_name(display_name, used_names)
            used_names.add(display_name)
            renamed_columns.append(display_name)
            mappings.append(
                {
                    "sheet_name": str(sheet_name),
                    "original_name": original_name,
                    "display_name": display_name,
                }
            )
        output_sheets[sheet_name] = frame.copy().set_axis(renamed_columns, axis="columns")
    return output_sheets, mappings


def _requires_readable_headers(instruction: str) -> bool:
    lowered = str(instruction or "").lower()
    return any(keyword in lowered for keyword in READABLE_HEADER_KEYWORDS)


def _readable_field_name(value: str) -> str:
    original = str(value or "").strip()
    if not original:
        return "未命名字段"
    if re.search(r"[\u4e00-\u9fff]", original):
        return original

    normalized = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", original)
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", normalized).strip("_").lower()
    if not normalized:
        return original
    if normalized in FINANCE_FIELD_LABELS:
        return FINANCE_FIELD_LABELS[normalized]

    parts = [item for item in normalized.split("_") if item]
    translated = [FINANCE_FIELD_PART_LABELS.get(item, item) for item in parts]
    if translated and any(item != source for item, source in zip(translated, parts)):
        return "".join(translated)
    return f"字段：{normalized.replace('_', ' ')}"


def _unique_display_name(value: str, used_names: set[str]) -> str:
    if value not in used_names:
        return value
    suffix = 2
    while f"{value}{suffix}" in used_names:
        suffix += 1
    return f"{value}{suffix}"


def _flatten_record(record: dict[str, Any]) -> dict[str, Any]:
    return {
        str(key): _clean_excel_value(value)
        for key, value in record.items()
    }


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_side = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            if cell.row != 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    if sheet.max_row > 1 and sheet.max_column > 1:
        sheet.auto_filter.ref = sheet.dimensions


def _auto_width(sheet, *, max_width: int = 38) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 8
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), max_width))
        sheet.column_dimensions[column_letter].width = max(10, min(max_length + 2, max_width))


def _clean_excel_value(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, (dict, list, tuple, set)):
        return str(value)

    if pd.isna(value):
        return ""

    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()

    if isinstance(value, (int, float, str, bool, datetime)):
        return value

    return str(value)


def _safe_filename(value: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._")
    return normalized or "finance_excel"


def _unique_sheet_title(title: str, used_titles: set[str]) -> str:
    clean_title = re.sub(r"[\[\]:*?/\\]", "_", title).strip() or "Sheet"
    clean_title = clean_title[:31]
    candidate = clean_title
    index = 2

    while candidate in used_titles:
        suffix = f"_{index}"
        candidate = f"{clean_title[:31 - len(suffix)]}{suffix}"
        index += 1

    used_titles.add(candidate)
    return candidate
