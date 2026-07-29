from __future__ import annotations

import calendar
import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.erp.base import ERPProviderError
from app.erp.providers import get_active_provider
from app.erp.resources import provider_fields_for, provider_resource_for


MAX_SALARY_ROWS = 500


@dataclass
class SalaryExportIntent:
    intent: str
    period_label: str
    start_date: date
    end_date: date
    output_format: str
    confidence: float
    matched_keywords: list[str]


@dataclass
class FinanceSalaryExportResult:
    filename: str
    content: bytes
    metadata: dict[str, Any]
    intent: SalaryExportIntent
    items: list[dict[str, Any]]
    provider: str
    provider_label: str
    provider_resource: str


def recognize_salary_export_intent(message: str, today: date | None = None) -> SalaryExportIntent:
    normalized = (message or "").strip()
    lowered = normalized.lower()
    current_day = today or date.today()
    matched_keywords: list[str] = []

    keyword_groups = {
        "salary": ["工资", "薪资", "工资表", "工资单", "薪水", "薪酬", "salary", "payroll"],
        "export": ["发给我", "导出", "生成", "下载", "excel", "xlsx", "表", "明细", "清单"],
        "all_staff": ["所有员工", "全部员工", "全体员工", "员工"],
    }
    for keywords in keyword_groups.values():
        for keyword in keywords:
            if keyword in lowered:
                matched_keywords.append(keyword)

    start_date, end_date, period_label = _resolve_period(normalized, current_day)
    has_salary = bool(set(matched_keywords) & set(keyword_groups["salary"]))
    has_export = bool(set(matched_keywords) & set(keyword_groups["export"]))
    confidence = 0.45
    if has_salary:
        confidence += 0.35
    if has_export:
        confidence += 0.15
    if start_date and end_date:
        confidence += 0.05

    return SalaryExportIntent(
        intent="finance_salary_export" if has_salary else "unknown",
        period_label=period_label,
        start_date=start_date,
        end_date=end_date,
        output_format="xlsx",
        confidence=round(min(confidence, 0.98), 2),
        matched_keywords=matched_keywords,
    )


def export_salary_workbook_from_erp(
    *,
    message: str,
    current_user: dict,
    today: date | None = None,
    intent: SalaryExportIntent | None = None,
) -> FinanceSalaryExportResult:
    intent = intent or recognize_salary_export_intent(message, today=today)
    if intent.intent != "finance_salary_export":
        raise ValueError("没有识别到工资表导出意图，请说明要导出哪个月份的员工工资表。")

    provider = get_active_provider()
    provider_resource = provider_resource_for("Salary Slip", provider.provider_id)
    if provider_resource is None:
        raise ValueError(f"{provider.provider_label} 暂未映射工资单资源。")

    filters = [
        ["start_date", ">=", intent.start_date.isoformat()],
        ["end_date", "<=", intent.end_date.isoformat()],
    ]
    try:
        result = provider.query_resource(
            resource="Salary Slip",
            provider_resource=provider_resource,
            query=None,
            filters=filters,
            fields=provider_fields_for("Salary Slip", provider.provider_id),
            limit=MAX_SALARY_ROWS,
        )
    except ERPProviderError as error:
        raise ValueError(error.message) from error
    items = result.get("items") if isinstance(result.get("items"), list) else []
    if not result.get("ok"):
        raise ValueError(str(result.get("message") or "ERP 工资单查询失败。"))
    if not items:
        raise ValueError(f"没有查到 {intent.period_label} 的员工工资单。")

    sorted_items = sorted(items, key=lambda item: str(item.get("employee") or item.get("employee_name") or ""))
    workbook = _build_salary_workbook(
        intent=intent,
        items=sorted_items,
        provider_label=provider.provider_label,
        username=str(current_user.get("username") or ""),
    )
    output = BytesIO()
    workbook.save(output)
    content = output.getvalue()
    safe_period = intent.start_date.strftime("%Y%m")
    filename = f"finance_salary_{safe_period}.xlsx"
    gross_total = _money_sum(sorted_items, "gross_pay")
    net_total = _money_sum(sorted_items, "net_pay")

    return FinanceSalaryExportResult(
        filename=filename,
        content=content,
        metadata={
            "output_filename": filename,
            "provider": provider.provider_id,
            "provider_label": provider.provider_label,
            "provider_resource": provider_resource,
            "resource": "Salary Slip",
            "period_label": intent.period_label,
            "start_date": intent.start_date.isoformat(),
            "end_date": intent.end_date.isoformat(),
            "employee_count": len(sorted_items),
            "gross_pay_total": gross_total,
            "net_pay_total": net_total,
            "intent": intent.intent,
            "intent_confidence": intent.confidence,
            "matched_keywords": intent.matched_keywords,
            "output_bytes": len(content),
            "input_preview": message[:500],
        },
        intent=intent,
        items=sorted_items,
        provider=provider.provider_id,
        provider_label=provider.provider_label,
        provider_resource=provider_resource,
    )


def _resolve_period(text: str, today: date) -> tuple[date, date, str]:
    month_match = re.search(r"(20\d{2})[-年/\.](0?[1-9]|1[0-2])", text)
    if month_match:
        year = int(month_match.group(1))
        month = int(month_match.group(2))
        return _month_range(year, month)

    chinese_month_match = re.search(r"(0?[1-9]|1[0-2])月", text)
    if chinese_month_match and "下个月" not in text and "上个月" not in text:
        month = int(chinese_month_match.group(1))
        return _month_range(today.year, month)

    if any(keyword in text for keyword in ["上个月", "上月", "last month"]):
        year = today.year
        month = today.month - 1
        if month == 0:
            year -= 1
            month = 12
        return _month_range(year, month)

    if any(keyword in text for keyword in ["下个月", "下月", "next month"]):
        year = today.year
        month = today.month + 1
        if month == 13:
            year += 1
            month = 1
        return _month_range(year, month)

    return _month_range(today.year, today.month)


def _month_range(year: int, month: int) -> tuple[date, date, str]:
    last_day = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)
    end_date = date(year, month, last_day)
    return start_date, end_date, f"{year}年{month:02d}月"


def _build_salary_workbook(
    *,
    intent: SalaryExportIntent,
    items: list[dict[str, Any]],
    provider_label: str,
    username: str,
) -> Workbook:
    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "工资明细"
    summary_sheet = workbook.create_sheet("自动化摘要")
    intent_sheet = workbook.create_sheet("意图识别")

    _write_detail_sheet(detail_sheet, items)
    _write_summary_sheet(
        summary_sheet,
        intent=intent,
        items=items,
        provider_label=provider_label,
        username=username,
    )
    _write_intent_sheet(intent_sheet, intent)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        _auto_width(sheet)

    return workbook


def _write_detail_sheet(sheet, items: list[dict[str, Any]]) -> None:
    headers = ["工资单号", "员工编号", "员工姓名", "开始日期", "结束日期", "应发工资", "实发工资", "状态"]
    sheet.append(headers)
    for item in items:
        sheet.append([
            item.get("name") or "",
            item.get("employee") or "",
            item.get("employee_name") or "",
            item.get("start_date") or "",
            item.get("end_date") or "",
            _clean_money(item.get("gross_pay")),
            _clean_money(item.get("net_pay")),
            item.get("status") or "",
        ])
    _style_table(sheet, money_columns={6, 7})


def _write_summary_sheet(
    sheet,
    *,
    intent: SalaryExportIntent,
    items: list[dict[str, Any]],
    provider_label: str,
    username: str,
) -> None:
    gross_total = _money_sum(items, "gross_pay")
    net_total = _money_sum(items, "net_pay")
    rows = [
        ("自动化结果", "已根据财务自然语言请求，自动识别工资表导出意图，查询 ERP 工资单并生成 Excel。"),
        ("请求期间", intent.period_label),
        ("日期范围", f"{intent.start_date.isoformat()} 至 {intent.end_date.isoformat()}"),
        ("ERP 数据源", provider_label),
        ("工资单数量", len(items)),
        ("应发工资合计", gross_total),
        ("实发工资合计", net_total),
        ("意图置信度", intent.confidence),
        ("导出人", username),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("复核建议", _summary_suggestion(items)),
    ]
    sheet.append(["项目", "内容"])
    for row in rows:
        sheet.append(list(row))
    _style_table(sheet, money_columns={2})
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")


def _write_intent_sheet(sheet, intent: SalaryExportIntent) -> None:
    sheet.append(["字段", "识别结果"])
    rows = [
        ("intent", intent.intent),
        ("period_label", intent.period_label),
        ("start_date", intent.start_date.isoformat()),
        ("end_date", intent.end_date.isoformat()),
        ("output_format", intent.output_format),
        ("confidence", intent.confidence),
        ("matched_keywords", "、".join(intent.matched_keywords)),
    ]
    for row in rows:
        sheet.append(list(row))
    _style_table(sheet)


def _summary_suggestion(items: list[dict[str, Any]]) -> str:
    if not items:
        return "未查询到工资单，请复核 ERP 工资期间。"

    missing_net = [item for item in items if item.get("net_pay") in (None, "")]
    negative_net = [item for item in items if _clean_money(item.get("net_pay")) < 0]
    if missing_net or negative_net:
        return "存在实发工资为空或负数的记录，建议财务人工复核后再发放。"

    return "工资单金额字段完整，建议财务复核员工名单、发放状态和审批流程后使用。"


def _style_table(sheet, money_columns: set[int] | None = None) -> None:
    money_columns = money_columns or set()
    header_fill = PatternFill("solid", fgColor="EEF4FF")
    header_font = Font(bold=True, color="1F2937")
    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            if cell.column in money_columns and cell.row > 1:
                cell.number_format = '#,##0.00'


def _auto_width(sheet) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 42)


def _money_sum(items: list[dict[str, Any]], key: str) -> float:
    return round(sum(_clean_money(item.get(key)) for item in items), 2)


def _clean_money(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return 0.0
