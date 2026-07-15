import http.client
import json
import os
import subprocess
import sys
import time
from pathlib import Path
from tempfile import TemporaryDirectory
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
FRONTEND_DIR = ROOT_DIR / "frontend"
DEFAULT_API_BASE_URL = "http://127.0.0.1:8001"
WORKSPACE_NODE_MODULES = Path(
    "/Users/xiaoxiang/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/node_modules"
)


class VerificationError(RuntimeError):
    pass


def main() -> None:
    api_base_url = os.getenv("VERIFY_API_BASE_URL", DEFAULT_API_BASE_URL).rstrip("/")
    python_bin = _python_bin()

    steps = [
        (
            "后端 Python 语法编译",
            lambda: _run([str(python_bin), "-m", "compileall", "app", "scripts"]),
        ),
        ("API 健康检查", lambda: _check_api_health(api_base_url)),
        ("ERP 管理员诊断", lambda: _check_erp_diagnostics(api_base_url)),
        (
            "ERP 对话权限回归",
            lambda: _run([str(python_bin), "scripts/verify_erp_chat.py"]),
        ),
        (
            "岗位越权权限回归",
            lambda: _run([str(python_bin), "scripts/verify_position_permissions.py"]),
        ),
        (
            "发布前稳定化回归",
            lambda: _run([str(python_bin), "scripts/verify_release_ready.py"]),
        ),
        ("财务 Excel 生成回归", lambda: _check_finance_excel_transform(api_base_url)),
        ("前端构建", lambda: _run(["npm", "run", "build"], cwd=FRONTEND_DIR)),
        ("前端权限可见性回归", lambda: _check_frontend_permissions()),
    ]

    started_at = time.monotonic()

    for index, (label, action) in enumerate(steps, start=1):
        print(f"\n[{index}/{len(steps)}] {label}", flush=True)
        step_started_at = time.monotonic()
        try:
            action()
        except Exception as error:
            raise SystemExit(f"\n验证失败：{label}\n原因：{error}") from error
        print(f"通过：{label}（{time.monotonic() - step_started_at:.1f}s）", flush=True)

    print(f"\n全部验证通过（总耗时 {time.monotonic() - started_at:.1f}s）", flush=True)


def _python_bin() -> Path:
    venv_python = ROOT_DIR / ".venv" / "bin" / "python"
    if venv_python.exists():
        return venv_python

    return Path(sys.executable)


def _run(command: list[str], cwd: Path | None = None) -> None:
    subprocess.run(command, cwd=cwd or ROOT_DIR, check=True)


def _run_node(command: list[str], cwd: Path | None = None) -> None:
    env = os.environ.copy()
    if WORKSPACE_NODE_MODULES.exists():
        env["NODE_PATH"] = str(WORKSPACE_NODE_MODULES)

    subprocess.run(command, cwd=cwd or ROOT_DIR, check=True, env=env)


def _check_frontend_permissions() -> None:
    _run_node(["node", "scripts/verify_frontend_permissions.mjs"])


def _check_api_health(api_base_url: str) -> None:
    last_error: Exception | None = None
    for _ in range(10):
        try:
            payload = _request_json(f"{api_base_url}/health", method="GET", timeout=8)
            break
        except (VerificationError, ConnectionResetError, http.client.RemoteDisconnected) as error:
            last_error = error
            time.sleep(1)
    else:
        raise VerificationError(f"API 健康检查失败：{last_error}") from last_error

    if payload.get("status") != "ok":
        raise VerificationError(f"/health 返回异常：{payload}")


def _check_erp_diagnostics(api_base_url: str) -> None:
    token = _login_admin(api_base_url)
    payload = _request_json(
        f"{api_base_url}/erp/diagnostics",
        method="GET",
        headers={"Authorization": f"Bearer {token}"},
        timeout=20,
    )

    active_health = payload.get("active_health") or {}
    status = active_health.get("status")
    ok = bool(active_health.get("ok"))

    if not ok or status != "ok":
        message = active_health.get("message") or "未知 ERP 诊断错误"
        raise VerificationError(f"ERP 诊断未通过：status={status}, message={message}")

    detail = active_health.get("detail")
    if isinstance(detail, dict):
        logged_user = detail.get("user") or detail.get("message") or "unknown"
    else:
        logged_user = detail or "unknown"
    print(
        json.dumps(
            {
                "active_provider": payload.get("active_provider"),
                "status": status,
                "logged_user": logged_user,
            },
            ensure_ascii=False,
        )
    )


def _login_admin(api_base_url: str) -> str:
    username = os.getenv("VERIFY_ADMIN_USERNAME", "admin_demo")
    password = os.getenv("VERIFY_ADMIN_PASSWORD", "Admin123456")
    body = urlencode({"username": username, "password": password}).encode()
    payload = _request_json(
        f"{api_base_url}/auth/login",
        method="POST",
        body=body,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=10,
    )

    token = payload.get("access_token")
    if not token:
        raise VerificationError("管理员登录成功响应中没有 access_token")

    return str(token)


def _login_user(api_base_url: str, username: str, password: str) -> str:
    body = urlencode({"username": username, "password": password}).encode()
    payload = _request_json(
        f"{api_base_url}/auth/login",
        method="POST",
        body=body,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=10,
    )

    token = payload.get("access_token")
    if not token:
        raise VerificationError(f"{username} 登录成功响应中没有 access_token")

    return str(token)


def _check_finance_excel_transform(api_base_url: str) -> None:
    token = _login_user(api_base_url, "finance_demo", "Finance123456")

    with TemporaryDirectory() as temp_dir:
        input_path = Path(temp_dir) / "finance_input.xlsx"
        output_path = Path(temp_dir) / "finance_output.xlsx"
        _build_sample_finance_workbook(input_path)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        boundary = f"----codex-finance-{int(time.time() * 1000)}"
        body = _build_multipart_body(
            boundary=boundary,
            fields={
                "instruction": "按店铺统计销售额、成本、利润，标记利润为负或金额为空的记录。",
            },
            files={
                "file": {
                    "filename": input_path.name,
                    "content_type": content_type,
                    "content": input_path.read_bytes(),
                },
            },
        )
        raw = _request_bytes(
            f"{api_base_url}/automation/finance/excel-transform",
            method="POST",
            body=body,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": f"multipart/form-data; boundary={boundary}",
            },
            timeout=60,
        )
        output_path.write_bytes(raw)

        workbook = load_workbook(output_path, data_only=True)
        expected_sheets = {"处理摘要", "数值汇总", "AI建议"}
        missing_sheets = expected_sheets.difference(workbook.sheetnames)
        if missing_sheets:
            raise VerificationError(f"生成 Excel 缺少 sheet：{', '.join(sorted(missing_sheets))}")

        if workbook["数值汇总"].max_row < 2:
            raise VerificationError("生成 Excel 的数值汇总为空")

        print(
            json.dumps(
                {
                    "filename": output_path.name,
                    "sheets": workbook.sheetnames,
                    "bytes": len(raw),
                },
                ensure_ascii=False,
            )
        )


def _build_sample_finance_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Amazon销售明细"
    sheet.append(["店铺", "订单号", "销售额", "成本", "利润"])
    sheet.append(["US Store", "AMZ-US-112-4589012-7783401", 59.99, 31.5, 28.49])
    sheet.append(["DE Store", "AMZ-DE-305-7712468-1290045", 42.5, 25.2, 17.3])
    sheet.append(["JP Store", "AMZ-JP-250-6630188-4402197", 41.97, 44.0, -2.03])
    workbook.save(path)


def _build_multipart_body(
    *,
    boundary: str,
    fields: dict[str, str],
    files: dict[str, dict[str, bytes | str]],
) -> bytes:
    chunks: list[bytes] = []

    for name, value in fields.items():
        chunks.extend(
            [
                f"--{boundary}\r\n".encode(),
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode(),
                value.encode("utf-8"),
                b"\r\n",
            ]
        )

    for name, file_item in files.items():
        filename = str(file_item["filename"])
        content_type = str(file_item["content_type"])
        content = file_item["content"]
        if not isinstance(content, bytes):
            raise TypeError("multipart file content must be bytes")

        chunks.extend(
            [
                f"--{boundary}\r\n".encode(),
                (
                    f'Content-Disposition: form-data; name="{name}"; '
                    f'filename="{filename}"\r\n'
                ).encode(),
                f"Content-Type: {content_type}\r\n\r\n".encode(),
                content,
                b"\r\n",
            ]
        )

    chunks.append(f"--{boundary}--\r\n".encode())
    return b"".join(chunks)


def _request_json(
    url: str,
    *,
    method: str,
    body: bytes | None = None,
    headers: dict[str, str] | None = None,
    timeout: int,
) -> dict:
    request = Request(url, data=body, headers=headers or {}, method=method)

    try:
        with urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8")
    except HTTPError as error:
        raw = error.read().decode("utf-8", errors="replace")
        raise VerificationError(f"{url} 返回 HTTP {error.code}: {_preview(raw)}") from error
    except URLError as error:
        raise VerificationError(f"{url} 无法连接：{error.reason}") from error
    except TimeoutError as error:
        raise VerificationError(f"{url} 请求超时") from error

    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as error:
        raise VerificationError(f"{url} 返回不是 JSON：{_preview(raw)}") from error

    if not isinstance(payload, dict):
        raise VerificationError(f"{url} 返回 JSON 不是对象：{payload!r}")

    return payload


def _request_bytes(
    url: str,
    *,
    method: str,
    body: bytes | None = None,
    headers: dict[str, str] | None = None,
    timeout: int,
) -> bytes:
    request = Request(url, data=body, headers=headers or {}, method=method)

    try:
        with urlopen(request, timeout=timeout) as response:
            return response.read()
    except HTTPError as error:
        raw = error.read().decode("utf-8", errors="replace")
        raise VerificationError(f"{url} 返回 HTTP {error.code}: {_preview(raw)}") from error
    except URLError as error:
        raise VerificationError(f"{url} 无法连接：{error.reason}") from error
    except TimeoutError as error:
        raise VerificationError(f"{url} 请求超时") from error


def _preview(value: str, length: int = 300) -> str:
    return " ".join(value.split())[:length]


if __name__ == "__main__":
    main()
