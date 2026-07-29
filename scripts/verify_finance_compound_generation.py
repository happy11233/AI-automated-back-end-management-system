from __future__ import annotations

from io import BytesIO
import base64
import json
import os
from pathlib import Path
import sys
import time
from typing import Any

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.services.chat_react_decision_service import decide_chat_action
from app.services.finance_compound_intent_service import (
    FINANCE_COMPOUND_INTENT,
    FINANCE_REPORT_OUTPUT,
    FINANCE_SALARY_OUTPUT,
    recognize_finance_compound_intent,
)


API_BASE_URL = os.getenv("VERIFY_API_BASE_URL", "http://127.0.0.1:8001").rstrip("/")


def main() -> None:
    verify_intent_synonyms()
    wait_for_api()
    token = login("finance_demo", "Finance123456")["access_token"]

    message = "帮我生成这个月的财务报表和工资表"
    chat_data = post_json(token, "/chat", {"message": message}, timeout=180)
    assert chat_data["intent"] == FINANCE_COMPOUND_INTENT, chat_data
    assert "财务报表" in chat_data["answer"], chat_data["answer"]
    assert "工资表" in chat_data["answer"], chat_data["answer"]
    assert "没有要求发送邮箱或微信" in chat_data["answer"], chat_data["answer"]
    assert len(chat_data["attachments"]) >= 2, chat_data

    filenames = [item["filename"] for item in chat_data["attachments"]]
    assert any(name.startswith("finance_monthly_report_") for name in filenames), filenames
    assert any(name.startswith("finance_salary_") for name in filenames), filenames
    for attachment in chat_data["attachments"]:
        decoded = base64.b64decode(attachment["content_base64"])
        assert decoded[:2] == b"PK", attachment["filename"]
        workbook = load_workbook(BytesIO(decoded), data_only=True)
        assert workbook.sheetnames, attachment["filename"]

    stream_data = post_sse(token, "/chat/stream", {"message": "生成本月薪资表和财务月报"}, timeout=180)
    assert stream_data["intent"] == FINANCE_COMPOUND_INTENT, stream_data
    assert len(stream_data["attachments"]) >= 2, stream_data

    print(json.dumps({
        "ok": True,
        "chat_intent": chat_data["intent"],
        "chat_files": filenames,
        "stream_files": [item["filename"] for item in stream_data["attachments"]],
        "note": "real API, real auth, real /chat and /chat/stream, real ERP finance resources, real xlsx attachments; no mock/stub/fake",
    }, ensure_ascii=False, indent=2))


def verify_intent_synonyms() -> None:
    current_user = {"role": "employee", "position": "finance"}
    messages = [
        "帮我生成这个月的财务报表和工资表",
        "生成本月薪资表和财务月报",
        "导出这个月薪水表和经营汇总",
        "做一份本月薪酬表和财报",
        "帮我生成本月 payroll 和月度经营汇总",
    ]
    for message in messages:
        intent = recognize_finance_compound_intent(message)
        assert intent.intent == FINANCE_COMPOUND_INTENT, (message, intent)
        assert FINANCE_REPORT_OUTPUT in intent.outputs, (message, intent.outputs)
        assert FINANCE_SALARY_OUTPUT in intent.outputs, (message, intent.outputs)
        decision = decide_chat_action(message, current_user)
        assert decision.action == FINANCE_COMPOUND_INTENT, (message, decision)


def wait_for_api() -> None:
    for _ in range(30):
        try:
            response = requests.get(f"{API_BASE_URL}/health", timeout=5)
            if response.status_code == 200:
                return
        except requests.RequestException:
            pass
        time.sleep(1)
    raise RuntimeError(f"API 不可用：{API_BASE_URL}")


def login(username: str, password: str) -> dict[str, Any]:
    response = requests.post(
        f"{API_BASE_URL}/auth/login",
        data={"username": username, "password": password},
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def post_json(token: str, path: str, payload: dict[str, Any], timeout: int = 60) -> dict[str, Any]:
    response = requests.post(
        f"{API_BASE_URL}{path}",
        headers={**auth_headers(token), "Content-Type": "application/json"},
        json=payload,
        timeout=timeout,
    )
    if response.status_code != 200:
        raise AssertionError(f"{path} expected 200, got {response.status_code}: {response.text[:1000]}")
    return response.json()


def post_sse(token: str, path: str, payload: dict[str, Any], timeout: int = 120) -> dict[str, Any]:
    response = requests.post(
        f"{API_BASE_URL}{path}",
        headers={**auth_headers(token), "Content-Type": "application/json"},
        json=payload,
        stream=True,
        timeout=timeout,
    )
    response.raise_for_status()
    done_payload: dict[str, Any] | None = None
    current_event = ""
    current_data = ""
    for raw_line in response.iter_lines(decode_unicode=True):
        line = raw_line or ""
        if line.startswith("event:"):
            current_event = line.split(":", 1)[1].strip()
        elif line.startswith("data:"):
            current_data = line.split(":", 1)[1].strip()
        elif line == "" and current_event:
            if current_event == "error":
                raise AssertionError(current_data)
            if current_event == "done":
                done_payload = json.loads(current_data)
                break
            current_event = ""
            current_data = ""
    assert done_payload, "stream did not emit done"
    return done_payload


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


if __name__ == "__main__":
    main()
