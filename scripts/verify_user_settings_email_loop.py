from __future__ import annotations

from contextlib import contextmanager
from email import policy
from email.parser import BytesParser
from io import BytesIO
import json
import os
import socket
import socketserver
import subprocess
import sys
import threading
import time
from pathlib import Path
from typing import Any

import openpyxl
import requests


ROOT = Path(__file__).resolve().parents[1]
API_BASE_URL = "http://127.0.0.1:8002"
FINANCE_EMAIL = "finance-loop-test@example.com"


class CapturedSMTPMessage:
    def __init__(self, mail_from: str, rcpt_to: list[str], data: bytes):
        self.mail_from = mail_from
        self.rcpt_to = rcpt_to
        self.data = data


class SMTPTestServer(socketserver.ThreadingTCPServer):
    allow_reuse_address = True

    def __init__(self, server_address):
        super().__init__(server_address, SMTPHandler)
        self.messages: list[CapturedSMTPMessage] = []


class SMTPHandler(socketserver.StreamRequestHandler):
    def handle(self) -> None:
        mail_from = ""
        rcpt_to: list[str] = []
        self._write("220 company-rag-test-smtp")
        while True:
            raw = self.rfile.readline()
            if not raw:
                return
            line = raw.decode("utf-8", errors="replace").rstrip("\r\n")
            command = line.upper()
            if command.startswith("EHLO") or command.startswith("HELO"):
                self._write("250-company-rag-test-smtp")
                self._write("250 OK")
            elif command.startswith("MAIL FROM:"):
                mail_from = line.split(":", 1)[1].strip("<> ")
                self._write("250 OK")
            elif command.startswith("RCPT TO:"):
                rcpt_to.append(line.split(":", 1)[1].strip("<> "))
                self._write("250 OK")
            elif command == "DATA":
                self._write("354 End data with <CR><LF>.<CR><LF>")
                chunks: list[bytes] = []
                while True:
                    data_line = self.rfile.readline()
                    if not data_line or data_line == b".\r\n":
                        break
                    chunks.append(data_line)
                self.server.messages.append(CapturedSMTPMessage(mail_from, rcpt_to, b"".join(chunks)))
                self._write("250 OK")
            elif command == "RSET":
                mail_from = ""
                rcpt_to = []
                self._write("250 OK")
            elif command == "QUIT":
                self._write("221 Bye")
                return
            else:
                self._write("250 OK")

    def _write(self, text: str) -> None:
        self.wfile.write(f"{text}\r\n".encode("utf-8"))


def main() -> None:
    smtp_server = SMTPTestServer(("127.0.0.1", 0))
    smtp_port = smtp_server.server_address[1]
    smtp_thread = threading.Thread(target=smtp_server.serve_forever, daemon=True)
    smtp_thread.start()

    use_docker = os.environ.get("VERIFY_USER_SETTINGS_EMAIL_DOCKER") == "1"
    context = run_docker_api(smtp_port) if use_docker else run_api(smtp_port)

    with context:
        finance_token = login("finance_demo", "Finance123456")
        customer_token = login("employee_demo", "Employee123456")

        settings = update_profile(finance_token, "财务测试用户", FINANCE_EMAIL)
        assert settings["email"] == FINANCE_EMAIL, settings
        assert settings["display_name"] == "财务测试用户", settings

        no_email_response = post_chat(
            finance_token,
            "把这个月所有员工的工资表发我",
        )
        assert no_email_response["intent"] == "finance_salary_export", no_email_response
        assert no_email_response["attachments"], no_email_response
        assert "没有要求发送邮箱" in no_email_response["answer"], no_email_response["answer"]
        time.sleep(0.3)
        assert len(smtp_server.messages) == 0, "未要求发邮箱时不应该发邮件"

        email_response = post_chat(
            finance_token,
            "把这个月所有员工的工资表发给我并发送到我的邮箱",
        )
        assert email_response["intent"] == "finance_salary_export", email_response
        assert "已按你的要求发送到邮箱" in email_response["answer"], email_response["answer"]
        assert email_response["attachments"], email_response

        wait_for(lambda: len(smtp_server.messages) == 1, "没有捕获到 SMTP 邮件")
        parsed = BytesParser(policy=policy.default).parsebytes(smtp_server.messages[0].data)
        assert parsed["To"] == FINANCE_EMAIL, parsed
        attachments = list(parsed.iter_attachments())
        assert len(attachments) == 1, f"expected one attachment, got {len(attachments)}"
        payload = attachments[0].get_payload(decode=True)
        assert payload and len(payload) > 1000, "工资表附件内容为空"
        workbook = openpyxl.load_workbook(BytesIO(payload))
        assert {"工资明细", "自动化摘要", "意图识别"}.issubset(set(workbook.sheetnames)), workbook.sheetnames

        forbidden = requests.post(
            f"{API_BASE_URL}/chat",
            headers=auth_headers(customer_token),
            json={
                "message": "把这个月所有员工的工资表发给我并发送到我的邮箱",
            },
            timeout=60,
        )
        assert forbidden.status_code == 403, forbidden.text[:500]

        files = requests.get(
            f"{API_BASE_URL}/files?search=finance_salary&date_range=30d&file_type=excel&limit=20",
            headers=auth_headers(finance_token),
            timeout=60,
        )
        assert files.status_code == 200, files.text[:500]
        assert files.json()["items"], files.json()

    smtp_server.shutdown()
    smtp_server.server_close()

    print(json.dumps({
        "ok": True,
        "settings_email": settings["email"],
        "no_email_chat_attachment_count": len(no_email_response["attachments"]),
        "email_chat_attachment_count": len(email_response["attachments"]),
        "captured_email_count": len(smtp_server.messages),
        "captured_attachment_sheets": workbook.sheetnames,
        "customer_service_forbidden": forbidden.status_code,
        "note": "real API, real ERP salary export, real local SMTP send; no production mock/stub",
    }, ensure_ascii=False, indent=2))


@contextmanager
def run_api(smtp_port: int):
    env = os.environ.copy()
    env.update({
        "DATABASE_URL": "postgresql://rag_user:rag_password@127.0.0.1:5433/rag_agent",
        "VECTOR_DATABASE_URL": "postgresql+psycopg://rag_user:rag_password@127.0.0.1:5433/rag_agent",
        "ERP_BASE_URL": "http://127.0.0.1:8080",
        "SMTP_HOST": "127.0.0.1",
        "SMTP_PORT": str(smtp_port),
        "SMTP_FROM_EMAIL": "ai-platform@example.com",
        "SMTP_USE_TLS": "false",
        "SMTP_USE_STARTTLS": "false",
        "SMTP_TIMEOUT_SECONDS": "5",
    })
    process = subprocess.Popen(
        [
            sys.executable,
            "-m",
            "uvicorn",
            "app.main:app",
            "--host",
            "127.0.0.1",
            "--port",
            "8002",
        ],
        cwd=ROOT,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
    )
    try:
        wait_for_api(process)
        yield
    finally:
        process.terminate()
        try:
            process.wait(timeout=8)
        except subprocess.TimeoutExpired:
            process.kill()
            process.wait(timeout=8)
        if process.stdout:
            output = process.stdout.read()
            if output:
                print(output[-2000:], file=sys.stderr)


@contextmanager
def run_docker_api(smtp_port: int):
    command = [
        "docker",
        "run",
        "--rm",
        "--name",
        "company-rag-api-email-verify",
        "--env-file",
        str(ROOT / ".env"),
        "-e",
        "DATABASE_URL=postgresql://rag_user:rag_password@host.docker.internal:5433/rag_agent",
        "-e",
        "VECTOR_DATABASE_URL=postgresql+psycopg://rag_user:rag_password@host.docker.internal:5433/rag_agent",
        "-e",
        "ERP_BASE_URL=http://host.docker.internal:8080",
        "-e",
        "SMTP_HOST=host.docker.internal",
        "-e",
        f"SMTP_PORT={smtp_port}",
        "-e",
        "SMTP_FROM_EMAIL=ai-platform@example.com",
        "-e",
        "SMTP_USE_TLS=false",
        "-e",
        "SMTP_USE_STARTTLS=false",
        "-e",
        "SMTP_TIMEOUT_SECONDS=5",
        "-p",
        "8002:8001",
        "company-rag-agent-api:latest",
    ]
    subprocess.run(
        ["docker", "rm", "-f", "company-rag-api-email-verify"],
        cwd=ROOT,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        check=False,
    )
    process = subprocess.Popen(
        command,
        cwd=ROOT,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
    )
    try:
        wait_for_docker_api(process)
        yield
    finally:
        subprocess.run(
            ["docker", "rm", "-f", "company-rag-api-email-verify"],
            cwd=ROOT,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
        try:
            process.wait(timeout=8)
        except subprocess.TimeoutExpired:
            process.kill()
            process.wait(timeout=8)
        if process.stdout:
            output = process.stdout.read()
            if output:
                print(output[-2000:], file=sys.stderr)


def wait_for_api(process: subprocess.Popen[str]) -> None:
    deadline = time.time() + 60
    while time.time() < deadline:
        if process.poll() is not None:
            output = process.stdout.read() if process.stdout else ""
            raise RuntimeError(f"API exited early: {output}")
        try:
            if requests.get(f"{API_BASE_URL}/health", timeout=2).status_code == 200:
                return
        except requests.RequestException:
            time.sleep(0.5)
    raise TimeoutError("API did not become healthy")


def wait_for_docker_api(process: subprocess.Popen[str]) -> None:
    deadline = time.time() + 90
    while time.time() < deadline:
        if process.poll() is not None:
            output = process.stdout.read() if process.stdout else ""
            raise RuntimeError(f"Docker API exited early: {output}")
        try:
            if requests.get(f"{API_BASE_URL}/health", timeout=2).status_code == 200:
                return
        except requests.RequestException:
            pass
        time.sleep(0.5)
    raise TimeoutError(f"API did not become healthy: {API_BASE_URL}")


def login(username: str, password: str) -> str:
    response = requests.post(
        f"{API_BASE_URL}/auth/login",
        data={"username": username, "password": password},
        timeout=60,
    )
    assert response.status_code == 200, response.text[:500]
    return response.json()["access_token"]


def update_profile(token: str, display_name: str, email: str) -> dict[str, Any]:
    response = requests.put(
        f"{API_BASE_URL}/settings/me/profile",
        headers=auth_headers(token),
        json={"display_name": display_name, "email": email},
        timeout=60,
    )
    assert response.status_code == 200, response.text[:500]
    return response.json()["item"]


def post_chat(token: str, message: str, thread_id: str | None = None) -> dict[str, Any]:
    payload: dict[str, Any] = {"message": message}
    if thread_id:
        payload["thread_id"] = thread_id

    response = requests.post(
        f"{API_BASE_URL}/chat",
        headers=auth_headers(token),
        json=payload,
        timeout=90,
    )
    assert response.status_code == 200, response.text[:500]
    return response.json()


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def wait_for(predicate, message: str) -> None:
    deadline = time.time() + 10
    while time.time() < deadline:
        if predicate():
            return
        time.sleep(0.2)
    raise AssertionError(message)


if __name__ == "__main__":
    main()
