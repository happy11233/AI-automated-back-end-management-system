from __future__ import annotations

from io import BytesIO
import json
import os
import time
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook


API_BASE_URL = os.getenv("VERIFY_API_BASE_URL", "http://127.0.0.1:8001").rstrip("/")

ACCOUNTS = {
    "admin": ("admin_demo", "Admin123456"),
    "finance": ("finance_demo", "Finance123456"),
    "customer_service": ("employee_demo", "Employee123456"),
}


def main() -> None:
    test_run_id = f"fin-rec-{int(time.time())}"
    tokens = {name: login(*account) for name, account in ACCOUNTS.items()}
    files = build_reconciliation_files(test_run_id)

    response = requests.post(
        f"{API_BASE_URL}/automation/finance/reconciliation",
        headers=auth_headers(tokens["finance"]),
        files=[
            (
                "files",
                (
                    filename,
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            )
            for filename, content in files.items()
        ],
        data={
            "instruction": f"{test_run_id} 生成订单利润表，标记亏损、低毛利、缺成本和未匹配费用。",
            "base_currency": "CNY",
        },
        timeout=180,
    )
    assert response.status_code == 200, response.text[:500]
    assert response.content[:2] == b"PK"

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    expected_sheets = {"对账摘要", "订单利润表", "异常账单", "字段识别", "源文件概览"}
    missing = expected_sheets.difference(workbook.sheetnames)
    assert not missing, f"missing sheets: {missing}; got={workbook.sheetnames}"

    profit_sheet = workbook["订单利润表"]
    anomaly_sheet = workbook["异常账单"]
    summary_sheet = workbook["对账摘要"]
    assert profit_sheet.max_row >= 4, "订单利润表应该至少包含 3 条订单明细"
    assert anomaly_sheet.max_row >= 2, "异常账单应该至少包含 1 条异常"
    assert find_cell_value(summary_sheet, "异常数量") is not None
    assert has_row_text(profit_sheet, "AMZ-US-REC-001")
    assert has_row_text(profit_sheet, "亏损")
    assert has_row_text(anomaly_sheet, "missing_purchase_cost") or has_row_text(anomaly_sheet, "negative_profit")

    forbidden = requests.post(
        f"{API_BASE_URL}/automation/finance/reconciliation",
        headers=auth_headers(tokens["customer_service"]),
        files=[
            (
                "files",
                (
                    "amazon_settlement.xlsx",
                    files["amazon_settlement.xlsx"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            )
        ],
        data={"instruction": "客服尝试越权对账", "base_currency": "CNY"},
        timeout=60,
    )
    assert forbidden.status_code == 403, forbidden.text

    flow_items = get_json(tokens["finance"], "/automation-flows")["items"]
    assert any(item["app_id"] == "finance-reconciliation" for item in flow_items), flow_items

    workflow_items = get_json(tokens["finance"], "/ai-workflows")["items"]
    assert any(item["id"] == "finance_reconciliation" for item in workflow_items), workflow_items

    run_items = get_json(tokens["admin"], "/run-records?run_type=finance_reconciliation&limit=20")["items"]
    assert run_items, "missing finance_reconciliation run record"

    raw_outputs = json.dumps([run_items], ensure_ascii=False)
    for sensitive_text in ["Bearer ", "api_secret", "password", "Authorization"]:
        assert sensitive_text not in raw_outputs, f"leaked {sensitive_text}"

    print(json.dumps({
        "ok": True,
        "download_bytes": len(response.content),
        "sheets": workbook.sheetnames,
        "profit_rows": profit_sheet.max_row - 1,
        "anomaly_rows": anomaly_sheet.max_row - 1,
        "latest_run_id": run_items[0]["id"],
        "note": "real API, real auth, real xlsx upload/download, real workbook inspection; no mock/stub/fake",
    }, ensure_ascii=False))


def build_reconciliation_files(test_run_id: str) -> dict[str, bytes]:
    return {
        "amazon_settlement.xlsx": workbook_bytes(
            "Amazon结算",
            [
                {"订单号": "AMZ-US-REC-001", "SKU": "BOTTLE-01", "数量": 2, "币种": "USD", "销售额": 100, "退款": 0, "平台手续费": 15},
                {"订单号": "AMZ-US-REC-002", "SKU": "BAG-02", "数量": 1, "币种": "USD", "销售额": 35, "退款": 0, "平台手续费": 7},
                {"订单号": "AMZ-US-REC-003", "SKU": "MISSING-COST", "数量": 1, "币种": "USD", "销售额": 20, "退款": 0, "平台手续费": 3},
            ],
        ),
        "logistics.xlsx": workbook_bytes(
            "物流账单",
            [
                {"订单号": "AMZ-US-REC-001", "物流费": 80, "币种": "CNY"},
                {"订单号": "AMZ-US-REC-002", "物流费": 30, "币种": "CNY"},
                {"订单号": "UNMATCHED-LOGISTICS", "物流费": 18, "币种": "CNY"},
            ],
        ),
        "purchase_cost.xlsx": workbook_bytes(
            "采购成本",
            [
                {"SKU": "BOTTLE-01", "采购成本": 120, "币种": "CNY"},
                {"SKU": "BAG-02", "采购成本": 250, "币种": "CNY"},
            ],
        ),
        "ads.xlsx": workbook_bytes(
            "广告费",
            [
                {"SKU": "BOTTLE-01", "广告费": 30, "币种": "CNY"},
                {"SKU": "BAG-02", "广告费": 50, "币种": "CNY"},
            ],
        ),
        "exchange_rate.xlsx": workbook_bytes(
            "汇率",
            [
                {"币种": "USD", "汇率": 7.2},
                {"币种": "CNY", "汇率": 1},
            ],
        ),
        "marker.xlsx": workbook_bytes(
            "测试标识",
            [{"测试ID": test_run_id, "说明": "运行记录应只保存摘要，不保存完整原文"}],
        ),
    }


def workbook_bytes(sheet_name: str, rows: list[dict[str, Any]]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name=sheet_name)
    return output.getvalue()


def login(username: str, password: str) -> str:
    response = requests.post(
        f"{API_BASE_URL}/auth/login",
        data={"username": username, "password": password},
        timeout=30,
    )
    assert response.status_code == 200, response.text
    return response.json()["access_token"]


def get_json(token: str, path: str) -> dict[str, Any]:
    response = requests.get(f"{API_BASE_URL}{path}", headers=auth_headers(token), timeout=60)
    assert response.status_code == 200, response.text
    return response.json()


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def has_row_text(sheet, text: str) -> bool:
    for row in sheet.iter_rows(values_only=True):
        if any(text in str(value) for value in row if value is not None):
            return True
    return False


def find_cell_value(sheet, key: str) -> Any:
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == key:
            return row[1] if len(row) > 1 else None
    return None


if __name__ == "__main__":
    main()
