from __future__ import annotations

from io import BytesIO
import base64
import json
import os

import requests
from openpyxl import load_workbook


API_BASE_URL = os.getenv("VERIFY_API_BASE_URL", "http://127.0.0.1:8001").rstrip("/")

ACCOUNTS = {
    "finance": ("finance_demo", "Finance123456"),
    "customer_service": ("employee_demo", "Employee123456"),
}


def main() -> None:
    tokens = {name: login(*account) for name, account in ACCOUNTS.items()}
    message = "把这个月所有员工的工资表发我"

    response = requests.post(
        f"{API_BASE_URL}/automation/finance/salary-export",
        headers={**auth_headers(tokens["finance"]), "Content-Type": "application/json"},
        json={"message": message},
        timeout=90,
    )
    assert response.status_code == 200, response.text[:500]
    assert response.content[:2] == b"PK"
    assert response.headers.get("x-automation-intent") == "finance_salary_export"
    assert response.headers.get("x-automation-employee-count") == "5"

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    expected_sheets = {"工资明细", "自动化摘要", "意图识别"}
    missing = expected_sheets.difference(workbook.sheetnames)
    assert not missing, f"missing sheets: {missing}; got={workbook.sheetnames}"

    detail_sheet = workbook["工资明细"]
    summary_sheet = workbook["自动化摘要"]
    assert detail_sheet.max_row == 6, f"expected 5 salary rows, got {detail_sheet.max_row - 1}"
    assert has_row_text(detail_sheet, "张晨")
    assert has_row_text(detail_sheet, "王静")
    assert find_cell_value(summary_sheet, "工资单数量") == 5
    assert float(find_cell_value(summary_sheet, "应发工资合计")) == 57400.0
    assert float(find_cell_value(summary_sheet, "实发工资合计")) == 47150.0

    chat_response = requests.post(
        f"{API_BASE_URL}/chat",
        headers={**auth_headers(tokens["finance"]), "Content-Type": "application/json"},
        json={"message": message},
        timeout=90,
    )
    assert chat_response.status_code == 200, chat_response.text[:500]
    chat_data = chat_response.json()
    assert chat_data["intent"] == "finance_salary_export", chat_data
    assert chat_data["attachments"], chat_data
    attachment = chat_data["attachments"][0]
    assert attachment["filename"].endswith(".xlsx"), attachment
    decoded = base64.b64decode(attachment["content_base64"])
    assert decoded[:2] == b"PK"

    forbidden = requests.post(
        f"{API_BASE_URL}/automation/finance/salary-export",
        headers={**auth_headers(tokens["customer_service"]), "Content-Type": "application/json"},
        json={"message": message},
        timeout=30,
    )
    assert forbidden.status_code == 403, forbidden.text

    forbidden_chat = requests.post(
        f"{API_BASE_URL}/chat",
        headers={**auth_headers(tokens["customer_service"]), "Content-Type": "application/json"},
        json={"message": message},
        timeout=60,
    )
    assert forbidden_chat.status_code == 403, forbidden_chat.text

    print(json.dumps({
        "ok": True,
        "endpoint_download_bytes": len(response.content),
        "chat_attachment_bytes": len(decoded),
        "sheets": workbook.sheetnames,
        "employee_count": detail_sheet.max_row - 1,
        "gross_total": find_cell_value(summary_sheet, "应发工资合计"),
        "net_total": find_cell_value(summary_sheet, "实发工资合计"),
        "customer_service_forbidden": forbidden.status_code,
        "customer_service_chat_forbidden": forbidden_chat.status_code,
        "note": "real API, real auth, real ERPNext Salary Slip query, real xlsx download/inspection; no mock/stub/fake",
    }, ensure_ascii=False))


def login(username: str, password: str) -> str:
    response = requests.post(
        f"{API_BASE_URL}/auth/login",
        data={"username": username, "password": password},
        timeout=30,
    )
    assert response.status_code == 200, response.text
    return response.json()["access_token"]


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def find_cell_value(sheet, label: str):
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == label:
            return row[1]
    return None


def has_row_text(sheet, text: str) -> bool:
    for row in sheet.iter_rows(values_only=True):
        if any(text in str(value) for value in row if value is not None):
            return True
    return False


if __name__ == "__main__":
    main()
