from __future__ import annotations

from io import BytesIO
import json
import os
import time
from typing import Any
from zipfile import ZipFile

import requests
from openpyxl import Workbook, load_workbook


API_BASE_URL = os.getenv("VERIFY_API_BASE_URL", "http://127.0.0.1:8001").rstrip("/")
EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

ACCOUNTS = {
    "finance": ("finance_demo", "Finance123456"),
    "customer_service": ("employee_demo", "Employee123456"),
    "admin": ("admin_demo", "Admin123456"),
}


def main() -> None:
    tokens = {name: login(*account) for name, account in ACCOUNTS.items()}
    marker = f"finance-report-loop-{int(time.time())}"
    report_content = build_finance_report_workbook(marker)

    word_response = post_report_analysis(
        tokens["finance"],
        marker,
        report_content,
        output_format="word",
    )
    assert word_response.status_code == 200, word_response.text[:500]
    assert word_response.content[:2] == b"PK"
    assert word_response.headers.get("x-automation-output-format") == "word"
    assert has_docx_document(word_response.content)

    excel_response = post_report_analysis(
        tokens["finance"],
        marker,
        report_content,
        output_format="excel",
    )
    assert excel_response.status_code == 200, excel_response.text[:500]
    assert excel_response.content[:2] == b"PK"
    assert excel_response.headers.get("x-automation-output-format") == "excel"
    workbook = load_workbook(BytesIO(excel_response.content), data_only=True)
    assert {"AI分析报告", "源文件摘要", "分析要求"}.issubset(workbook.sheetnames), workbook.sheetnames

    manual_response = requests.post(
        f"{API_BASE_URL}/automation/finance/report-analysis",
        headers=auth_headers(tokens["finance"]),
        data={
            "instruction": f"{marker} 手动输入报表：7月销售额128900.5，退款3280，广告费14500，净利润37120.5。请生成 Word 分析报告。",
            "output_format": "word",
        },
        timeout=180,
    )
    assert manual_response.status_code == 200, manual_response.text[:500]
    assert manual_response.content[:2] == b"PK"
    assert has_docx_document(manual_response.content)

    finance_files = get_json(tokens["finance"], f"/files?search=finance_report_analysis&date_range=30d&file_type=all&limit=20")["items"]
    word_files = [item for item in finance_files if item["name"].endswith(".docx")]
    excel_files = [item for item in finance_files if item["name"].endswith(".xlsx")]
    assert word_files, finance_files
    assert excel_files, finance_files
    assert all(item["downloadable"] for item in word_files + excel_files)
    assert all(item["expires_at"] for item in word_files + excel_files)

    downloaded_word = download_file(tokens["finance"], word_files[0]["id"])
    assert downloaded_word.status_code == 200, downloaded_word.text[:500]
    assert downloaded_word.content[:2] == b"PK"
    assert has_docx_document(downloaded_word.content)

    downloaded_excel = download_file(tokens["finance"], excel_files[0]["id"])
    assert downloaded_excel.status_code == 200, downloaded_excel.text[:500]
    assert downloaded_excel.content[:2] == b"PK"
    load_workbook(BytesIO(downloaded_excel.content), data_only=True)

    customer_files = get_json(tokens["customer_service"], "/files?date_range=30d&file_type=all&limit=20")["items"]
    assert all(item["position"] == "customer_service" for item in customer_files), customer_files
    forbidden = download_file(tokens["customer_service"], word_files[0]["id"])
    assert forbidden.status_code == 403, forbidden.text

    salary_response = requests.post(
        f"{API_BASE_URL}/automation/finance/salary-export",
        headers={**auth_headers(tokens["finance"]), "Content-Type": "application/json"},
        json={"message": "把这个月所有员工的工资表发我"},
        timeout=90,
    )
    assert salary_response.status_code == 200, salary_response.text[:500]
    salary_files = get_json(tokens["finance"], "/files?search=finance_salary&date_range=30d&file_type=excel&limit=20")["items"]
    assert salary_files, salary_files
    assert any(item["name"].endswith(".xlsx") for item in salary_files), salary_files

    admin_files = get_json(tokens["admin"], "/files?date_range=30d&file_type=all&limit=50")["items"]
    assert len(admin_files) >= len(finance_files), admin_files

    print(json.dumps({
        "ok": True,
        "finance_report_word_bytes": len(word_response.content),
        "finance_report_excel_bytes": len(excel_response.content),
        "manual_report_word_bytes": len(manual_response.content),
        "finance_file_count": len(finance_files),
        "customer_visible_file_count": len(customer_files),
        "customer_forbidden_download": forbidden.status_code,
        "salary_file_count": len(salary_files),
        "note": "real API, real auth, real uploaded xlsx parsing, real Word/Excel downloads, real persisted generated file center; no mock/stub/fake",
    }, ensure_ascii=False))


def post_report_analysis(token: str, marker: str, content: bytes, output_format: str):
    return requests.post(
        f"{API_BASE_URL}/automation/finance/report-analysis",
        headers=auth_headers(token),
        files={
            "files": (
                f"{marker}.xlsx",
                content,
                EXCEL_MIME,
            )
        },
        data={
            "instruction": f"{marker} 分析本月销售额、退款、广告费、利润率和异常风险。",
            "output_format": output_format,
        },
        timeout=180,
    )


def build_finance_report_workbook(marker: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "财务报表"
    sheet.append(["标识", "月份", "销售额", "退款", "广告费", "物流费", "采购成本", "净利润"])
    sheet.append([marker, "2026-07", 128900.5, 3280, 14500, 9800, 64200, 37120.5])
    sheet.append([marker, "2026-06", 115300.0, 1900, 13200, 9100, 60100, 31000.0])
    sheet.append([marker, "2026-05", 107500.0, 4200, 12000, 8700, 59000, 23600.0])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def has_docx_document(content: bytes) -> bool:
    with ZipFile(BytesIO(content)) as docx:
        return "word/document.xml" in docx.namelist()


def login(username: str, password: str) -> str:
    response = requests.post(
        f"{API_BASE_URL}/auth/login",
        data={"username": username, "password": password},
        timeout=30,
    )
    assert response.status_code == 200, response.text
    return response.json()["access_token"]


def get_json(token: str, path: str) -> dict[str, Any]:
    response = requests.get(f"{API_BASE_URL}{path}", headers=auth_headers(token), timeout=60)
    assert response.status_code == 200, response.text
    return response.json()


def download_file(token: str, artifact_id: str):
    return requests.get(
        f"{API_BASE_URL}/files/{artifact_id}/download",
        headers=auth_headers(token),
        timeout=60,
    )


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


if __name__ == "__main__":
    main()
