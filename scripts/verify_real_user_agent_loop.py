from __future__ import annotations

import base64
import json
import os
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

import requests
from openpyxl import Workbook, load_workbook


API_BASE_URL = os.getenv("VERIFY_API_BASE_URL", "http://127.0.0.1:8001").rstrip("/")
REPORT_DIR = Path(os.getenv("REAL_USER_AGENT_REPORT_DIR", "/tmp/company-rag-real-user-agent-loop"))
REPORT_JSON = REPORT_DIR / "report.json"
REPORT_MD = REPORT_DIR / "report.md"
EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

ACCOUNTS = {
    "admin": ("admin_demo", "Admin123456", "管理员"),
    "operations": ("operations_demo", "Operations123456", "运营员工"),
    "customer_service": ("employee_demo", "Employee123456", "客服员工"),
    "finance": ("finance_demo", "Finance123456", "财务员工"),
}

SENSITIVE_TEXTS = [
    "Authorization",
    "Bearer ",
    "access_token",
    "api_secret",
    "api_key",
    "password",
    "JWT_SECRET",
    "DATABASE_URL",
]


@dataclass
class AgentResult:
    agent: str
    position: str
    account: str
    scenario: str
    ok: bool
    automation_elapsed_seconds: float
    traditional_minutes: float
    estimated_saved_minutes: float
    saved_ratio: float
    manual_steps_replaced: list[str] = field(default_factory=list)
    automation_evidence: dict[str, Any] = field(default_factory=dict)
    output_summary: dict[str, Any] = field(default_factory=dict)


def main() -> None:
    started_at = time.perf_counter()
    tokens: dict[str, str] = {}
    users: dict[str, dict[str, Any]] = {}

    for key, (username, password, _) in ACCOUNTS.items():
        payload = login(username, password)
        tokens[key] = payload["access_token"]
        users[key] = {
            "username": payload["username"],
            "role": payload["role"],
            "position": payload.get("position"),
            "department": payload.get("department"),
            "capabilities": payload.get("capabilities") or [],
            "erp_scopes": payload.get("erp_scopes") or [],
            "allowed_ai_app_count": len(payload.get("allowed_ai_app_ids") or []),
        }

    report: dict[str, Any] = {
        "ok": True,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "api_base_url": API_BASE_URL,
        "mode": "real_position_user_agent_loop",
        "baseline_source": "脚本内置传统人工耗时基准，用于回归测试对比；可按企业真实统计调整。",
        "real_path_note": "使用真实账号、真实鉴权、真实 API、真实 ERP/RAG/LLM/Excel 下载链路；不向生产代码写入 mock 或 stub。",
        "accounts": users,
        "results": [],
    }

    cases: list[tuple[str, str, str, str, float, list[str], Callable[[], dict[str, Any]]]] = [
        (
            "岗位权限 Agent",
            "all",
            "admin_demo",
            "按真实岗位登录后检查应用、工作流和 ERP 权限范围",
            5,
            ["人工逐个登录岗位账号", "人工核对菜单和权限列表"],
            lambda: check_position_visibility(tokens),
        ),
        (
            "运营 Agent",
            "operations",
            "operations_demo",
            "运营用自然语言生成 Amazon Listing 上架草稿",
            35,
            ["整理产品卖点", "写标题和五点描述", "整理后台关键词", "撰写优化备注"],
            lambda: run_operations_listing_agent(tokens),
        ),
        (
            "客服 Agent",
            "customer_service",
            "employee_demo",
            "外部客户物流消息进入收件箱后自动识别、查 ERP/RAG 并生成回复",
            8,
            ["复制订单号", "登录 ERP 查物流", "翻售后规则", "写英文回复草稿"],
            lambda: run_customer_logistics_agent(tokens),
        ),
        (
            "客服 Agent",
            "customer_service",
            "employee_demo",
            "退款和差评高风险消息自动转人工",
            12,
            ["判断投诉风险", "整理订单上下文", "写升级原因", "创建人工处理记录"],
            lambda: run_customer_high_risk_agent(tokens),
        ),
        (
            "财务 Agent",
            "finance",
            "finance_demo",
            "财务用模糊问题让 AI 自动导出本月员工工资 Excel",
            45,
            ["理解财务口语请求", "登录 ERP 筛选工资单", "复制员工薪资明细", "汇总应发和实发工资", "整理 Excel"],
            lambda: run_finance_salary_chat_agent(tokens),
        ),
        (
            "财务 Agent",
            "finance",
            "finance_demo",
            "财务上传 Excel 并选择权限内 ERP 表生成新表",
            40,
            ["选择结算表", "查询销售发票", "查询收付款单", "跨表核对金额", "生成新工作簿"],
            lambda: run_finance_excel_agent(tokens),
        ),
        (
            "财务 Agent",
            "finance",
            "finance_demo",
            "财务多表对账生成订单利润表和异常账单",
            90,
            ["合并平台结算表", "匹配物流费用", "匹配采购成本", "核算利润", "标记异常账单"],
            lambda: run_finance_reconciliation_agent(tokens),
        ),
        (
            "安全 Agent",
            "all",
            "multi_account",
            "跨岗位越权访问必须被拦截",
            15,
            ["人工切换账号尝试访问敏感功能", "人工记录越权状态码"],
            lambda: run_permission_guard_agent(tokens),
        ),
    ]

    for agent, position, account, scenario, traditional_minutes, manual_steps, runner in cases:
        result = execute_case(
            agent=agent,
            position=position,
            account=account,
            scenario=scenario,
            traditional_minutes=traditional_minutes,
            manual_steps_replaced=manual_steps,
            runner=runner,
        )
        report["results"].append(asdict(result))
        print(json.dumps({
            "ok": True,
            "agent": agent,
            "scenario": scenario,
            "elapsed_seconds": result.automation_elapsed_seconds,
            "estimated_saved_minutes": result.estimated_saved_minutes,
            "summary": result.output_summary,
        }, ensure_ascii=False), flush=True)

    report["total_automation_elapsed_seconds"] = round(time.perf_counter() - started_at, 2)
    report["total_traditional_minutes"] = round(
        sum(item["traditional_minutes"] for item in report["results"]),
        2,
    )
    report["total_estimated_saved_minutes"] = round(
        sum(item["estimated_saved_minutes"] for item in report["results"]),
        2,
    )

    assert_no_sensitive_leaks(report)
    write_reports(report)
    print(json.dumps({
        "ok": True,
        "report_json": str(REPORT_JSON),
        "report_markdown": str(REPORT_MD),
        "total_automation_elapsed_seconds": report["total_automation_elapsed_seconds"],
        "total_estimated_saved_minutes": report["total_estimated_saved_minutes"],
        "note": "真实岗位用户 Agent Loop 已完成；无 mock/stub/fake 写入业务代码。",
    }, ensure_ascii=False), flush=True)


def execute_case(
    *,
    agent: str,
    position: str,
    account: str,
    scenario: str,
    traditional_minutes: float,
    manual_steps_replaced: list[str],
    runner: Callable[[], dict[str, Any]],
) -> AgentResult:
    started_at = time.perf_counter()
    try:
        payload = runner()
    except Exception as error:
        elapsed = round(time.perf_counter() - started_at, 2)
        partial = {
            "agent": agent,
            "position": position,
            "account": account,
            "scenario": scenario,
            "ok": False,
            "automation_elapsed_seconds": elapsed,
            "error": str(error),
        }
        REPORT_DIR.mkdir(parents=True, exist_ok=True)
        (REPORT_DIR / "failed_case.json").write_text(json.dumps(partial, ensure_ascii=False, indent=2), encoding="utf-8")
        raise AssertionError(f"{agent} 场景失败：{scenario}；原因：{error}") from error

    elapsed_seconds = round(time.perf_counter() - started_at, 2)
    saved_minutes = round(max(traditional_minutes - elapsed_seconds / 60, 0), 2)
    saved_ratio = round(saved_minutes / traditional_minutes, 4) if traditional_minutes else 0
    return AgentResult(
        agent=agent,
        position=position,
        account=account,
        scenario=scenario,
        ok=True,
        automation_elapsed_seconds=elapsed_seconds,
        traditional_minutes=traditional_minutes,
        estimated_saved_minutes=saved_minutes,
        saved_ratio=saved_ratio,
        manual_steps_replaced=manual_steps_replaced,
        automation_evidence=payload.get("automation_evidence") or {},
        output_summary=payload.get("output_summary") or {},
    )


def check_position_visibility(tokens: dict[str, str]) -> dict[str, Any]:
    expected = {
        "operations": {"operations_listing_launch", "operations_competitor_analysis"},
        "customer_service": {
            "customer_service_refund_reply",
            "customer_service_logistics_reply",
            "customer_service_message_loop",
        },
        "finance": {
            "finance_report_analysis",
            "finance_salary_summary",
            "finance_excel_settlement",
            "finance_reconciliation",
        },
    }
    visible: dict[str, Any] = {}
    for position, expected_ids in expected.items():
        workflows = get_json(tokens[position], "/ai-workflows")["items"]
        workflow_ids = {item["id"] for item in workflows}
        missing = expected_ids.difference(workflow_ids)
        assert not missing, f"{position} 缺少工作流：{sorted(missing)}"
        assert all(item["position"] == position for item in workflows), workflows
        visible[position] = {
            "workflow_count": len(workflows),
            "workflow_ids": sorted(workflow_ids),
            "erp_scopes": get_json(tokens[position], "/auth/me")["erp_scopes"],
        }

    admin_workflows = get_json(tokens["admin"], "/ai-workflows")["items"]
    assert len(admin_workflows) >= sum(len(items) for items in expected.values()), admin_workflows
    return {
        "automation_evidence": {
            "position_workflows_are_isolated": True,
            "admin_can_view_all_positions": True,
        },
        "output_summary": visible,
    }


def run_operations_listing_agent(tokens: dict[str, str]) -> dict[str, Any]:
    payload = post_json(
        tokens["operations"],
        "/ai-workflows/operations_listing_launch/run",
        {
            "input_text": (
                "新品 SKU BOTTLE-01，32oz 不锈钢保温杯，防漏盖，双层真空保温，"
                "可替换吸管，目标站点 Amazon US，受众为通勤和户外用户。"
                "竞品差评集中在漏水、杯身掉漆、关键词覆盖不足。请生成可上架 Listing。"
            )
        },
        timeout=240,
    )
    assert payload["status"] == "succeeded", payload
    assert payload["run_id"], payload
    answer = payload["answer"]
    assert len(answer.strip()) >= 120, answer
    assert_contains_any(answer, ["Title", "Bullet", "Search Terms", "关键词", "五点"])
    step_names = [step["step_name"] for step in payload["steps"]]
    assert {"trigger_validate", "ai_generate_decision", "write_run_record"}.issubset(step_names), step_names

    run_detail = get_json(tokens["admin"], f"/run-records/{payload['run_id']}")
    assert run_detail["run"]["run_type"] == "ai_workflow", run_detail["run"]
    return {
        "automation_evidence": {
            "workflow_id": payload["workflow"]["id"],
            "run_id": payload["run_id"],
            "step_names": step_names,
            "run_record_status": run_detail["run"]["status"],
        },
        "output_summary": {
            "answer_chars": len(answer),
            "answer_preview": compact_text(answer, 220),
        },
    }


def run_customer_logistics_agent(tokens: dict[str, str]) -> dict[str, Any]:
    external_id = f"real-user-agent-logistics-{int(time.time())}"
    payload = post_json(
        tokens["customer_service"],
        "/customer-service/webhooks/messages",
        {
            "channel": "email",
            "external_id": external_id,
            "buyer_name": "John Buyer",
            "buyer_language": "English",
            "marketplace": "Amazon US",
            "order_no": "AMZ-US-001",
            "subject": "Where is my order?",
            "message": "Where is my order? My order number is AMZ-US-001. Please reply in English.",
            "auto_process": True,
        },
        timeout=240,
    )
    assert payload["processed"] is True, payload
    item = payload["item"]
    assert item["intent"] in {"logistics", "general_question"}, item
    assert item["risk_level"] == "low", item
    assert item["status"] in {"auto_reply_ready", "drafted"}, item
    assert item["reply_draft"], item
    assert payload["run_id"], payload
    step_names = [step["step_name"] for step in payload["steps"]]
    assert {"classify_intent_and_risk", "erp_permission_query", "rag_policy_lookup", "generate_reply_draft"}.issubset(step_names), step_names

    detail = get_json(tokens["customer_service"], f"/customer-service/messages/{item['id']}")
    assert len(detail["events"]) >= 4, detail["events"]
    run_detail = get_json(tokens["admin"], f"/run-records/{payload['run_id']}")
    assert run_detail["run"]["run_type"] == "customer_service_automation", run_detail["run"]
    return {
        "automation_evidence": {
            "message_id": item["id"],
            "run_id": payload["run_id"],
            "intent": item["intent"],
            "risk_level": item["risk_level"],
            "automation_decision": item["automation_decision"],
            "erp_reference_count": len(item.get("erp_references") or []),
            "citation_count": len(item.get("citations") or []),
            "event_count": len(detail["events"]),
            "step_names": step_names,
        },
        "output_summary": {
            "status": item["status"],
            "reply_preview": compact_text(item["reply_draft"], 220),
        },
    }


def run_customer_high_risk_agent(tokens: dict[str, str]) -> dict[str, Any]:
    detail = post_json(
        tokens["customer_service"],
        "/customer-service/messages",
        {
            "channel": "amazon",
            "buyer_name": "Mary Buyer",
            "buyer_language": "English",
            "marketplace": "Amazon US",
            "order_no": "AMZ-US-001",
            "subject": "Refund complaint",
            "message": "I want a refund now, otherwise I will leave a bad review.",
        },
        timeout=60,
    )
    item_id = detail["item"]["id"]
    payload = post_json(
        tokens["customer_service"],
        f"/customer-service/messages/{item_id}/process",
        {},
        timeout=240,
    )
    item = payload["item"]
    assert item["risk_level"] == "high", item
    assert item["status"] == "human_handoff", item
    assert item["approval_id"], item
    assert item["handoff_reason"], item
    return {
        "automation_evidence": {
            "message_id": item["id"],
            "run_id": payload["run_id"],
            "intent": item["intent"],
            "risk_level": item["risk_level"],
            "automation_decision": item["automation_decision"],
            "approval_id": item["approval_id"],
            "step_names": [step["step_name"] for step in payload["steps"]],
        },
        "output_summary": {
            "status": item["status"],
            "handoff_reason": item["handoff_reason"],
            "reply_preview": compact_text(item["reply_draft"] or "", 180),
        },
    }


def run_finance_salary_chat_agent(tokens: dict[str, str]) -> dict[str, Any]:
    message = "把这个月所有员工的工资表发我"
    payload = post_json(
        tokens["finance"],
        "/chat",
        {"message": message},
        timeout=180,
    )
    assert payload["intent"] == "finance_salary_export", payload
    assert payload["attachments"], payload
    attachment = payload["attachments"][0]
    assert attachment["filename"].endswith(".xlsx"), attachment
    content = base64.b64decode(attachment["content_base64"])
    assert content[:2] == b"PK", attachment
    workbook = load_workbook(BytesIO(content), data_only=True)
    expected_sheets = {"工资明细", "自动化摘要", "意图识别"}
    missing = expected_sheets.difference(workbook.sheetnames)
    assert not missing, f"工资 Excel 缺少 sheet：{missing}；got={workbook.sheetnames}"
    employee_count = workbook["工资明细"].max_row - 1
    assert employee_count > 0, "工资明细没有员工数据"
    metadata = attachment.get("metadata") or {}
    return {
        "automation_evidence": {
            "intent": payload["intent"],
            "thread_id": payload["thread_id"],
            "erp_resource": "Salary Slip",
            "filename": attachment["filename"],
            "employee_count": employee_count,
            "gross_pay_total": metadata.get("gross_pay_total"),
            "net_pay_total": metadata.get("net_pay_total"),
            "workbook_sheets": workbook.sheetnames,
        },
        "output_summary": {
            "answer_preview": compact_text(payload["answer"], 220),
            "download_bytes": len(content),
        },
    }


def run_finance_excel_agent(tokens: dict[str, str]) -> dict[str, Any]:
    content = workbook_bytes(
        "订单金额",
        [
            {"订单号": f"AMZ-US-XLS-{int(time.time())}-001", "店铺": "US Store", "SKU": "BOTTLE-01", "销售额": 120.5, "币种": "USD"},
            {"订单号": f"AMZ-DE-XLS-{int(time.time())}-002", "店铺": "DE Store", "SKU": "BAG-02", "销售额": 88.0, "币种": "EUR"},
            {"订单号": f"AMZ-JP-XLS-{int(time.time())}-003", "店铺": "JP Store", "SKU": "MUG-03", "销售额": -18.0, "币种": "JPY"},
        ],
    )
    response = requests.post(
        f"{API_BASE_URL}/automation/finance/excel-transform",
        headers=auth_headers(tokens["finance"]),
        files={"file": ("finance_agent_orders.xlsx", content, EXCEL_MIME)},
        data={
            "instruction": "按站点和 SKU 汇总销售额，并结合销售发票和收付款单检查异常金额。",
            "erp_resources": json.dumps(["Sales Invoice", "Payment Entry"], ensure_ascii=False),
        },
        timeout=180,
    )
    assert_status(response)
    assert response.content[:2] == b"PK", response.headers
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    expected_sheets = {"处理摘要", "数值汇总", "AI建议", "ERP数据摘要"}
    missing = expected_sheets.difference(workbook.sheetnames)
    assert not missing, f"财务 Excel 缺少 sheet：{missing}；got={workbook.sheetnames}"
    assert any(name.startswith("ERP_Sales Invoice") for name in workbook.sheetnames), workbook.sheetnames
    assert any(name.startswith("ERP_Payment Entry") for name in workbook.sheetnames), workbook.sheetnames
    latest = latest_run_record(tokens["admin"], "finance_excel_transform")
    return {
        "automation_evidence": {
            "selected_erp_resources": ["Sales Invoice", "Payment Entry"],
            "latest_run_id": latest.get("id"),
            "workbook_sheets": workbook.sheetnames,
        },
        "output_summary": {
            "download_bytes": len(response.content),
            "erp_summary_rows": workbook["ERP数据摘要"].max_row,
        },
    }


def run_finance_reconciliation_agent(tokens: dict[str, str]) -> dict[str, Any]:
    test_run_id = f"real-user-agent-rec-{int(time.time())}"
    files = build_reconciliation_files(test_run_id)
    response = requests.post(
        f"{API_BASE_URL}/automation/finance/reconciliation",
        headers=auth_headers(tokens["finance"]),
        files=[
            ("files", (filename, content, EXCEL_MIME))
            for filename, content in files.items()
        ],
        data={
            "instruction": f"{test_run_id} 自动生成订单利润表，标记亏损、低毛利、缺成本和未匹配费用。",
            "base_currency": "CNY",
        },
        timeout=240,
    )
    assert_status(response)
    assert response.content[:2] == b"PK", response.headers
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    expected_sheets = {"对账摘要", "订单利润表", "异常账单", "字段识别", "源文件概览"}
    missing = expected_sheets.difference(workbook.sheetnames)
    assert not missing, f"财务对账缺少 sheet：{missing}；got={workbook.sheetnames}"
    profit_rows = workbook["订单利润表"].max_row - 1
    anomaly_rows = workbook["异常账单"].max_row - 1
    assert profit_rows >= 3, "订单利润表应该至少包含 3 条订单"
    assert anomaly_rows >= 1, "异常账单应该至少包含 1 条异常"
    latest = latest_run_record(tokens["admin"], "finance_reconciliation")
    return {
        "automation_evidence": {
            "input_file_count": len(files),
            "latest_run_id": latest.get("id"),
            "workbook_sheets": workbook.sheetnames,
            "profit_rows": profit_rows,
            "anomaly_rows": anomaly_rows,
        },
        "output_summary": {
            "download_bytes": len(response.content),
            "summary_rows": workbook["对账摘要"].max_row,
        },
    }


def run_permission_guard_agent(tokens: dict[str, str]) -> dict[str, Any]:
    operations_cross_run = requests.post(
        f"{API_BASE_URL}/ai-workflows/finance_salary_summary/run",
        headers={**auth_headers(tokens["operations"]), "Content-Type": "application/json"},
        json={"input_text": "尝试查询本月员工工资"},
        timeout=30,
    )
    finance_customer_create = requests.post(
        f"{API_BASE_URL}/customer-service/messages",
        headers={**auth_headers(tokens["finance"]), "Content-Type": "application/json"},
        json={"channel": "manual", "message": "Where is my order?"},
        timeout=30,
    )
    customer_salary = requests.post(
        f"{API_BASE_URL}/automation/finance/salary-export",
        headers={**auth_headers(tokens["customer_service"]), "Content-Type": "application/json"},
        json={"message": "把这个月所有员工的工资表发我"},
        timeout=30,
    )
    finance_issue_resource = requests.post(
        f"{API_BASE_URL}/automation/finance/excel-transform",
        headers=auth_headers(tokens["finance"]),
        files={"file": ("finance_agent_orders.xlsx", workbook_bytes("订单金额", [{"订单号": "AMZ-US-XLS-403", "销售额": 12}]), EXCEL_MIME)},
        data={
            "instruction": "尝试选择客服售后工单表。",
            "erp_resources": json.dumps(["Issue"], ensure_ascii=False),
        },
        timeout=60,
    )
    assert operations_cross_run.status_code == 403, operations_cross_run.text
    assert finance_customer_create.status_code == 403, finance_customer_create.text
    assert customer_salary.status_code == 403, customer_salary.text
    assert finance_issue_resource.status_code == 403, finance_issue_resource.text
    return {
        "automation_evidence": {
            "operations_run_finance_workflow": operations_cross_run.status_code,
            "finance_create_customer_message": finance_customer_create.status_code,
            "customer_service_salary_export": customer_salary.status_code,
            "finance_select_customer_erp_resource": finance_issue_resource.status_code,
        },
        "output_summary": {
            "blocked_cases": 4,
            "expected_status": 403,
        },
    }


def login(username: str, password: str) -> dict[str, Any]:
    response = requests.post(
        f"{API_BASE_URL}/auth/login",
        data={"username": username, "password": password},
        timeout=30,
    )
    assert_status(response)
    payload = response.json()
    assert payload.get("access_token"), payload
    return payload


def get_json(token: str, path: str, timeout: int = 60) -> dict[str, Any]:
    response = requests.get(f"{API_BASE_URL}{path}", headers=auth_headers(token), timeout=timeout)
    assert_status(response)
    return response.json()


def post_json(token: str, path: str, payload: dict[str, Any], timeout: int = 60) -> dict[str, Any]:
    response = requests.post(
        f"{API_BASE_URL}{path}",
        headers={**auth_headers(token), "Content-Type": "application/json"},
        json=payload,
        timeout=timeout,
    )
    assert_status(response)
    return response.json()


def latest_run_record(token: str, run_type: str) -> dict[str, Any]:
    payload = get_json(token, f"/run-records?run_type={run_type}&limit=10")
    items = payload.get("items") or []
    assert items, f"没有找到 {run_type} 运行记录"
    return items[0]


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def assert_status(response: requests.Response, expected: int = 200) -> None:
    if response.status_code != expected:
        raise AssertionError(f"HTTP {response.status_code}, expected {expected}: {response.text[:800]}")


def assert_contains_any(text: str, needles: list[str]) -> None:
    lowered = text.lower()
    if not any(needle.lower() in lowered for needle in needles):
        raise AssertionError(f"输出未包含预期关键词 {needles}: {text[:500]}")


def workbook_bytes(sheet_name: str, rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]
    if rows:
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_reconciliation_files(test_run_id: str) -> dict[str, bytes]:
    return {
        "amazon_settlement.xlsx": workbook_bytes(
            "Amazon结算",
            [
                {"订单号": "AMZ-US-REC-001", "SKU": "BOTTLE-01", "数量": 2, "币种": "USD", "销售额": 100, "退款": 0, "平台手续费": 15},
                {"订单号": "AMZ-US-REC-002", "SKU": "BAG-02", "数量": 1, "币种": "USD", "销售额": 35, "退款": 0, "平台手续费": 7},
                {"订单号": "AMZ-US-REC-003", "SKU": "MISSING-COST", "数量": 1, "币种": "USD", "销售额": 20, "退款": 0, "平台手续费": 3},
            ],
        ),
        "logistics.xlsx": workbook_bytes(
            "物流账单",
            [
                {"订单号": "AMZ-US-REC-001", "物流费": 80, "币种": "CNY"},
                {"订单号": "AMZ-US-REC-002", "物流费": 30, "币种": "CNY"},
                {"订单号": "UNMATCHED-LOGISTICS", "物流费": 18, "币种": "CNY"},
            ],
        ),
        "purchase_cost.xlsx": workbook_bytes(
            "采购成本",
            [
                {"SKU": "BOTTLE-01", "采购成本": 120, "币种": "CNY"},
                {"SKU": "BAG-02", "采购成本": 250, "币种": "CNY"},
            ],
        ),
        "ads.xlsx": workbook_bytes(
            "广告费",
            [
                {"SKU": "BOTTLE-01", "广告费": 30, "币种": "CNY"},
                {"SKU": "BAG-02", "广告费": 50, "币种": "CNY"},
            ],
        ),
        "exchange_rate.xlsx": workbook_bytes(
            "汇率",
            [
                {"币种": "USD", "汇率": 7.2},
                {"币种": "CNY", "汇率": 1},
            ],
        ),
        "marker.xlsx": workbook_bytes(
            "测试标识",
            [{"测试ID": test_run_id, "说明": "真实用户 Agent Loop 上传的临时测试文件"}],
        ),
    }


def compact_text(value: str, limit: int) -> str:
    text = " ".join((value or "").split())
    return text if len(text) <= limit else f"{text[:limit - 3]}..."


def assert_no_sensitive_leaks(report: dict[str, Any]) -> None:
    raw = json.dumps(report, ensure_ascii=False)
    for sensitive_text in SENSITIVE_TEXTS:
        assert sensitive_text not in raw, f"报告中泄露敏感字段：{sensitive_text}"


def write_reports(report: dict[str, Any]) -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    REPORT_MD.write_text(build_markdown_report(report), encoding="utf-8")


def build_markdown_report(report: dict[str, Any]) -> str:
    lines = [
        "# 真实岗位用户 Agent Loop 报告",
        "",
        f"- 生成时间：{report['generated_at']}",
        f"- API：{report['api_base_url']}",
        f"- 说明：{report['real_path_note']}",
        f"- 传统耗时基准：{report['baseline_source']}",
        "",
        "| Agent | 岗位 | 场景 | 自动耗时 | 传统估算 | 预计节省 | 结果摘要 |",
        "| --- | --- | --- | ---: | ---: | ---: | --- |",
    ]
    for item in report["results"]:
        summary = compact_text(json.dumps(item["output_summary"], ensure_ascii=False), 120)
        lines.append(
            "| {agent} | {position} | {scenario} | {elapsed:.2f}s | {traditional:.1f}min | {saved:.2f}min | {summary} |".format(
                agent=item["agent"],
                position=item["position"],
                scenario=item["scenario"],
                elapsed=item["automation_elapsed_seconds"],
                traditional=item["traditional_minutes"],
                saved=item["estimated_saved_minutes"],
                summary=summary.replace("|", "/"),
            )
        )
    lines.extend([
        "",
        f"- 自动化总耗时：{report['total_automation_elapsed_seconds']:.2f}s",
        f"- 传统人工总估算：{report['total_traditional_minutes']:.2f}min",
        f"- 预计节省：{report['total_estimated_saved_minutes']:.2f}min",
        "",
    ])
    return "\n".join(lines)


if __name__ == "__main__":
    main()
