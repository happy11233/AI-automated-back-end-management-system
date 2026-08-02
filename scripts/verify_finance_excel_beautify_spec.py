from __future__ import annotations

import ast
from io import BytesIO
from pathlib import Path
import re
import sys
import types

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]


def load_request_matcher():
    source = (ROOT / "app/main.py").read_text(encoding="utf-8")
    module = ast.parse(source)
    function = next(
        node
        for node in module.body
        if isinstance(node, ast.FunctionDef)
        and node.name == "_looks_like_existing_finance_excel_transform_request"
    )
    namespace: dict[str, object] = {"re": re}
    exec(compile(ast.Module(body=[function], type_ignores=[]), "app/main.py", "exec"), namespace)
    return namespace[function.name]


def main() -> None:
    fake_llm = types.ModuleType("app.llm")
    fake_llm.chat = lambda _prompt: "固定测试建议：请复核金额汇总。"
    sys.modules["app.llm"] = fake_llm

    import app.services.finance_excel_service as finance_excel_service

    readable_field_name = finance_excel_service._readable_field_name
    transform_finance_excel = finance_excel_service.transform_finance_excel
    finance_excel_service.chat = lambda _prompt: "固定测试建议：请复核金额汇总。"
    looks_like_transform_request = load_request_matcher()

    beautify_request = "将财务报表里面的内容给美化一下，每个字段都是英文单词不好读"
    assert looks_like_transform_request(beautify_request)
    assert not looks_like_transform_request("生成一份财务报表")
    assert not looks_like_transform_request("查询并生成本月财务报表")
    assert not looks_like_transform_request("生成财务报表并美化一下")

    assert readable_field_name("posting_date") == "记账日期"
    assert readable_field_name("gross_pay") == "应发工资"

    source = BytesIO()
    with pd.ExcelWriter(source, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {
                    "posting_date": "2026-07-01",
                    "gross_pay": 10000,
                    "employee_name": "测试员工",
                }
            ]
        ).to_excel(writer, index=False, sheet_name="Monthly Report")

    result = transform_finance_excel(
        source_filename="finance_monthly_report_202607.xlsx",
        content=source.getvalue(),
        instruction=beautify_request,
    )
    workbook = load_workbook(BytesIO(result.content), data_only=True)

    assert "字段说明" in workbook.sheetnames, workbook.sheetnames
    assert result.metadata["readable_headers_applied"] is True
    assert result.metadata["field_mapping_count"] == 3
    assert "记账日期" in workbook["整理_Monthly Report"].cell(row=1, column=1).value
    assert "应发工资" in workbook["整理_Monthly Report"].cell(row=1, column=2).value

    stream_source = (ROOT / "app/main.py").read_text(encoding="utf-8")
    stream_marker = "if _looks_like_existing_finance_excel_transform_request(request.message):"
    assert stream_source.count(stream_marker) >= 2
    assert "workflow_id=\"finance_excel_transform\"" in stream_source

    print("finance excel beautify spec checks passed")


if __name__ == "__main__":
    main()
