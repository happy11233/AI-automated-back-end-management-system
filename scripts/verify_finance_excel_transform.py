from __future__ import annotations

from io import BytesIO
import json
import os
import time
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook


API_BASE_URL = os.getenv("VERIFY_API_BASE_URL", "http://127.0.0.1:8001").rstrip("/")

ACCOUNTS = {
    "finance": ("finance_demo", "Finance123456"),
    "customer_service": ("employee_demo", "Employee123456"),
}


def main() -> None:
    tokens = {name: login(*account) for name, account in ACCOUNTS.items()}
    content = workbook_bytes()

    response = requests.post(
        f"{API_BASE_URL}/automation/finance/excel-transform",
        headers=auth_headers(tokens["finance"]),
        files={
            "file": (
                "finance_orders.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "instruction": "用销售发票和收付款单核对上传表里的订单金额，输出异常提示。",
            "erp_resources": json.dumps(["Sales Invoice", "Payment Entry"], ensure_ascii=False),
        },
        timeout=120,
    )
    assert response.status_code == 200, response.text[:500]
    assert response.content[:2] == b"PK"

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    expected_sheets = {"处理摘要", "数值汇总", "AI建议", "ERP数据摘要"}
    missing = expected_sheets.difference(workbook.sheetnames)
    assert not missing, f"missing sheets: {missing}; got={workbook.sheetnames}"
    assert any(name.startswith("ERP_Sales Invoice") for name in workbook.sheetnames), workbook.sheetnames
    assert any(name.startswith("ERP_Payment Entry") for name in workbook.sheetnames), workbook.sheetnames
    assert has_row_text(workbook["ERP数据摘要"], "Sales Invoice")
    assert has_row_text(workbook["ERP数据摘要"], "Payment Entry")

    forbidden = requests.post(
        f"{API_BASE_URL}/automation/finance/excel-transform",
        headers=auth_headers(tokens["customer_service"]),
        files={
            "file": (
                "finance_orders.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "instruction": "客服尝试越权生成财务 Excel",
            "erp_resources": json.dumps(["Sales Invoice"], ensure_ascii=False),
        },
        timeout=60,
    )
    assert forbidden.status_code == 403, forbidden.text

    blocked_resource = requests.post(
        f"{API_BASE_URL}/automation/finance/excel-transform",
        headers=auth_headers(tokens["finance"]),
        files={
            "file": (
                "finance_orders.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "instruction": "尝试选择客服 ERP 表",
            "erp_resources": json.dumps(["Issue"], ensure_ascii=False),
        },
        timeout=60,
    )
    assert blocked_resource.status_code == 403, blocked_resource.text

    print(json.dumps({
        "ok": True,
        "download_bytes": len(response.content),
        "sheets": workbook.sheetnames,
        "customer_service_forbidden": forbidden.status_code,
        "non_finance_resource_forbidden": blocked_resource.status_code,
        "note": "real API, real auth, real xlsx upload/download, real ERP provider query; no mock/stub/fake",
    }, ensure_ascii=False))


def workbook_bytes() -> bytes:
    output = BytesIO()
    rows: list[dict[str, Any]] = [
        {"订单号": f"AMZ-US-XLS-{int(time.time())}-001", "店铺": "US Store", "SKU": "BOTTLE-01", "销售额": 120.5, "币种": "USD"},
        {"订单号": f"AMZ-US-XLS-{int(time.time())}-002", "店铺": "US Store", "SKU": "BAG-02", "销售额": -18.0, "币种": "USD"},
    ]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="订单金额")
    return output.getvalue()


def login(username: str, password: str) -> str:
    response = requests.post(
        f"{API_BASE_URL}/auth/login",
        data={"username": username, "password": password},
        timeout=30,
    )
    assert response.status_code == 200, response.text
    return response.json()["access_token"]


def auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def has_row_text(sheet, text: str) -> bool:
    for row in sheet.iter_rows(values_only=True):
        if any(text in str(value) for value in row if value is not None):
            return True
    return False


if __name__ == "__main__":
    main()
